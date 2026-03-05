import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import threading
import queue
import time
import os
import sys
import json
import random
import numpy as np
import openpyxl
from openpyxl.styles import Font


class LinearityTab(ttk.Frame):
    """Linearity試験タブ - DACの直線性を測定し、最小二乗法で誤差を算出する"""

    # DAC仕様定数
    DAC_SPECS = {
        'Position': {'bits': 20, 'span': 20.0, 'ci': 'ci', 'center': '80000', 'dmm_range': '10'},
        'LBC':      {'bits': 16, 'span': 2.0,  'ci': 'cii', 'center': '80000', 'dmm_range': '10'},
    }

    # 出荷試験 (Ship) NG判定基準 (Excelマクロ DacTestBench5K 準拠)
    SHIP_CRITERIA = {
        'Position': {'inl': 0.75, 'dnl': 0.50},
        'LBC':      {'inl': 0.50, 'dnl': 0.25},
    }

    def __init__(self, parent, gpib_3458a, gpib_3499b, datagen_manager, test_tab):
        super().__init__(parent)
        self.gpib_dmm = gpib_3458a
        self.gpib_scanner = gpib_3499b
        self.datagen = datagen_manager
        self.test_tab = test_tab
        self.scanner_slot = "1"

        self.is_running = False
        self._stop_event = threading.Event()
        self._worker_thread = None
        self._update_queue = queue.Queue()
        self._log_window = None

        # 設定変数
        self.pattern_mode = tk.StringVar(value='Ship')
        self.num_points = tk.StringVar(value='64')
        self.pattern_file = tk.StringVar(value='')
        self.dac_var = tk.StringVar(value='Position')
        self.settle_time_var = tk.DoubleVar(value=0.2)
        self.th_gain = tk.DoubleVar(value=0.01)
        self.th_offset = tk.DoubleVar(value=10.0)
        self.th_error = tk.DoubleVar(value=1.5)
        self.save_dir = tk.StringVar(value='linearity_data')

        self._load_settings()
        self._create_widgets()

    # ==================== UI ====================
    def _create_widgets(self):
        main = ttk.Frame(self)
        main.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 左パネル (設定)
        left = ttk.Frame(main, width=340)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        left.pack_propagate(False)
        self._create_left_panel(left)

        # 右パネル (結果)
        right = ttk.Frame(main)
        right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        self._create_right_panel(right)

    def _create_left_panel(self, parent):
        # 色バー付き見出し
        header = tk.Frame(parent, bg='#27ae60', height=28)
        header.pack(fill=tk.X, pady=(0, 8))
        header.pack_propagate(False)
        tk.Label(header, text="Linearity試験", bg='#27ae60', fg='white',
                 font=('', 11, 'bold')).pack(expand=True)

        # --- パターン設定 ---
        pat_frame = ttk.LabelFrame(parent, text="パターン設定", padding=8)
        pat_frame.pack(fill=tk.X, pady=(0, 5))

        mode_frame = ttk.Frame(pat_frame)
        mode_frame.pack(fill=tk.X, pady=2)
        ttk.Label(mode_frame, text="モード:").pack(side=tk.LEFT)
        for value, label in [('Ship', '出荷シーケンス'), ('Random', 'Random'), ('Linear', 'Linear'), ('File', 'File')]:
            ttk.Radiobutton(mode_frame, text=label, variable=self.pattern_mode,
                            value=value, command=self._on_mode_changed).pack(side=tk.LEFT, padx=3)

        pts_frame = ttk.Frame(pat_frame)
        pts_frame.pack(fill=tk.X, pady=2)
        ttk.Label(pts_frame, text="計測点数:").pack(side=tk.LEFT)
        self.pts_combo = ttk.Combobox(pts_frame, textvariable=self.num_points,
                                       values=['32', '64', '128', '256', '512', '1024', '2048', '4096'],
                                       width=8, state='readonly')
        self.pts_combo.pack(side=tk.LEFT, padx=5)

        file_frame = ttk.Frame(pat_frame)
        file_frame.pack(fill=tk.X, pady=2)
        ttk.Label(file_frame, text="ファイル:").pack(side=tk.LEFT)
        self.file_entry = ttk.Entry(file_frame, textvariable=self.pattern_file, width=16)
        self.file_entry.pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)
        self.file_browse_btn = ttk.Button(file_frame, text="参照",
                                           command=self._browse_pattern_file, width=5)
        self.file_browse_btn.pack(side=tk.LEFT)
        self._on_mode_changed()

        # --- DAC設定 ---
        dac_frame = ttk.LabelFrame(parent, text="DAC設定", padding=8)
        dac_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Radiobutton(dac_frame, text="Position (20bit)", variable=self.dac_var, value='Position').pack(anchor=tk.W)
        ttk.Radiobutton(dac_frame, text="LBC (16bit)", variable=self.dac_var, value='LBC').pack(anchor=tk.W)

        settle_frame = ttk.Frame(dac_frame)
        settle_frame.pack(fill=tk.X, pady=2)
        ttk.Label(settle_frame, text="DAC安定待ち:").pack(side=tk.LEFT)
        ttk.Spinbox(settle_frame, from_=0.05, to=5.0, increment=0.05,
                     textvariable=self.settle_time_var, width=6, format="%.2f").pack(side=tk.LEFT, padx=2)
        ttk.Label(settle_frame, text="sec").pack(side=tk.LEFT)

        # --- DEF選択&スキャナーCH (test_tabの変数を直接使用→両画面同期) ---
        def_frame = ttk.LabelFrame(parent, text="DEF選択&スキャナーCH", padding=4)
        def_frame.pack(fill=tk.X, pady=(0, 5))

        channel_options = ['ー'] + [f'CH{i:02d}' for i in range(10)]
        save_cmd = lambda e=None: self.test_tab.save_settings()

        # 2列レイアウト (DEF0-2 | DEF3-5)
        col_container = ttk.Frame(def_frame)
        col_container.pack(fill=tk.X)

        for col_idx, start in enumerate([0, 3]):
            col = ttk.Frame(col_container)
            col.pack(side=tk.LEFT, padx=(0, 8) if col_idx == 0 else (0, 0))
            # ヘッダー
            hdr = ttk.Frame(col)
            hdr.pack(fill=tk.X)
            ttk.Label(hdr, text="", width=5).pack(side=tk.LEFT)
            ttk.Label(hdr, text="Pos", width=5, font=('', 8)).pack(side=tk.LEFT, padx=1)
            ttk.Label(hdr, text="Neg", width=5, font=('', 8)).pack(side=tk.LEFT, padx=1)

            for i in range(start, start + 3):
                row = ttk.Frame(col)
                row.pack(fill=tk.X, pady=1)
                ttk.Checkbutton(row, text=f"DEF{i}",
                                variable=self.test_tab.def_check_vars[i],
                                command=self.test_tab.save_settings,
                                width=5).pack(side=tk.LEFT)
                pos_cb = ttk.Combobox(row, textvariable=self.test_tab.scanner_channels_pos[i],
                                       values=channel_options, width=5, state='readonly')
                pos_cb.pack(side=tk.LEFT, padx=2)
                pos_cb.bind('<<ComboboxSelected>>', save_cmd)
                neg_cb = ttk.Combobox(row, textvariable=self.test_tab.scanner_channels_neg[i],
                                       values=channel_options, width=5, state='readonly')
                neg_cb.pack(side=tk.LEFT, padx=2)
                neg_cb.bind('<<ComboboxSelected>>', save_cmd)

        # --- NG判定閾値 ---
        th_frame = ttk.LabelFrame(parent, text="NG判定閾値", padding=4)
        th_frame.pack(fill=tk.X, pady=(0, 5))
        for label, var, unit in [("Gain:", self.th_gain, "LSB/LSB"),
                                  ("Offset:", self.th_offset, "LSB"),
                                  ("Error:", self.th_error, "LSB")]:
            row = ttk.Frame(th_frame)
            row.pack(fill=tk.X, pady=1)
            ttk.Label(row, text=label, width=7).pack(side=tk.LEFT)
            ttk.Entry(row, textvariable=var, width=8).pack(side=tk.LEFT, padx=2)
            ttk.Label(row, text=unit).pack(side=tk.LEFT)

        # --- 実行制御 ---
        ctrl_frame = ttk.LabelFrame(parent, text="実行制御", padding=4)
        ctrl_frame.pack(fill=tk.X, pady=(0, 5))

        btn_row = ttk.Frame(ctrl_frame)
        btn_row.pack(fill=tk.X, pady=2)
        self.start_btn = ttk.Button(btn_row, text="開始", command=self.start_measurement, width=10)
        self.start_btn.pack(side=tk.LEFT, padx=2)
        self.stop_btn = ttk.Button(btn_row, text="停止", command=self.stop_measurement,
                                    state=tk.DISABLED, width=10)
        self.stop_btn.pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_row, text="実行ログ", command=self._open_log_window, width=10).pack(side=tk.LEFT, padx=2)

        dir_frame = ttk.Frame(ctrl_frame)
        dir_frame.pack(fill=tk.X, pady=2)
        ttk.Label(dir_frame, text="保存先:").pack(side=tk.LEFT)
        ttk.Entry(dir_frame, textvariable=self.save_dir, width=18).pack(
            side=tk.LEFT, padx=2, fill=tk.X, expand=True)
        ttk.Button(dir_frame, text="参照", command=self._browse_save_dir, width=5).pack(side=tk.LEFT)

    def _create_right_panel(self, parent):
        # --- 進捗 ---
        prog_frame = ttk.LabelFrame(parent, text="進捗", padding=8)
        prog_frame.pack(fill=tk.X, pady=(0, 5))

        self.target_label = ttk.Label(prog_frame, text="待機中", font=('', 10, 'bold'))
        self.target_label.pack(anchor=tk.W)

        voltage_row = ttk.Frame(prog_frame)
        voltage_row.pack(fill=tk.X, pady=2)
        ttk.Label(voltage_row, text="計測値:", font=('', 10)).pack(side=tk.LEFT)
        self.voltage_label = ttk.Label(voltage_row, text="--- V",
                                        font=('', 14, 'bold'), foreground='red')
        self.voltage_label.pack(side=tk.LEFT, padx=5)

        prog_row = ttk.Frame(prog_frame)
        prog_row.pack(fill=tk.X, pady=2)
        self.progress_label = ttk.Label(prog_row, text="0 / 0")
        self.progress_label.pack(side=tk.LEFT)
        self.progressbar = ttk.Progressbar(prog_row, length=200, mode='determinate')
        self.progressbar.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)

        # --- 結果サマリー ---
        summary_frame = ttk.LabelFrame(parent, text="結果サマリー", padding=5)
        summary_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        cols = ('def', 'dac', 'pole', 'gain', 'offset', 'maxerr', 'judge')
        self.summary_tree = ttk.Treeview(summary_frame, columns=cols, show='headings', height=10)
        for col, text, w in [('def', 'DEF', 50), ('dac', 'DAC', 70), ('pole', 'Pole', 45),
                              ('gain', 'Gain', 85), ('offset', 'Offset', 70),
                              ('maxerr', 'MaxErr', 70), ('judge', '判定', 45)]:
            self.summary_tree.heading(col, text=text)
            self.summary_tree.column(col, width=w, anchor=tk.CENTER)
        self.summary_tree.pack(fill=tk.BOTH, expand=True)
        self.summary_tree.tag_configure('ng', background='#f5b7b1')
        self.summary_tree.tag_configure('ok', background='#d5f5e3')

    # ==================== 実行ログウィンドウ ====================
    def _open_log_window(self):
        """実行ログを別ウィンドウで表示"""
        if self._log_window and self._log_window.winfo_exists():
            self._log_window.lift()
            self._log_window.focus_force()
            return

        self._log_window = tk.Toplevel(self)
        self._log_window.title("Linearity 実行ログ")

        # メインウィンドウの右隣に配置
        root = self.winfo_toplevel()
        root.update_idletasks()
        rx = root.winfo_x() + root.winfo_width() + 8
        ry = root.winfo_y()
        self._log_window.geometry(f"600x{root.winfo_height()}+{rx}+{ry}")

        toolbar = ttk.Frame(self._log_window)
        toolbar.pack(fill=tk.X, padx=5, pady=2)
        ttk.Button(toolbar, text="クリア", command=self._clear_log, width=8).pack(side=tk.LEFT)

        self.log_text = ScrolledText(self._log_window, wrap=tk.WORD,
                                      state=tk.DISABLED, font=('Courier', 9))
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
        self.log_text.tag_config('INFO', foreground='black')
        self.log_text.tag_config('SUCCESS', foreground='green')
        self.log_text.tag_config('WARNING', foreground='orange')
        self.log_text.tag_config('ERROR', foreground='red')

        # 既存ログがあれば復元
        if hasattr(self, '_log_buffer'):
            self.log_text.config(state=tk.NORMAL)
            for msg, level in self._log_buffer:
                ms = 0
                self.log_text.insert(tk.END, f"{msg}\n", level)
            self.log_text.see(tk.END)
            self.log_text.config(state=tk.DISABLED)

    def _clear_log(self):
        """ログをクリア"""
        if hasattr(self, 'log_text') and self.log_text.winfo_exists():
            self.log_text.config(state=tk.NORMAL)
            self.log_text.delete('1.0', tk.END)
            self.log_text.config(state=tk.DISABLED)
        if hasattr(self, '_log_buffer'):
            self._log_buffer.clear()

    # ==================== 設定 ====================
    def _load_settings(self):
        try:
            if os.path.exists('app_settings.json'):
                with open('app_settings.json', 'r', encoding='utf-8') as f:
                    config = json.load(f)
                lin = config.get('linearity', {})
                self.pattern_mode.set(lin.get('pattern_mode', 'Ship'))
                self.num_points.set(lin.get('num_points', '64'))
                dac_type = lin.get('dac_type')
                if dac_type:
                    self.dac_var.set(dac_type)
                elif lin.get('lbc', False):
                    self.dac_var.set('LBC')
                else:
                    self.dac_var.set('Position')
                self.settle_time_var.set(lin.get('settle_time', 0.2))
                self.th_gain.set(lin.get('th_gain', 0.01))
                self.th_offset.set(lin.get('th_offset', 10.0))
                self.th_error.set(lin.get('th_error', 1.5))
                self.save_dir.set(lin.get('save_dir', 'linearity_data'))
        except Exception:
            pass

    def _save_settings(self):
        try:
            config = {}
            if os.path.exists('app_settings.json'):
                with open('app_settings.json', 'r', encoding='utf-8') as f:
                    config = json.load(f)
            config['linearity'] = {
                'pattern_mode': self.pattern_mode.get(),
                'num_points': self.num_points.get(),
                'dac_type': self.dac_var.get(),
                'settle_time': self.settle_time_var.get(),
                'th_gain': self.th_gain.get(),
                'th_offset': self.th_offset.get(),
                'th_error': self.th_error.get(),
                'save_dir': self.save_dir.get(),
            }
            with open('app_settings.json', 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
        except Exception:
            pass

    # ==================== UIコールバック ====================
    def _on_mode_changed(self):
        mode = self.pattern_mode.get()
        is_file = mode == 'File'
        is_ship = mode == 'Ship'
        self.file_entry.config(state=tk.NORMAL if is_file else tk.DISABLED)
        self.file_browse_btn.config(state=tk.NORMAL if is_file else tk.DISABLED)
        if is_file or is_ship:
            self.pts_combo.config(state=tk.NORMAL)
            self.num_points.set('')
            self.pts_combo.config(state=tk.DISABLED)
        else:
            if not self.num_points.get():
                self.num_points.set('1024')
            self.pts_combo.config(state='readonly')

    def _browse_pattern_file(self):
        path = filedialog.askopenfilename(
            title="パターンファイルを選択",
            filetypes=[("テキストファイル", "*.txt *.csv"), ("すべて", "*.*")])
        if path:
            self.pattern_file.set(path)

    def _browse_save_dir(self):
        d = filedialog.askdirectory(title="保存先フォルダを選択")
        if d:
            self.save_dir.set(d)
            self._save_settings()

    # ==================== パターン生成 ====================

    # 出荷試験用パターン (Ship mode) - POS/NEGで異なるビット境界テストポイント
    # Position POS: 0x00000→0xFFFFF (コード昇順, ビット境界はLSB側)
    SHIP_PATTERN_POSITION_POS = [
        0x00000, 0x00001, 0x00002, 0x00003, 0x00004, 0x00007, 0x00008,
        0x0000F, 0x00010, 0x0001F, 0x00020, 0x0003F, 0x00040, 0x0007F,
        0x00080, 0x000FF, 0x00100, 0x001FF, 0x00200, 0x003FF, 0x00400,
        0x007FF, 0x00800, 0x00FFF, 0x01000, 0x01FFF, 0x02000, 0x03FFF,
        0x04000, 0x07FFF, 0x08000, 0x0FFFF, 0x10000, 0x1FFFF, 0x20000,
        0x3FFFF, 0x40000, 0x5FFFF, 0x60000, 0x7FFFF, 0x80000, 0x9FFFF,
        0xA0000, 0xBFFFF, 0xC0000, 0xDFFFF, 0xE0000, 0xFFFFF,
    ]
    # Position NEG: ビット境界はMSB側 (コード昇順スイープ)
    SHIP_PATTERN_POSITION_NEG = [
        0x00000, 0x1FFFF, 0x20000, 0x3FFFF, 0x40000, 0x5FFFF, 0x60000,
        0x7FFFF, 0x80000, 0x9FFFF, 0xA0000, 0xBFFFF, 0xC0000, 0xDFFFF,
        0xE0000, 0xEFFFF, 0xF0000, 0xF7FFF, 0xF8000, 0xFBFFF, 0xFC000,
        0xFDFFF, 0xFE000, 0xFEFFF, 0xFF000, 0xFF7FF, 0xFF800, 0xFFBFF,
        0xFFC00, 0xFFDFF, 0xFFE00, 0xFFEFF, 0xFFF00, 0xFFF7F, 0xFFF80,
        0xFFFBF, 0xFFFC0, 0xFFFDF, 0xFFFE0, 0xFFFEF, 0xFFFF0, 0xFFFF7,
        0xFFFF8, 0xFFFFB, 0xFFFFC, 0xFFFFD, 0xFFFFE, 0xFFFFF,
    ]
    # LBC POS: ビット境界はMSB側 (コード昇順スイープ)
    # LBC POS/NEG共通パターン (54点, LSB側ビット境界テスト)
    SHIP_PATTERN_LBC = [
        0x0000, 0x0001, 0x0002, 0x0003, 0x0004, 0x0007, 0x0008,
        0x000F, 0x0010, 0x001F, 0x0020, 0x003F, 0x0040, 0x007F,
        0x0080, 0x00FF, 0x0100, 0x01FF, 0x0200, 0x03FF, 0x0400,
        0x07FF, 0x0800, 0x0FFF, 0x1000, 0x1FFF, 0x2000, 0x2FFF,
        0x3000, 0x3FFF, 0x4000, 0x4FFF, 0x5000, 0x5FFF, 0x6000,
        0x6FFF, 0x7000, 0x7FFF, 0x8000, 0x8FFF, 0x9000, 0x9FFF,
        0xA000, 0xAFFF, 0xB000, 0xBFFF, 0xC000, 0xCFFF, 0xD000,
        0xDFFF, 0xE000, 0xEFFF, 0xF000, 0xFFFF,
    ]

    def _generate_pattern(self, bits, pole='POS'):
        mode = self.pattern_mode.get()
        max_val = (1 << bits) - 1

        if mode == 'File':
            values = []
            try:
                with open(self.pattern_file.get(), 'r') as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith('#'):
                            values.append(int(line))
            except Exception as e:
                raise ValueError(f"パターンファイル読込エラー: {e}")
            return values

        if mode == 'Ship':
            if bits == 20:
                if pole == 'POS':
                    return list(self.SHIP_PATTERN_POSITION_POS)
                else:
                    return list(reversed(self.SHIP_PATTERN_POSITION_NEG))
            else:
                if pole == 'POS':
                    max_val = (1 << bits) - 1
                    return [max_val - v for v in self.SHIP_PATTERN_LBC]
                else:
                    return list(self.SHIP_PATTERN_LBC)

        n = int(self.num_points.get())
        if mode == 'Linear':
            if n <= 1:
                return [0]
            step = max_val / (n - 1)
            return [min(int(round(i * step)), max_val) for i in range(n)]
        else:  # Random (VBAマクロ準拠: 未ソート、重複許可)
            return [random.randint(0, max_val) for _ in range(n)]

    # ==================== 計測制御 ====================
    def start_measurement(self):
        # 接続チェック
        if not self.datagen.is_connected():
            messagebox.showerror("エラー", "DataGenが未接続です")
            return
        if not self.gpib_dmm.connected:
            messagebox.showerror("エラー", "DMM (3458A) が未接続です")
            return
        if not self.gpib_scanner.connected:
            messagebox.showerror("エラー", "スキャナー (3499B) が未接続です")
            return
        defs = self._get_selected_defs()
        if not defs:
            messagebox.showwarning("警告", "DEFを選択してください")
            return

        self._save_settings()
        self.is_running = True
        self._stop_event.clear()
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)

        # サマリーをクリア & ヘッダー切替
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)
        if self.pattern_mode.get() == 'Ship':
            for col, text in [('gain', 'INL(max)'), ('offset', 'DNL(max)'), ('maxerr', '')]:
                self.summary_tree.heading(col, text=text)
        else:
            for col, text in [('gain', 'Gain'), ('offset', 'Offset'), ('maxerr', 'MaxErr')]:
                self.summary_tree.heading(col, text=text)

        self.log("=== Linearity試験 開始 ===", "INFO")

        self._worker_thread = threading.Thread(target=self._measurement_worker, daemon=True)
        self._worker_thread.start()
        self._poll_updates()

    def stop_measurement(self):
        self.log("=== 停止要求 ===", "WARNING")
        self._stop_event.set()

    def _finish(self):
        self.is_running = False
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.target_label.config(text="完了")

    # ==================== DEF情報取得 ====================
    def _get_selected_defs(self):
        selected = []
        if hasattr(self.test_tab, 'def_check_vars'):
            for i, var in enumerate(self.test_tab.def_check_vars):
                if var.get():
                    pos_ch = self.test_tab.scanner_channels_pos[i].get() \
                        if i < len(self.test_tab.scanner_channels_pos) else "ー"
                    neg_ch = self.test_tab.scanner_channels_neg[i].get() \
                        if i < len(self.test_tab.scanner_channels_neg) else "ー"
                    if (pos_ch == "ー" or not pos_ch) and (neg_ch == "ー" or not neg_ch):
                        continue
                    selected.append({
                        'index': i, 'name': f"DEF{i}",
                        'pos_channel': pos_ch, 'neg_channel': neg_ch
                    })
        return selected

    def _get_serial_number(self, def_index):
        try:
            with open('app_settings.json', 'r', encoding='utf-8') as f:
                config = json.load(f)
            sn = config.get('comm_profiles', {}).get('1', {}).get('serial_numbers', {})
            return sn.get(f'DEF{def_index}_sn', f'DEF{def_index}')
        except Exception:
            return f'DEF{def_index}'

    # ==================== ワーカースレッド ====================
    def _measurement_worker(self):
        try:
            selected_defs = self._get_selected_defs()
            switch_delay = self._load_switch_delay()
            settle_time = self.settle_time_var.get()

            # スキャナー初期化
            self._queue_update('log', ("スキャナー初期化 (cpon)...", "INFO"))
            self._scanner_cpon()
            time.sleep(0.5)

            dac_types = [self.dac_var.get()]
            is_ship = self.pattern_mode.get() == 'Ship'

            for dac_name in dac_types:
                if self._stop_event.is_set():
                    break

                spec = self.DAC_SPECS[dac_name]
                bits = spec['bits']
                span = spec['span']
                ci_cmd = spec['ci']
                center_hex = spec['center']
                dmm_range = spec['dmm_range']

                self._queue_update('log', (
                    f"--- {dac_name} ({bits}bit) 計測開始 ---", "INFO"))

                # DataGen: A固定モード設定
                self._datagen_send(f"alt s sa {ci_cmd}")
                time.sleep(0.1)

                # DMM Range設定
                self._queue_update('log', (f"DMM Range: DCV {dmm_range}", "INFO"))
                self.gpib_dmm.write(f"DCV {dmm_range}")
                time.sleep(0.3)

                for def_info in selected_defs:
                    if self._stop_event.is_set():
                        break

                    serial_no = self._get_serial_number(def_info['index'])
                    ship_pole_results = {}
                    linear_pole_results = {}
                    common_timestamp = time.strftime('%Y%m%d_%H%M%S')

                    for pole in ['POS', 'NEG']:
                        if self._stop_event.is_set():
                            break

                        ch_key = 'pos_channel' if pole == 'POS' else 'neg_channel'
                        channel = def_info[ch_key]
                        if not channel or channel == 'ー':
                            self._queue_update('log', (
                                f"{def_info['name']} {pole}: CH未設定、スキップ", "WARNING"))
                            continue

                        pole_cmd = 'p' if pole == 'POS' else 'n'

                        # パターン生成 (出荷シーケンスはPOS/NEGで異なるパターン)
                        try:
                            sweep_values = self._generate_pattern(bits, pole)
                        except ValueError as e:
                            self._queue_update('log', (str(e), "ERROR"))
                            continue

                        self._queue_update('log', (
                            f"  {pole} パターン: {len(sweep_values)}点, "
                            f"安定待ち: {settle_time}秒", "INFO"))

                        ch_number = channel.replace("CH", "")
                        channel_addr = f"@{self.scanner_slot}{ch_number}"

                        self._queue_update('log', (
                            f"▼ {def_info['name']} {dac_name} {pole} (CH:{channel_addr})", "INFO"))
                        self._queue_update('target',
                            f"{def_info['name']} {dac_name} {pole}")

                        # スキャナーCH切替
                        self._switch_scanner(channel_addr, switch_delay)

                        # Center設定
                        self._datagen_set_value(center_hex, ci_cmd, pole_cmd)
                        time.sleep(settle_time)

                        # 計測ループ (POS/NEGそれぞれ専用パターンでスイープ)
                        x_vals = []
                        y_vals = []
                        total_pts = len(sweep_values)

                        for idx, val in enumerate(sweep_values):
                            if self._stop_event.is_set():
                                break

                            mask = (1 << bits) - 1
                            shift = 20 - bits  # LBC(16bit)は上詰め(4bit左シフト)
                            hex_str = f"{(val & mask) << shift:05X}"

                            # DAC値設定
                            self._datagen_set_value(hex_str, ci_cmd, pole_cmd)
                            time.sleep(settle_time)

                            # DMM計測
                            voltage = self._measure_voltage()
                            if voltage is not None:
                                x_vals.append(val)
                                y_vals.append(voltage)
                                self._queue_update('voltage', f"{voltage:.6f} V")
                                self._queue_update('log', (
                                    f"  [{idx+1}/{total_pts}] {hex_str} → {voltage:.6f} V",
                                    "INFO"))
                            else:
                                self._queue_update('voltage', "--- V")
                                self._queue_update('log', (
                                    f"  [{idx+1}/{total_pts}] {hex_str} → 計測失敗",
                                    "WARNING"))

                            # 進捗更新
                            self._queue_update('progress', (idx + 1, total_pts))

                        # Center復帰
                        self._datagen_set_value(center_hex, ci_cmd, pole_cmd)

                        if self._stop_event.is_set():
                            break

                        if len(x_vals) < 2:
                            self._queue_update('log', (
                                "計測点不足、解析スキップ", "WARNING"))
                            continue

                        if is_ship:
                            # --- 出荷シーケンスモード: INL/DNL (Excelマクロ準拠) ---
                            # コード昇順にソート (降順パターンの場合に対応)
                            paired = sorted(zip(x_vals, y_vals))
                            x_vals = [p[0] for p in paired]
                            y_vals = [p[1] for p in paired]

                            ship_results = self._calculate_linearity_ship(
                                x_vals, y_vals, bits, dac_name)

                            criteria = self.SHIP_CRITERIA[dac_name]
                            inl_arr = np.array(ship_results['inl'])
                            dnl_vals = [d for d in ship_results['dnl'] if d is not None]
                            dnl_arr = np.array(dnl_vals) if dnl_vals else np.array([0.0])

                            inl_max_p = float(np.max(inl_arr))
                            inl_max_n = float(np.min(inl_arr))
                            dnl_max_p = float(np.max(dnl_arr))
                            dnl_max_n = float(np.min(dnl_arr))

                            inl_ok = max(abs(inl_max_p), abs(inl_max_n)) <= criteria['inl']
                            dnl_ok = max(abs(dnl_max_p), abs(dnl_max_n)) <= criteria['dnl']
                            judge = 'OK' if (inl_ok and dnl_ok) else 'NG'

                            inl_worst = inl_max_n if abs(inl_max_n) >= abs(inl_max_p) else inl_max_p
                            dnl_worst = dnl_max_n if abs(dnl_max_n) >= abs(dnl_max_p) else dnl_max_p

                            self._queue_update('log', (
                                f"  INL: {inl_max_n:+.4f} ~ {inl_max_p:+.4f} LSB "
                                f"(基準: \u00b1{criteria['inl']}), "
                                f"DNL: {dnl_max_n:+.4f} ~ {dnl_max_p:+.4f} LSB "
                                f"(基準: \u00b1{criteria['dnl']}) \u2192 {judge}",
                                "ERROR" if judge == 'NG' else "SUCCESS"
                            ))

                            # XLSX保存 (POS: 最終ファイル名, NEG: 一時ファイル)
                            save_dir = self.save_dir.get()
                            os.makedirs(save_dir, exist_ok=True)
                            base_name = (f"{serial_no}_{dac_name}_linearity_"
                                         f"出荷シーケンス_{common_timestamp}")
                            if pole == 'POS':
                                xlsx_filepath = os.path.join(
                                    save_dir, base_name + '.xlsx')
                            else:
                                xlsx_filepath = os.path.join(
                                    save_dir, base_name + '_tmp.xlsx')

                            xlsx_path, _ = self._save_xlsx_ship(
                                ship_results, x_vals, y_vals,
                                dac_name, pole, def_info, serial_no, bits,
                                filepath=xlsx_filepath)
                            if xlsx_path:
                                self._queue_update('log', (
                                    f"  XLSX保存: {xlsx_path}", "SUCCESS"))

                            pole_data = {
                                'def_name': def_info['name'],
                                'dac_name': dac_name,
                                'pole': pole,
                                'inl_worst': inl_worst,
                                'dnl_worst': dnl_worst,
                                'judge': judge,
                                'ship_results': ship_results,
                                'x_vals': x_vals,
                                'serial_no': serial_no,
                                'bits': bits,
                                'xlsx_path': xlsx_path,
                            }
                            ship_pole_results[pole] = pole_data

                            # サマリー更新 + PNG表示を即時キュー
                            self._queue_update('ship_pole_done', pole_data)

                        else:
                            # --- 通常モード: XLSX保存 + Gain/Offset/Error ---
                            save_dir = self.save_dir.get()
                            os.makedirs(save_dir, exist_ok=True)
                            mode = self.pattern_mode.get().lower()
                            base_name = (f"{serial_no}_{dac_name}_linearity_"
                                         f"{mode}_{common_timestamp}")
                            if pole == 'POS':
                                xlsx_filepath = os.path.join(
                                    save_dir, base_name + '.xlsx')
                            else:
                                xlsx_filepath = os.path.join(
                                    save_dir, base_name + '_tmp.xlsx')

                            xlsx_path, results = self._save_xlsx_linear(
                                x_vals, y_vals, dac_name, pole,
                                def_info, serial_no, bits, span,
                                filepath=xlsx_filepath)

                            if results:
                                ng_flags = results.get('ng_detail', [])
                                detail_str = (f" ({', '.join(ng_flags)})"
                                              if ng_flags else "")
                                self._queue_update('log', (
                                    f"  Gain={results['gain']:.6f}, "
                                    f"Offset={results['offset']:.3f} LSB, "
                                    f"MaxErr={results['max_error']:.3f} LSB "
                                    f"\u2192 {results['judge']}{detail_str}",
                                    "ERROR" if ng_flags else "SUCCESS"
                                ))

                            if xlsx_path:
                                self._queue_update('log', (
                                    f"  XLSX保存: {xlsx_path}", "SUCCESS"))

                            pole_data = {
                                'def_name': def_info['name'],
                                'dac_name': dac_name,
                                'pole': pole,
                                'results': results or {},
                                'serial_no': serial_no,
                                'xlsx_path': xlsx_path,
                            }
                            linear_pole_results[pole] = pole_data

                            # サマリー + PNG表示
                            self._queue_update('linear_pole_done', pole_data)

                    # POS/NEGを1ファイルに統合
                    pole_results = ship_pole_results if is_ship \
                        else linear_pole_results
                    merge_msg = 'merge_ship_xlsx' if is_ship \
                        else 'merge_linear_xlsx'
                    if len(pole_results) == 2:
                        self._queue_update(merge_msg, pole_results)
                    elif len(pole_results) == 1:
                        only_data = list(pole_results.values())[0]
                        xp = only_data.get('xlsx_path', '')
                        if '_tmp.xlsx' in xp:
                            final = xp.replace('_tmp.xlsx', '.xlsx')
                            try:
                                os.rename(xp, final)
                                only_data['xlsx_path'] = final
                            except Exception:
                                pass

            # 終了処理
            self._scanner_cpon()
            self._queue_update('log', ("=== Linearity試験 完了 ===", "INFO"))
            self._queue_update('done', None)

        except Exception as e:
            import traceback
            self._queue_update('log', (
                f"エラー: {e}\n{traceback.format_exc()}", "ERROR"))
            self._queue_update('done', None)

    # ==================== ハードウェア操作 ====================
    def _scanner_cpon(self):
        """スキャナー全チャンネルOPEN (cpon)"""
        orig_timeout = self.gpib_scanner.instrument.timeout
        self.gpib_scanner.instrument.timeout = 5000
        try:
            self.gpib_scanner.write(f":system:cpon {self.scanner_slot}")
        finally:
            self.gpib_scanner.instrument.timeout = orig_timeout

    def _switch_scanner(self, channel_addr, switch_delay):
        """スキャナーCH切替 (cpon方式 - measurement_windowと同じ)"""
        orig_timeout = self.gpib_scanner.instrument.timeout
        self.gpib_scanner.instrument.timeout = 5000
        try:
            # 1. cpon: 全CH OPEN
            self.gpib_scanner.write(f":system:cpon {self.scanner_slot}")
            # 2. リレー安定待ち
            time.sleep(switch_delay / 2)
            # 3. *OPC?: cpon完了確認
            self.gpib_scanner.query("*OPC?")
            # 4. CLOSE: 対象CH選択
            self.gpib_scanner.write(f"CLOSE ({channel_addr})")
            # 5. CLOSE後安定待ち
            time.sleep(switch_delay / 2)
        finally:
            self.gpib_scanner.instrument.timeout = orig_timeout

    def _measure_voltage(self):
        """DMM計測 (TRIG SGL)"""
        orig_timeout = self.gpib_dmm.instrument.timeout
        self.gpib_dmm.instrument.timeout = 5000
        try:
            success, response = self.gpib_dmm.query("TRIG SGL")
            if success:
                return float(response.strip())
        except Exception:
            pass
        finally:
            self.gpib_dmm.instrument.timeout = orig_timeout
        return None

    def _datagen_send(self, cmd):
        """DataGenコマンド送信"""
        self._queue_update('log', (f"  DG SEND: {cmd}", "INFO"))
        self.datagen.send_command(cmd)
        time.sleep(0.05)

    def _datagen_set_value(self, hex_str, ci_cmd, pole_cmd):
        """DAC値設定（Linearモード時はp/n両方に同じ値を設定）"""
        if self.pattern_mode.get() == 'Linear':
            self._datagen_send(f"alt a {hex_str} {ci_cmd} p")
            self._datagen_send(f"alt a {hex_str} {ci_cmd} n")
        elif ci_cmd == 'cii':
            # LBCはNEGでも常にcii pにセット
            self._datagen_send(f"alt a {hex_str} {ci_cmd} p")
        else:
            self._datagen_send(f"alt a {hex_str} {ci_cmd} {pole_cmd}")

    def _load_switch_delay(self):
        """スキャナー切替時間を設定ファイルから読み込み"""
        try:
            with open('app_settings.json', 'r', encoding='utf-8') as f:
                config = json.load(f)
            return max(0.4, config.get('measurement_window', {}).get('switch_delay_sec', 0.4))
        except Exception:
            return 0.4

    # ==================== 最小二乗法・誤差計算 ====================
    def _calculate_linearity(self, x, y, bits, span, is_neg):
        """最小二乗法で直線性を計算

        Args:
            x: DAC設定値 (numpy array)
            y: 測定電圧 (numpy array)
            bits: DACビット数 (20 or 16)
            span: 電圧スパン (V)
            is_neg: NEG極性フラグ

        Returns:
            dict: gain, offset, max_error, errors, etc.
        """
        n = len(x)
        max_val = (1 << bits) - 1
        v_per_lsb = span / max_val

        # 最小二乗法: y = m*x + b
        sx = np.sum(x)
        sy = np.sum(y)
        sxx = np.sum(x * x)
        sxy = np.sum(x * y)
        denom = n * sxx - sx * sx

        m = (n * sxy - sx * sy) / denom
        b = (sy * sxx - sx * sxy) / denom

        # 理論値: コード0 → -span/2, コードmax → +span/2
        # 理論傾き = span / max_val = v_per_lsb
        # 理論切片 = -span / 2
        # POS: コード000000→-10V, コードFFFFF→+10V (傾き正)
        # NEG: コード000000→+10V, コードFFFFF→-10V (傾き負)
        ideal_slope = v_per_lsb if not is_neg else -v_per_lsb
        ideal_intercept = -span / 2 if not is_neg else span / 2

        # Gain (LSB/LSB): 実測傾き / 理論傾き
        gain = m / ideal_slope
        # Offset (LSB): 実測切片と理論切片の差をLSB換算
        offset = (b - ideal_intercept) / v_per_lsb

        # 各点の誤差 (LSB)
        y_fit = m * x + b
        errors = (y - y_fit) / v_per_lsb
        max_error = float(np.max(np.abs(errors)))

        return {
            'gain': float(gain),
            'offset': float(offset),
            'max_error': max_error,
            'm': float(m),
            'b': float(b),
            'v_per_lsb': v_per_lsb,
            'errors': errors.tolist(),
            'y_fit': y_fit.tolist(),
        }

    def _calculate_linearity_ship(self, unsigned_vals, measured_v, bits, dac_name):
        """出荷試験用 INL/DNL計算 (Excelマクロ DacTestBench5K 準拠)

        - 係数: 端点フィット (last_V - first_V) / (last_signed - first_signed)
        - 理論電圧: first_V + coeff * (signed - first_signed)
        - INL: (measured - theoretical) / v_per_lsb
        - DNL: INL[i+1] - INL[i] (コードステップ=1 の場合のみ)
        """
        offset_val = 2 ** (bits - 1)
        signed = np.array([v - offset_val for v in unsigned_vals], dtype=float)

        y = np.array(measured_v, dtype=float)

        # 端点フィット: coeff = (V_last - V_first) / (signed_last - signed_first)
        coeff = float((y[-1] - y[0]) / (signed[-1] - signed[0]))

        # 理論電圧: first_V + coeff * (signed - first_signed)
        theoretical_v = y[0] + coeff * (signed - signed[0])

        # 誤差計算用 V/LSB
        if dac_name == 'Position':
            v_per_lsb = coeff  # Position: 係数 = V/LSB (常に正)
        else:  # LBC: Position 1LSB = 10V / 2^15 = 20V / 2^16
            v_per_lsb = 10.0 / (2 ** (bits - 1))

        # INL
        inl = ((y - theoretical_v) / v_per_lsb).tolist()

        # DNL: INL[i+1] - INL[i] (コードステップ=1 の時のみ)
        dnl = [None] * len(unsigned_vals)
        for i in range(len(unsigned_vals) - 1):
            if unsigned_vals[i + 1] - unsigned_vals[i] == 1:
                dnl[i] = inl[i + 1] - inl[i]

        return {
            'coeff': coeff,
            'signed': signed.tolist(),
            'theoretical_v': theoretical_v.tolist(),
            'inl': inl,
            'dnl': dnl,
            'v_per_lsb': v_per_lsb,
        }

    # ==================== CSV保存 ====================
    def _save_csv(self, results, x_vals, y_vals, dac_name, pole,
                  def_info, serial_no, bits, span):
        """計測結果をCSVファイルに保存"""
        save_dir = self.save_dir.get()
        os.makedirs(save_dir, exist_ok=True)

        timestamp = time.strftime('%Y%m%d_%H%M%S')
        mode = '出荷シーケンス' if self.pattern_mode.get() == 'Ship' else self.pattern_mode.get()
        filename = f"{serial_no}_{dac_name}_{pole}_linearity_{mode}_{timestamp}.csv"
        filepath = os.path.join(save_dir, filename)

        v_per_lsb = results['v_per_lsb']
        m = results['m']
        b = results['b']

        try:
            with open(filepath, 'w', encoding='utf-8', newline='') as f:
                # ヘッダ情報
                f.write(f"# Linearity Test Result\n")
                f.write(f"# Serial: {serial_no}\n")
                f.write(f"# DEF: {def_info['name']}\n")
                f.write(f"# DAC: {dac_name} ({bits}bit)\n")
                f.write(f"# Pole: {pole}\n")
                f.write(f"# Gain: {results['gain']:.8f}\n")
                f.write(f"# Offset: {results['offset']:.6f} LSB\n")
                f.write(f"# Max Error: {results['max_error']:.6f} LSB\n")
                f.write(f"# Judge: {results['judge']}\n")
                f.write(f"# Date: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"# V_per_LSB: {v_per_lsb:.10e}\n")
                f.write(f"# Slope(m): {m:.10e}\n")
                f.write(f"# Intercept(b): {b:.10e}\n")
                f.write(f"#\n")
                f.write("Order,DAC_Set,DAC_HEX,Theoretical_V,Measured_V,"
                        "FIT_V,Error_LSB,NG\n")

                mask = (1 << bits) - 1
                for i, (x, y_meas) in enumerate(zip(x_vals, y_vals)):
                    hex_str = f"{x & mask:05X}"
                    theoretical_v = -span / 2 + x * v_per_lsb
                    fit_v = m * x + b
                    error_lsb = results['errors'][i]
                    ng_mark = "*" if abs(error_lsb) > self.th_error.get() else ""
                    f.write(f"{i+1},{x},{hex_str},{theoretical_v:.6f},"
                            f"{y_meas:.6f},{fit_v:.6f},{error_lsb:.4f},{ng_mark}\n")

            return filepath
        except Exception as e:
            self._queue_update('log', (f"CSV保存エラー: {e}", "ERROR"))
            return None

    def _save_xlsx_linear(self, x_vals, y_vals, dac_name, pole,
                          def_info, serial_no, bits, span,
                          filepath=None):
        """Random/Linear/File計測結果をテンプレートXLSXに保存 (DacTestBench5K準拠)

        テンプレート (linearity_linear.xlsx) をコピーし、データ部分を書き換え。
        グラフシートはテンプレートの書式をそのまま保持。
        シート名を "{serial}{P/N}" に変更し、チャート参照も更新。
        """
        import shutil
        import re
        import zipfile

        template_path = self._get_template_path('linear')
        if not template_path:
            self._queue_update('log', (
                "テンプレートが見つかりません: linearity_linear.xlsx", "ERROR"))
            return None, None

        offset_val = 2 ** (bits - 1)
        max_val = (1 << bits) - 1
        vgain = span / max_val  # V/LSB
        if pole == 'NEG':
            vgain = -vgain  # NEG: 入力コードに対して電圧反転

        # Signed DAC values
        n = len(x_vals)
        signed_vals = [int(x - offset_val) for x in x_vals]

        # 最小二乗法 (signed値ベース): y = m*x + b
        x_arr = np.array(signed_vals, dtype=float)
        y_arr = np.array(y_vals, dtype=float)
        sx = float(np.sum(x_arr))
        sy = float(np.sum(y_arr))
        sxx = float(np.sum(x_arr * x_arr))
        sxy = float(np.sum(x_arr * y_arr))
        denom = n * sxx - sx * sx
        m = (n * sxy - sx * sy) / denom
        b = (sy * sxx - sx * sxy) / denom

        # GAIN (LSB/LSB), OFFSET (LSB) — VBA準拠
        gain_lsb = m / vgain
        offset_lsb = b / vgain

        # Sort by signed DAC value (ascending), preserving original order
        order = list(range(1, n + 1))
        data_sorted = sorted(
            zip(order, signed_vals, y_vals), key=lambda r: r[1])

        # テンプレートコピー → 保存先
        if filepath is None:
            save_dir = self.save_dir.get()
            os.makedirs(save_dir, exist_ok=True)
            timestamp = time.strftime('%Y%m%d_%H%M%S')
            mode = self.pattern_mode.get().lower()
            filename = (f"{serial_no}_{dac_name}_linearity_"
                        f"{mode}_{timestamp}.xlsx")
            filepath = os.path.join(save_dir, filename)
        shutil.copy2(template_path, filepath)

        # データシート書き換え
        mode = self.pattern_mode.get().lower()
        sheet_name = f"{serial_no}{pole[0]}"[:31]
        wb = openpyxl.load_workbook(filepath)
        ws = wb['計算データ']
        old_sheet_name = ws.title
        ws.title = sheet_name

        # チャートシートのタブ名も変更
        chart_tab_name = f"{serial_no}{pole[0]}_chart"[:31]
        for cs in wb.chartsheets:
            cs.title = chart_tab_name

        # 旧データクリア (Row 7+)
        for r in range(7, ws.max_row + 1):
            for c in range(1, 9):
                ws.cell(row=r, column=c).value = None

        # ヘッダー
        title = f"{serial_no} {pole} {dac_name} ({n}点{mode})"
        ws['A1'] = title
        ws['B2'] = mode
        ws['B3'] = gain_lsb
        ws['B4'] = offset_lsb
        ws['C3'] = None
        ws['C4'] = None

        # GAIN NG判定
        red_font = Font(color="FF0000")
        if abs(gain_lsb - 1.0) > self.th_gain.get():
            ws['C3'] = "NG"
            for cell in ['A3', 'B3', 'C3']:
                ws[cell].font = red_font

        # OFFSET NG判定
        if abs(offset_lsb) > self.th_offset.get():
            ws['C4'] = "NG"
            for cell in ['A4', 'B4', 'C4']:
                ws[cell].font = red_font

        # データ行 (Row 7+)
        err_threshold = self.th_error.get()
        max_err = 0.0
        for i, (orig_order, s_val, m_val) in enumerate(data_sorted):
            r = 7 + i
            theo_v = s_val * vgain
            meas_dac = m_val / vgain
            fit_val = (m * s_val + b) / vgain
            err_lsb = m_val / vgain - fit_val
            max_err = max(max_err, abs(err_lsb))

            ws.cell(row=r, column=1, value=orig_order)
            ws.cell(row=r, column=2, value=s_val)
            ws.cell(row=r, column=3, value=theo_v)
            ws.cell(row=r, column=4, value=m_val)
            ws.cell(row=r, column=5, value=meas_dac)
            ws.cell(row=r, column=6, value=fit_val)
            ws.cell(row=r, column=7, value=err_lsb)

            if abs(err_lsb) > err_threshold:
                ws.cell(row=r, column=8, value="NG")
                for c in range(1, 9):
                    ws.cell(row=r, column=c).font = red_font

        # NG判定 (サマリー用)
        ng_flags = []
        if abs(gain_lsb - 1.0) > self.th_gain.get():
            ng_flags.append('Gain')
        if abs(offset_lsb) > self.th_offset.get():
            ng_flags.append('Offset')
        if max_err > err_threshold:
            ng_flags.append('Error')
        calc_results = {
            'gain': gain_lsb,
            'offset': offset_lsb,
            'max_error': max_err,
            'ng': bool(ng_flags),
            'judge': 'NG' if ng_flags else 'OK',
            'ng_detail': ng_flags,
        }

        try:
            wb.save(filepath)
        except Exception as e:
            self._queue_update('log', (f"XLSX保存エラー: {e}", "ERROR"))
            return None, calc_results

        # チャートXML更新: シート名参照 + データ範囲 + タイトル + キャッシュ削除
        last_row = 6 + n
        try:
            tmp = filepath + '.tmp'
            with zipfile.ZipFile(filepath, 'r') as zin:
                with zipfile.ZipFile(tmp, 'w') as zout:
                    for item in zin.infolist():
                        raw = zin.read(item.filename)
                        if item.filename == 'xl/charts/chart1.xml':
                            content = raw.decode('utf-8')
                            # シート名参照を更新
                            content = content.replace(
                                old_sheet_name, sheet_name)
                            # セル参照の行範囲を更新
                            content = re.sub(
                                r'([$][A-G][$]7:[$][A-G][$])\d+',
                                rf'\g<1>{last_row}',
                                content)
                            # チャートタイトルを更新
                            content = re.sub(
                                r'(<a:t>)[^<]*(</a:t>)',
                                rf'\g<1>{title}\g<2>',
                                content, count=1)
                            # numCacheを削除 (Excel再計算)
                            content = re.sub(
                                r'<numCache>.*?</numCache>',
                                '', content, flags=re.DOTALL)
                            raw = content.encode('utf-8')
                        zout.writestr(item, raw)
            shutil.move(tmp, filepath)
        except Exception as e:
            self._queue_update('log', (
                f"  チャート範囲更新エラー: {e}", "WARNING"))

        return filepath, calc_results

    def _save_xlsx_ship(self, ship_results, unsigned_vals, measured_v,
                        dac_name, pole, def_info, serial_no, bits,
                        filepath=None):
        """出荷試験結果をテンプレートXLSXに書き込み保存"""
        template_path = self._get_template_path(dac_name)
        if not template_path:
            self._queue_update('log', (
                f"テンプレートが見つかりません: linearity_{dac_name.lower()}.xlsx",
                "ERROR"))
            return None, False

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # シート名変更 + チャート参照更新
        old_name = ws.title
        new_name = f"{serial_no}{pole[0]}"[:31]
        ws.title = new_name
        self._update_chart_refs(ws, old_name, new_name)

        # A-D列はテンプレートのまま、E列(測定電圧)のみ書き込み
        # POS出力: コード昇順で電圧上昇(-V→+V)
        # NEG出力: コード昇順で電圧下降(+V→-V) ※cii入力の補数関係
        # Position: -V先頭 → NEGのみ逆順、LBC: +V先頭 → POSのみ逆順
        n_pts = len(measured_v)
        if (bits == 20 and pole == 'NEG') or (bits < 20 and pole == 'POS'):
            measured_v = list(reversed(measured_v))

        for i in range(n_pts):
            ws.cell(row=7 + i, column=5, value=measured_v[i])

        # チャートタイトル変更
        chart = ws._charts[0]
        runs = chart.title.tx.rich.paragraphs[0].r
        runs[0].t = f"{serial_no} {pole} {dac_name} "
        runs[2].t = f"({n_pts}"

        # ファイル保存
        if filepath is None:
            save_dir = self.save_dir.get()
            os.makedirs(save_dir, exist_ok=True)
            timestamp = time.strftime('%Y%m%d_%H%M%S')
            filename = (f"{serial_no}_{dac_name}_{pole}_linearity_"
                        f"出荷シーケンス_{timestamp}.xlsx")
            filepath = os.path.join(save_dir, filename)

        try:
            wb.save(filepath)
            return filepath, True
        except Exception as e:
            self._queue_update('log', (f"XLSX保存エラー: {e}", "ERROR"))
            return None, False

    @staticmethod
    def _update_chart_refs(ws, old_name, new_name):
        """チャート内のシート参照を更新"""
        old_unquoted = f"{old_name}!"
        old_quoted = f"'{old_name}'!"
        new_ref = f"'{new_name}'!"

        for chart in ws._charts:
            for child in chart._charts:
                for series in child.series:
                    if (hasattr(series.val, 'numRef') and series.val.numRef
                            and series.val.numRef.f):
                        f = series.val.numRef.f
                        f = f.replace(old_quoted, new_ref)
                        f = f.replace(old_unquoted, new_ref)
                        series.val.numRef.f = f
                    if (hasattr(series, 'title')
                            and hasattr(series.title, 'strRef')
                            and series.title.strRef
                            and series.title.strRef.f):
                        f = series.title.strRef.f
                        f = f.replace(old_quoted, new_ref)
                        f = f.replace(old_unquoted, new_ref)
                        series.title.strRef.f = f

    def _get_template_path(self, dac_name):
        """テンプレートXLSXのパスを取得 (PyInstaller対応)"""
        filename = f'linearity_{dac_name.lower()}.xlsx'

        candidates = [
            os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         '..', 'template', filename),
        ]

        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
            candidates.insert(0, os.path.join(base_path, 'template', filename))
            exe_dir = os.path.dirname(sys.executable)
            candidates.insert(0, os.path.join(exe_dir, 'template', filename))

        for path in candidates:
            if os.path.exists(path):
                return path

        return None

    # ==================== グラフ表示 ====================
    def _export_png_async(self, export_func, data):
        """PNGエクスポートをバックグラウンドスレッドで実行"""
        def _worker():
            import pythoncom
            pythoncom.CoInitialize()
            try:
                png_path = export_func(data)
                if png_path:
                    self.after(0, lambda: self.log(
                        f"  PNG保存: {png_path}", "SUCCESS"))
                    self.after(0, lambda: self._show_png(png_path))
            finally:
                pythoncom.CoUninitialize()
        threading.Thread(target=_worker, daemon=True).start()

    def _show_png(self, png_path):
        """PNGファイルをウィンドウで表示"""
        try:
            from PIL import Image, ImageTk
            img = Image.open(png_path)

            win = tk.Toplevel(self)
            win.title(os.path.basename(png_path))
            win.resizable(True, True)

            photo = ImageTk.PhotoImage(img)
            label = tk.Label(win, image=photo)
            label.image = photo  # 参照保持
            label.pack()

            win.update_idletasks()
        except Exception as e:
            self.log(f"  PNG表示エラー: {e}", "WARNING")

    def _save_graph_ship_png(self, data):
        """XLSX内のグラフと表をExcel COM経由でPNGエクスポート"""
        xlsx_path = data.get('xlsx_path')
        if not xlsx_path or not os.path.exists(xlsx_path):
            return None

        pole = data.get('pole', '')
        base = os.path.splitext(xlsx_path)[0]
        if base.endswith('_tmp'):
            base = base[:-4]
        chart_png = f"{base}_{pole}_chart.png"
        table_png = f"{base}_{pole}_table.png"

        try:
            import win32com.client
            from PIL import ImageGrab

            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            try:
                wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
                ws = wb.Sheets(1)

                # グラフPNG
                ws.ChartObjects(1).Chart.Export(os.path.abspath(chart_png))

                # 表PNG (Row3～判定結果行)
                n_pts = len(data['x_vals'])
                judge_row = 7 + n_pts - 1 + 4
                rng = ws.Range(f"A5:H{judge_row}")
                rng.CopyPicture(Appearance=1, Format=2)  # xlScreen, xlBitmap
                img = ImageGrab.grabclipboard()
                if img:
                    img.save(os.path.abspath(table_png))

                wb.Close(False)
            finally:
                excel.Quit()

            if os.path.exists(table_png):
                self._queue_update('log', (
                    f"  表PNG保存: {table_png}", "SUCCESS"))

            return chart_png
        except Exception as e:
            self._queue_update('log', (
                f"  PNGエクスポートエラー: {e}", "WARNING"))
            return None

    def _save_graph_linear_png(self, data):
        """Linear/Random/File XLSX内のチャートシートをExcel COM経由でPNGエクスポート"""
        xlsx_path = data.get('xlsx_path')
        if not xlsx_path or not os.path.exists(xlsx_path):
            return None

        pole = data.get('pole', '')
        base = os.path.splitext(xlsx_path)[0]
        if base.endswith('_tmp'):
            base = base[:-4]
        chart_png = f"{base}_{pole}_chart.png"
        png_abs = os.path.normpath(os.path.abspath(chart_png))

        # 前回の0バイトファイルが残っている可能性
        if os.path.exists(png_abs):
            os.remove(png_abs)

        try:
            import win32com.client
            from PIL import ImageGrab

            excel = win32com.client.DispatchEx("Excel.Application")
            excel.DisplayAlerts = False
            try:
                xlsx_abs = os.path.normpath(os.path.abspath(xlsx_path))
                wb = excel.Workbooks.Open(xlsx_abs)
                chart = wb.Charts(1)

                # NG時はチャート背景を薄い黄色に
                if data.get('results', {}).get('ng', False):
                    chart.PlotArea.Interior.Color = 13434879  # RGB(255,255,204)
                    wb.Save()

                # チャートシートはレンダリングが必要なため一時的に表示
                excel.Visible = True
                chart.Activate()
                time.sleep(0.5)

                # 方法1: Export
                chart.Export(png_abs, "PNG")

                # 0バイトなら方法2: CopyPicture + クリップボード
                if (not os.path.exists(png_abs)
                        or os.path.getsize(png_abs) == 0):
                    if os.path.exists(png_abs):
                        os.remove(png_abs)
                    chart.CopyPicture(Appearance=1, Format=2)
                    time.sleep(0.5)
                    img = ImageGrab.grabclipboard()
                    if img:
                        img.save(png_abs)

                wb.Close(False)
            finally:
                excel.Visible = False
                excel.Quit()

            if os.path.exists(png_abs) and os.path.getsize(png_abs) > 0:
                return chart_png
            self.log("  PNG: Export/CopyPicture両方失敗", "WARNING")
            return None
        except Exception as e:
            self.log(f"  PNGエクスポートエラー: {e}", "WARNING")
            return None

    def _merge_linear_xlsx(self, pole_results):
        """Linear/Random/File: POS/NEGのXLSXを1ファイルに統合
        チャートシート + データシートをまとめてコピー"""
        pos_xlsx = pole_results.get('POS', {}).get('xlsx_path')
        neg_xlsx = pole_results.get('NEG', {}).get('xlsx_path')

        if not pos_xlsx or not neg_xlsx:
            return

        try:
            import win32com.client

            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            try:
                wb_pos = excel.Workbooks.Open(os.path.abspath(pos_xlsx))
                wb_neg = excel.Workbooks.Open(os.path.abspath(neg_xlsx))

                # NEGの全シート（データシート + チャートシート）をPOSファイルにコピー
                for i in range(1, wb_neg.Sheets.Count + 1):
                    wb_neg.Sheets(i).Copy(
                        None, wb_pos.Sheets(wb_pos.Sheets.Count))

                wb_pos.Save()
                wb_neg.Close(False)
                wb_pos.Close(False)
            finally:
                excel.Quit()

            # NEG一時ファイル削除
            try:
                if os.path.exists(neg_xlsx):
                    os.remove(neg_xlsx)
            except Exception:
                pass

            self.log(f"  XLSX統合完了: {pos_xlsx}", "SUCCESS")

        except Exception as e:
            self.log(f"  XLSX統合エラー: {e}", "WARNING")

    def _merge_ship_xlsx(self, pole_results):
        """POS/NEGのXLSXを1ファイルに統合 (PNGは各pole完了時に既にエクスポート済み)"""
        pos_xlsx = pole_results.get('POS', {}).get('xlsx_path')
        neg_xlsx = pole_results.get('NEG', {}).get('xlsx_path')

        if not pos_xlsx or not neg_xlsx:
            return

        try:
            import win32com.client

            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            try:
                wb_pos = excel.Workbooks.Open(os.path.abspath(pos_xlsx))
                wb_neg = excel.Workbooks.Open(os.path.abspath(neg_xlsx))

                # NEGシートをPOSファイルにコピー
                wb_neg.Sheets(1).Copy(None, wb_pos.Sheets(wb_pos.Sheets.Count))

                wb_pos.Save()
                wb_neg.Close(False)
                wb_pos.Close(False)
            finally:
                excel.Quit()

            # NEG一時ファイル削除
            try:
                if os.path.exists(neg_xlsx):
                    os.remove(neg_xlsx)
            except Exception:
                pass

            self.log(f"  XLSX統合完了: {pos_xlsx}", "SUCCESS")

        except Exception as e:
            self.log(f"  XLSX統合エラー: {e}", "WARNING")

    # ==================== 更新キュー・ポーリング ====================
    def _queue_update(self, msg_type, data):
        """ワーカースレッドからUI更新をキューに追加"""
        self._update_queue.put((msg_type, data))

    def _poll_updates(self):
        """UIスレッドで更新キューをポーリング"""
        try:
            while not self._update_queue.empty():
                msg_type, data = self._update_queue.get_nowait()

                if msg_type == 'log':
                    self.log(data[0], data[1])
                elif msg_type == 'target':
                    self.target_label.config(text=data)
                elif msg_type == 'voltage':
                    self.voltage_label.config(text=data)
                elif msg_type == 'progress':
                    current, total = data
                    self.progress_label.config(text=f"{current} / {total}")
                    self.progressbar['maximum'] = total
                    self.progressbar['value'] = current
                elif msg_type == 'linear_pole_done':
                    res = data['results']
                    tag = 'ng' if res['ng'] else 'ok'
                    self.summary_tree.insert('', 'end', values=(
                        data['def_name'], data['dac_name'], data['pole'],
                        f"{res['gain']:.6f}", f"{res['offset']:.3f}",
                        f"{res['max_error']:.3f}", res['judge']
                    ), tags=(tag,))
                    self._export_png_async(
                        self._save_graph_linear_png, data)
                elif msg_type == 'merge_linear_xlsx':
                    self._merge_linear_xlsx(data)
                elif msg_type == 'ship_pole_done':
                    tag = 'ng' if data['judge'] == 'NG' else 'ok'
                    self.summary_tree.insert('', 'end', values=(
                        data['def_name'], data['dac_name'], data['pole'],
                        f"{data['inl_worst']:.4f}", f"{data['dnl_worst']:.4f}",
                        '', data['judge']
                    ), tags=(tag,))
                    self._export_png_async(
                        self._save_graph_ship_png, data)
                elif msg_type == 'merge_ship_xlsx':
                    self._merge_ship_xlsx(data)
                elif msg_type == 'done':
                    self._finish()
                    return
        except Exception:
            pass

        if self.is_running:
            self.after(50, self._poll_updates)

    # ==================== ログ ====================
    def log(self, message, level="INFO"):
        """ログ出力（バッファに蓄積 + ウィンドウが開いていれば表示）"""
        ms = int(time.time() * 1000) % 1000
        timestamp = time.strftime("%H:%M:%S") + f".{ms:03d}"
        formatted = f"[{timestamp}] {message}"

        # バッファに蓄積
        if not hasattr(self, '_log_buffer'):
            self._log_buffer = []
        self._log_buffer.append((formatted, level))

        # ログウィンドウが開いていれば書き込み
        if self._log_window and self._log_window.winfo_exists():
            try:
                self.log_text.config(state=tk.NORMAL)
                self.log_text.insert(tk.END, formatted + "\n", level)
                self.log_text.see(tk.END)
                self.log_text.config(state=tk.DISABLED)
            except tk.TclError:
                pass
