import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import queue
import time
import os
import json
import copy
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from dc_char_definitions import (
    POSITION_TEST_POINTS, POSITION_DISPLAY_ORDER,
    LBC_TEST_POINTS,
    MONI_TEST_POINTS, MONI_DISPLAY_ORDER,
)


class DCCharTab(ttk.Frame):
    """DC特性タブ: POSTION/LBC/moni の自動計測・XLSX保存・PNG出力"""

    SETTINGS_KEY = "dc_char"

    def __init__(self, parent, gpib_3458a, gpib_3499b, datagen_manager, serial_manager, test_tab):
        super().__init__(parent)
        self.gpib_dmm = gpib_3458a
        self.gpib_scanner = gpib_3499b
        self.datagen = datagen_manager
        self.serial_mgr = serial_manager  # DEFシリアル通信（CAL用）
        self.test_tab = test_tab
        self.scanner_slot = "1"

        # Threading
        self.is_running = False
        self._stop_event = threading.Event()
        self._worker_thread = None
        self._update_queue = queue.Queue()

        # Settings
        self.save_dir = tk.StringVar(value='dc_char_data')
        self.settle_time_var = tk.DoubleVar(value=0.3)
        self.switch_delay_sec = tk.DoubleVar(value=1.0)  # Pattern Testと共有
        self.test_type = tk.StringVar(value='Position')  # Position / LBC / moni

        # Results
        self._results = {}  # {def_index: {'position': [...], 'lbc': [...], 'moni': [...]}}

        self._load_settings()
        self._create_widgets()

        # Auto-save settings on change
        for var in [self.save_dir, self.settle_time_var]:
            var.trace_add("write", lambda *_: self._save_settings())

    # ==================== Settings ====================
    def _load_settings(self):
        try:
            with open('app_settings.json', 'r', encoding='utf-8') as f:
                config = json.load(f)
            s = config.get(self.SETTINGS_KEY, {})
            if 'save_dir' in s:
                self.save_dir.set(s['save_dir'])
            if 'settle_time' in s:
                self.settle_time_var.set(s['settle_time'])
            # スキャナ切替時間はPattern Testと共有（measurement_window.switch_delay_sec）
            sw = config.get("measurement_window", {}).get("switch_delay_sec", 1.0)
            self.switch_delay_sec.set(max(0.4, sw))
        except Exception:
            pass

    def _save_settings(self):
        try:
            try:
                with open('app_settings.json', 'r', encoding='utf-8') as f:
                    config = json.load(f)
            except Exception:
                config = {}
            config[self.SETTINGS_KEY] = {
                'save_dir': self.save_dir.get(),
                'settle_time': self.settle_time_var.get(),
            }
            with open('app_settings.json', 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
        except Exception:
            pass

    # ==================== UI ====================
    def _create_widgets(self):
        # 左右分割
        paned = ttk.PanedWindow(self, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=5, pady=5)

        left = ttk.Frame(paned, width=340)
        right = ttk.Frame(paned)
        paned.add(left, weight=0)
        paned.add(right, weight=1)

        # === 左パネル ===
        # 試験種別選択（排他）
        type_frame = ttk.LabelFrame(left, text="試験種別", padding=5)
        type_frame.pack(fill="x", padx=5, pady=(5, 2))

        type_row = ttk.Frame(type_frame)
        type_row.pack(fill="x")
        for t in ["Position", "LBC", "moni"]:
            ttk.Radiobutton(type_row, text=t, variable=self.test_type,
                            value=t).pack(side="left", padx=(0, 15))

        # 試験設定
        settings_frame = ttk.LabelFrame(left, text="試験設定", padding=5)
        settings_frame.pack(fill="x", padx=5, pady=2)

        row = ttk.Frame(settings_frame)
        row.pack(fill="x", pady=2)
        ttk.Label(row, text="DAC安定待ち:").pack(side="left")
        ttk.Entry(row, textvariable=self.settle_time_var, width=6).pack(side="left", padx=(5, 2))
        ttk.Label(row, text="sec").pack(side="left")

        row2 = ttk.Frame(settings_frame)
        row2.pack(fill="x", pady=2)
        ttk.Label(row2, text="スキャナ切替:").pack(side="left")
        ttk.Label(row2, textvariable=self.switch_delay_sec,
                  font=("Arial", 9, "bold")).pack(side="left", padx=(5, 2))
        ttk.Label(row2, text="sec (Pattern Test共有)").pack(side="left")

        # DEF選択 & スキャナCH (test_tabの変数を共有)
        def_frame = ttk.LabelFrame(left, text="DEF選択 & スキャナCH", padding=5)
        def_frame.pack(fill="x", padx=5, pady=2)

        if hasattr(self.test_tab, 'def_check_vars'):
            ch_values = ["ー"] + [f"CH{i:02d}" for i in range(1, 21)]
            for i in range(len(self.test_tab.def_check_vars)):
                row = ttk.Frame(def_frame)
                row.pack(fill="x", pady=1)
                ttk.Checkbutton(row, text=f"DEF{i}",
                                variable=self.test_tab.def_check_vars[i]).pack(side="left")
                ttk.Label(row, text="POS:").pack(side="left", padx=(10, 0))
                ttk.Combobox(row, textvariable=self.test_tab.scanner_channels_pos[i],
                             values=ch_values, width=5, state="readonly").pack(side="left", padx=2)
                ttk.Label(row, text="NEG:").pack(side="left")
                ttk.Combobox(row, textvariable=self.test_tab.scanner_channels_neg[i],
                             values=ch_values, width=5, state="readonly").pack(side="left", padx=2)

        # ログ表示（DEF/DataGen操作ログ）
        from tkinter.scrolledtext import ScrolledText
        log_header = ttk.Frame(left)
        log_header.pack(fill="x", padx=5, pady=(2, 0))
        ttk.Label(log_header, text="操作ログ").pack(side="left")
        ttk.Button(log_header, text="クリア", width=5,
                   command=self._clear_log).pack(side="right")
        log_frame = ttk.Frame(left, padding=3)
        log_frame.pack(fill="both", expand=True, padx=5, pady=(0, 2))
        self.log_area = ScrolledText(log_frame, font=("Consolas", 8), wrap="none",
                                      height=8, width=38)
        self.log_area.pack(fill="both", expand=True)
        self.log_area.config(state="disabled")

        # === 右パネル ===
        # 実行制御（右上に配置、保存先パスを広く表示）
        ctrl_frame = ttk.LabelFrame(right, text="実行制御", padding=5)
        ctrl_frame.pack(fill="x", padx=5, pady=(5, 2))

        btn_row = ttk.Frame(ctrl_frame)
        btn_row.pack(fill="x", pady=2)
        self.btn_start = tk.Button(btn_row, text="開始", command=self._start_measurement,
                                    width=6, bg="#4CAF50", fg="white", font=("Arial", 10, "bold"),
                                    relief="raised", cursor="hand2")
        self.btn_start.pack(side="left", padx=2)
        self.btn_stop = tk.Button(btn_row, text="停止", command=self._stop_measurement,
                                   width=6, font=("Arial", 10, "bold"),
                                   relief="raised", state="disabled")
        self.btn_stop.pack(side="left", padx=2)
        self.var_progress_label = tk.StringVar(value="待機中")
        ttk.Label(btn_row, textvariable=self.var_progress_label,
                  font=("Arial", 9)).pack(side="left", padx=(15, 0))

        save_row = ttk.Frame(ctrl_frame)
        save_row.pack(fill="x", pady=2)
        self.var_save_enabled = tk.BooleanVar(value=False)
        ttk.Checkbutton(save_row, text="Excel&PNG保存",
                        variable=self.var_save_enabled).pack(side="left")
        ttk.Label(save_row, text="保存先:").pack(side="left", padx=(10, 0))
        ttk.Entry(save_row, textvariable=self.save_dir).pack(side="left", fill="x", expand=True, padx=2)
        ttk.Button(save_row, text="参照", width=4,
                   command=self._browse_save_dir).pack(side="left")

        opt_row = ttk.Frame(ctrl_frame)
        opt_row.pack(fill="x", pady=2)
        self.var_cal_enabled = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt_row, text="計測前CAL(LBC)",
                        variable=self.var_cal_enabled).pack(side="left")

        # 結果サマリー（Excelと同じ列順）
        result_frame = ttk.LabelFrame(right, text="結果サマリー", padding=5)
        result_frame.pack(fill="both", expand=True, padx=5, pady=2)

        columns = ('def_name', 'part', 'code', 'voltage', 'expected', 'error', 'error_pct', 'judge')
        self.tree = ttk.Treeview(result_frame, columns=columns, show='headings', height=18)
        self.tree.heading('def_name', text='ﾕﾆｯﾄNo.')
        self.tree.heading('part', text='極')
        self.tree.heading('code', text='入力ｺｰﾄﾞ')
        self.tree.heading('voltage', text='測定電圧(V)')
        self.tree.heading('expected', text='許容誤差(V)')
        self.tree.heading('error', text='誤差(V)')
        self.tree.heading('error_pct', text='誤差(%)')
        self.tree.heading('judge', text='判定')
        self.tree.column('def_name', width=50, anchor='center')
        self.tree.column('part', width=35, anchor='center')
        self.tree.column('code', width=60, anchor='e')
        self.tree.column('voltage', width=75, anchor='e')
        self.tree.column('expected', width=100, anchor='e')
        self.tree.column('error', width=65, anchor='e')
        self.tree.column('error_pct', width=55, anchor='e')
        self.tree.column('judge', width=35, anchor='center')
        self.tree.tag_configure('ok', background='#d5f5e3')
        self.tree.tag_configure('ng', background='#f5b7b1')

        scrollbar = ttk.Scrollbar(result_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

    def _browse_save_dir(self):
        d = filedialog.askdirectory(title="保存先ディレクトリ選択")
        if d:
            self.save_dir.set(d)

    def _clear_log(self):
        self.log_area.config(state="normal")
        self.log_area.delete("1.0", "end")
        self.log_area.config(state="disabled")

    def _append_log(self, text):
        """操作ログにテキスト追加"""
        self.log_area.config(state="normal")
        self.log_area.insert("end", text.rstrip() + "\n")
        self.log_area.see("end")
        self.log_area.config(state="disabled")

    # ==================== 計測制御 ====================
    def _start_measurement(self):
        selected = self._get_selected_defs()
        if not selected:
            messagebox.showerror("エラー", "DEFを1つ以上選択してください")
            return

        # 接続確認
        if not self.gpib_dmm or not self.gpib_dmm.connected:
            messagebox.showerror("エラー", "DMM (3458A) が未接続です")
            return
        if not self.gpib_scanner or not self.gpib_scanner.connected:
            messagebox.showerror("エラー", "スキャナー (3499B) が未接続です")
            return
        if not self.datagen or not self.datagen.is_connected():
            messagebox.showerror("エラー", "DataGenが未接続です")
            return

        # 初期化
        self.is_running = True
        self._stop_event.clear()
        self._results = {}
        self.tree.delete(*self.tree.get_children())
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")

        self._worker_thread = threading.Thread(
            target=self._measurement_worker, args=(selected,), daemon=True
        )
        self._worker_thread.start()
        self._poll_updates()

    def _stop_measurement(self):
        self._stop_event.set()

    def _poll_updates(self):
        try:
            for _ in range(50):
                msg_type, data = self._update_queue.get_nowait()
                if msg_type == 'progress':
                    self.var_progress_label.set(data)
                elif msg_type == 'log':
                    self._append_log(data)
                elif msg_type == 'row':
                    tag = 'ok' if data[-1] == 'OK' else 'ng'
                    self.tree.insert('', 'end', values=data, tags=(tag,))
                    self.tree.see(self.tree.get_children()[-1])
                elif msg_type == 'done':
                    self.var_progress_label.set("完了")
                    self.is_running = False
                    self.btn_start.config(state="normal")
                    self.btn_stop.config(state="disabled")
                    if data:
                        messagebox.showinfo("完了", data)
                    return
                elif msg_type == 'error':
                    self.var_progress_label.set("エラー")
                    self.is_running = False
                    self.btn_start.config(state="normal")
                    self.btn_stop.config(state="disabled")
                    messagebox.showerror("エラー", str(data))
                    return
        except queue.Empty:
            pass
        if self.is_running:
            self.after(100, self._poll_updates)

    # ==================== DEF選択 ====================
    def _get_selected_defs(self):
        selected = []
        if hasattr(self.test_tab, 'def_check_vars'):
            for i, var in enumerate(self.test_tab.def_check_vars):
                if var.get():
                    pos_ch = self.test_tab.scanner_channels_pos[i].get() \
                        if i < len(self.test_tab.scanner_channels_pos) else "ー"
                    neg_ch = self.test_tab.scanner_channels_neg[i].get() \
                        if i < len(self.test_tab.scanner_channels_neg) else "ー"
                    if (pos_ch == "ー" or not pos_ch) and (neg_ch == "ー" or not neg_ch):
                        continue
                    selected.append({
                        'index': i, 'name': f"DEF{i}",
                        'pos_channel': pos_ch, 'neg_channel': neg_ch
                    })
        return selected

    def _get_serial_number(self, def_index):
        try:
            with open('app_settings.json', 'r', encoding='utf-8') as f:
                config = json.load(f)
            sn = config.get('comm_profiles', {}).get('1', {}).get('serial_numbers', {})
            return sn.get(f'DEF{def_index}_sn', f'DEF{def_index}')
        except Exception:
            return f'DEF{def_index}'

    # ==================== ハードウェア操作 ====================
    def _scanner_cpon(self):
        """スキャナー全チャンネルOPEN (cpon)"""
        orig_timeout = self.gpib_scanner.instrument.timeout
        self.gpib_scanner.instrument.timeout = 5000
        try:
            self.gpib_scanner.write(f":system:cpon {self.scanner_slot}")
        finally:
            self.gpib_scanner.instrument.timeout = orig_timeout

    def _switch_scanner(self, channel_addr, switch_delay):
        """スキャナCH切替（Pattern Testと同じcpon方式）
        cpon → sleep(delay/2) → *OPC? → CLOSE → sleep(delay/2)
        """
        orig_timeout = self.gpib_scanner.instrument.timeout
        self.gpib_scanner.instrument.timeout = 5000
        try:
            self._update_queue.put(('log', f"[Scanner] cpon → CLOSE ({channel_addr})"))
            self.gpib_scanner.write(f":system:cpon {self.scanner_slot}")
            time.sleep(switch_delay / 2)
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                self.gpib_scanner.query("*OPC?")
            self.gpib_scanner.write(f"CLOSE ({channel_addr})")
            time.sleep(switch_delay / 2)
        finally:
            self.gpib_scanner.instrument.timeout = orig_timeout

    def _measure_voltage(self):
        orig_timeout = self.gpib_dmm.instrument.timeout
        self.gpib_dmm.instrument.timeout = 3000
        try:
            success, response = self.gpib_dmm.query("TRIG SGL")
            if success and response:
                val = float(response.strip())
                self._update_queue.put(('log', f"[DMM] TRIG SGL → {val}"))
                return val
        except Exception:
            pass
        finally:
            self.gpib_dmm.instrument.timeout = orig_timeout
        self._update_queue.put(('log', "[DMM] TRIG SGL → 計測失敗"))
        return None

    def _datagen_send(self, cmd):
        self._update_queue.put(('log', f"[DG] SEND: {cmd}"))
        self.datagen.send_command(cmd)
        time.sleep(0.05)

    def _ch_addr(self, ch_str):
        """CH番号文字列からスキャナアドレスを生成 (例: 'CH01' → '@101')"""
        num = ch_str.replace("CH", "")
        return f"@{self.scanner_slot}{num}"

    # ==================== 計測ワーカー ====================
    def _measurement_worker(self, selected_defs):
        try:
            settle = self.settle_time_var.get()
            switch_delay = self.switch_delay_sec.get()
            test_type = self.test_type.get()

            # DEF remoteモード設定
            if self.serial_mgr and self.serial_mgr.is_connected():
                for def_info in selected_defs:
                    cmd = f"DEF {def_info['index']} remote"
                    self._update_queue.put(('log', f"[DEF] SEND: {cmd}"))
                    response = self.serial_mgr.send_command_with_response(
                        cmd, wait_sec=0.1, read_timeout=0.5
                    )
                    if response:
                        self._update_queue.put(('log', f"[DEF] RECV: {response}"))

            # 初期化
            self._scanner_cpon()
            self._datagen_send("gen stop")
            self._datagen_send("func alt")
            self._datagen_send("alt s sa")

            # DMM設定: NPLC 10 + レンジ設定
            self.gpib_dmm.write("NPLC 10")
            time.sleep(0.1)
            if test_type == "Position":
                self.gpib_dmm.write("DCV 1000")
            else:
                self.gpib_dmm.write("DCV AUTO")
            time.sleep(0.1)

            self._datagen_send("gen start")

            # 結果領域を初期化
            for def_info in selected_defs:
                self._results[def_info['index']] = {'position': [], 'lbc': [], 'moni': []}

            if test_type == "LBC":
                # LBC: ATTが外側ループ（ATT切替+CALは3回のみ）
                self._run_lbc_all(selected_defs, settle, switch_delay)
            else:
                # Position/moni: DEFが外側ループ
                for def_info in selected_defs:
                    if self._stop_event.is_set():
                        break
                    if test_type == "Position":
                        self._run_position_test(def_info, settle, switch_delay)
                    else:
                        self._run_moni_test(def_info, settle, switch_delay)

            # 後片付け: コードをcenterに戻す
            self._scanner_cpon()
            self._datagen_send("alt a 80000 ci p")
            self._datagen_send("alt a 80000 ci n")
            self._datagen_send("alt a 80000 cii p")
            self._datagen_send("alt a 80000 cii n")
            self._datagen_send("alt s sa")

            # LBC計測後: 全DEFをLATT 1/2に戻す
            if test_type == "LBC" and self.serial_mgr and self.serial_mgr.is_connected():
                for def_info in selected_defs:
                    latt_cmd = f"DEF {def_info['index']} LATT 2"
                    self._update_queue.put(('log', f"[DEF] SEND: {latt_cmd}"))
                    response = self.serial_mgr.send_command_with_response(
                        latt_cmd, wait_sec=0.1, read_timeout=0.5
                    )
                    if response:
                        self._update_queue.put(('log', f"[DEF] RECV: {response}"))

            if self._stop_event.is_set():
                self._update_queue.put(('done', "計測を中断しました"))
                return

            # 保存（チェックON時のみ）
            if not self.var_save_enabled.get():
                self._update_queue.put(('done', "計測完了（保存なし）"))
                return

            save_dir = self.save_dir.get()
            os.makedirs(save_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            saved_files = []

            sheet_key = test_type.lower()
            # 選択 DEF を 1 ファイルに集約するため slots_data を先に作る
            # slots_data: [(serial_no, display_data), ...] 最大 4 枠
            slots_data: list = []
            for def_idx in sorted(self._results.keys()):
                results = self._results[def_idx]
                sn = self._get_serial_number(def_idx)
                data = results[sheet_key]
                if not data:
                    continue

                # Excel表示順にリオーダー
                if sheet_key == 'position':
                    display_data = self._reorder_for_display(
                        data, POSITION_DISPLAY_ORDER)
                elif sheet_key == 'moni':
                    display_data = self._reorder_for_display(
                        data, MONI_DISPLAY_ORDER)
                else:
                    display_data = data

                slots_data.append((sn, display_data))

            # テンプレートは 4 枠固定。それ以上は載らないので警告
            if len(slots_data) > 4:
                self._update_queue.put((
                    'log',
                    f"[警告] 選択 DEF が {len(slots_data)} 本。"
                    f"テンプレートは 4 枠まで。5 本目以降は記述されません。"))
                slots_data = slots_data[:4]

            if slots_data:
                # XLSX 1 ファイル
                xlsx_path = self._save_xlsx_single(
                    sheet_key, slots_data, save_dir, timestamp)
                saved_files.append(xlsx_path)

                # PNG 1 ファイル (Excel COM でスクリーンショット)
                png_path = self._generate_table_png(
                    sheet_key, slots_data, save_dir, timestamp)
                saved_files.append(png_path)

            msg = "保存完了:\n" + "\n".join(saved_files)
            self._update_queue.put(('done', msg))

        except Exception as e:
            self._update_queue.put(('error', str(e)))

    def _run_position_test(self, def_info, settle, switch_delay):
        """POSTION計測"""
        def_name = def_info['name']
        def_idx = def_info['index']
        current_pole = None
        for pi, tp in enumerate(POSITION_TEST_POINTS):
            if self._stop_event.is_set():
                break
            self._update_queue.put(('progress', f"{def_name} / POSTION / {tp.part} {tp.display_code}"))

            pole = "p" if tp.part == "POS" else "n"
            self._datagen_send(f"alt a {tp.address_code} ci {pole}")
            time.sleep(settle)

            if tp.part != current_pole:
                ch = def_info['pos_channel'] if tp.part == "POS" else def_info['neg_channel']
                self._switch_scanner(self._ch_addr(ch), switch_delay)
                current_pole = tp.part

            voltage = self._measure_voltage()

            if voltage is not None:
                error = voltage - tp.expected
                error_pct = (error / tp.expected * 100) if tp.expected != 0 else 0
                judge = "OK" if abs(error) <= tp.tolerance else "NG"
            else:
                error, error_pct, judge = None, None, "NG"

            self._results[def_idx]['position'].append({
                'part': tp.part, 'code': tp.display_code,
                'voltage': voltage, 'expected_str': tp.expected_str,
                'error': error, 'error_pct': error_pct, 'judge': judge,
            })

            v_str = f"{voltage:.3f}" if voltage is not None else "---"
            e_str = f"{error:.3f}" if error is not None else "---"
            is_center = '80000' in tp.display_code
            ep_str = "" if is_center else (f"{error_pct:.3f}" if error_pct is not None else "---")
            self._update_queue.put(('row', (
                def_name, tp.part, tp.display_code, v_str,
                tp.expected_str, e_str, ep_str, judge
            )))

    def _cal_all_defs(self, selected_defs):
        """全DEFに一斉CAL送信 → ラウンドロビンポーリングで全完了待ち。
        戻り値: True=全OK, False=いずれかNG/TIMEOUT

        [DBG] cal プレフィックスで実時系列を可視化:
          - 各 cal/cal s コマンド送信直前のタイムスタンプ
          - 装置の生応答 (repr で改行・空白を可視化)
          - 完了判定したトリガ応答と所要秒
          - 全 DEF 完了時の総所要秒
        """
        if not self.serial_mgr or not self.serial_mgr.is_connected():
            self._update_queue.put(('log', "[CAL] シリアル未接続"))
            return False

        def _now():
            ms = int(time.time() * 1000) % 1000
            return time.strftime("%H:%M:%S") + f".{ms:03d}"

        def _dbg(msg):
            self._update_queue.put(('log', f"[DBG] cal {_now()} {msg}"))

        t_start = time.time()
        _dbg(f"BEGIN _cal_all_defs DEFs={[d['index'] for d in selected_defs]}")

        # -1. Talkingモード中止 (reticent-mode へ切替)
        # ※ Talking のままだと CAL 実行中に大量の進捗ログを
        #    シリアルに垂れ流し、他 DEF の通信が壊れる
        _dbg("--- reticent-mode ---")
        for def_info in selected_defs:
            di = def_info['index']
            r_cmd = f"DEF {di} r"
            _dbg(f"SEND DEF{di}: {r_cmd}")
            self._update_queue.put(('log', f"[DEF] SEND: {r_cmd}"))
            r_resp = self.serial_mgr.send_command_with_response(
                r_cmd, wait_sec=0.1, read_timeout=0.5)
            _dbg(f"RECV DEF{di}: {r_resp!r}")
            if r_resp:
                self._update_queue.put(('log', f"[DEF] RECV: {r_resp}"))

        # 0. 開始前の sticky 状態を確認
        # CAL 送信前に cal s が complete を返してきたら、装置側に
        # 前回 CAL の sticky flag が残っている可能性が高い
        # (部分一致判定が誤発火する原因の検証)
        _dbg("--- pre-send sticky check ---")
        for def_info in selected_defs:
            di = def_info['index']
            pre_cmd = f"DEF {di} cal s"
            _dbg(f"PRE-SEND DEF{di}: {pre_cmd}")
            pre_resp = self.serial_mgr.send_command_with_response(
                pre_cmd, wait_sec=0.1, read_timeout=1.0)
            _dbg(f"PRE-RECV DEF{di}: {pre_resp!r}")

        # 1. 全DEFにcal一斉送信
        _dbg("--- send cal commands ---")
        pending = {}  # {def_index: t_send}
        for def_info in selected_defs:
            di = def_info['index']
            cmd = f"DEF {di} cal"
            _dbg(f"SEND DEF{di}: {cmd}")
            self._update_queue.put(('log', f"[DEF] SEND: {cmd}"))
            self.serial_mgr.send_command(cmd)
            time.sleep(0.3)
            pending[di] = time.time()

        # 2. ラウンドロビンポーリング (3秒間隔、最大240秒)
        _dbg("--- poll loop start ---")
        max_wait = 240
        elapsed = 0
        poll_interval = 3
        poll_round = 0
        while elapsed < max_wait and pending:
            if self._stop_event.is_set():
                _dbg("STOP event during poll loop")
                return False
            time.sleep(poll_interval)
            elapsed += poll_interval
            poll_round += 1
            _dbg(f"poll round #{poll_round} elapsed={elapsed}s "
                 f"pending={list(pending.keys())}")

            for di in list(pending.keys()):
                if self._stop_event.is_set():
                    _dbg("STOP event in poll inner loop")
                    return False
                status_cmd = f"DEF {di} cal s"
                _dbg(f"SEND DEF{di}: {status_cmd}")
                response = self.serial_mgr.send_command_with_response(
                    status_cmd, wait_sec=0.1, read_timeout=1.0
                )
                _dbg(f"RECV DEF{di}: {response!r}")
                self._update_queue.put(('log', f"[DEF] SEND: {status_cmd}"))
                if response:
                    self._update_queue.put(('log', f"[DEF] RECV: {response}"))
                    resp_lower = response.lower()
                    if "complete" in resp_lower:
                        dt = time.time() - pending[di]
                        _dbg(f"DEF{di} matched 'complete' in {dt:.2f}s")
                        self._update_queue.put((
                            'log', f"[CAL] DEF{di} → 完了 (OK, {dt:.1f}s)"))
                        del pending[di]
                    elif "failed" in resp_lower or "error" in resp_lower:
                        _dbg(f"DEF{di} matched failed/error")
                        self._update_queue.put((
                            'log', f"[CAL] DEF{di} → 失敗 (NG)"))
                        self._update_queue.put(('progress', f"CAL DEF{di} NG"))
                        return False
                    elif ("in excution" in resp_lower
                          or "executing" in resp_lower):
                        _dbg(f"DEF{di} still executing")
                    else:
                        _dbg(f"DEF{di} unmatched response (continue)")
                else:
                    _dbg(f"DEF{di} empty response (continue)")

            # 進捗表示
            done = len(selected_defs) - len(pending)
            names = [f"DEF{di}" for di in pending]
            self._update_queue.put(('progress',
                f"CAL {done}/{len(selected_defs)}完了 "
                f"待機中:{','.join(names)} ({elapsed}s/{max_wait}s)"))

        if pending:
            for di in pending:
                _dbg(f"DEF{di} TIMEOUT after {max_wait}s")
                self._update_queue.put((
                    'log', f"[CAL] DEF{di} → タイムアウト ({max_wait}s)"))
            self._update_queue.put(('progress', f"CAL タイムアウト"))
            return False

        total = time.time() - t_start
        _dbg(f"END _cal_all_defs OK total={total:.2f}s")
        self._update_queue.put((
            'log', f"[CAL] 全DEF完了 (OK, total {total:.1f}s)"))
        self._update_queue.put(('progress', f"CAL 全DEF完了"))
        return True

    def _run_lbc_all(self, selected_defs, settle, switch_delay):
        """LBC計測: ATTが外側ループ（CALはATT毎に1回 = 計3回）
        ATT1/1 → CAL → DEF0計測 → DEF1計測 → ...
        ATT1/2 → CAL → DEF0計測 → DEF1計測 → ...
        ATT1/4 → CAL → DEF0計測 → DEF1計測 → ...
        """
        att_list = ['1/1', '1/2', '1/4']
        att_map = {'1/1': '1', '1/2': '2', '1/4': '4'}
        # ATTごとのテストポイント（FFFF, 0000の2点）
        att_points = {}
        for tp in LBC_TEST_POINTS:
            att_points.setdefault(tp.att, []).append(tp)

        for att in att_list:
            if self._stop_event.is_set():
                break

            points = att_points.get(att, [])
            latt_val = att_map.get(att, '1')

            # ===== 1. LATT変更 =====
            self._update_queue.put(('log', f""))
            self._update_queue.put(('log', f"===== LBC ATT {att} ====="))
            self._update_queue.put(('progress', f"LBC ATT {att} に切替"))

            def _latt_dbg(msg):
                ms = int(time.time() * 1000) % 1000
                ts = time.strftime("%H:%M:%S") + f".{ms:03d}"
                self._update_queue.put((
                    'log', f"[DBG] latt {ts} {msg}"))

            _latt_dbg(f"BEGIN ATT={att} latt_val={latt_val}")
            # 全DEFにLATT送信
            if self.serial_mgr and self.serial_mgr.is_connected():
                for def_info in selected_defs:
                    di = def_info['index']
                    latt_cmd = f"DEF {di} LATT {latt_val}"
                    _latt_dbg(f"SEND DEF{di}: {latt_cmd}")
                    self._update_queue.put(('log', f"[DEF] SEND: {latt_cmd}"))
                    response = self.serial_mgr.send_command_with_response(
                        latt_cmd, wait_sec=0.1, read_timeout=0.5
                    )
                    _latt_dbg(f"RECV DEF{di}: {response!r}")
                    if response:
                        self._update_queue.put((
                            'log', f"[DEF] RECV: {response}"))
            _latt_dbg("sleep 0.2 before CAL")
            time.sleep(0.2)
            _latt_dbg("END LATT phase, entering CAL phase")

            # ===== 2. CAL（一斉送信→ラウンドロビンポーリング） =====
            cal_failed = False
            if self.var_cal_enabled.get():
                if not self._cal_all_defs(selected_defs):
                    self._update_queue.put(('log', f"[CAL] ATT{att} CAL失敗"))
                    cal_failed = True
                    # CAL NG結果を全DEF・全ポイントに記録
                    self._update_queue.put(('row', ("", f"ATT{att}", "", "", "", "", "", "")))
                    for def_info in selected_defs:
                        for tp in points:
                            self._results[def_info['index']]['lbc'].append({
                                'att': tp.att, 'code': tp.display_code,
                                'voltage_pos': None, 'voltage_neg': None,
                                'expected_str': tp.expected_str,
                                'error_pos': None, 'error_neg': None, 'judge': 'CAL NG',
                            })
                            self._update_queue.put(('row', (
                                def_info['name'], "", tp.display_code, "CAL NG",
                                tp.expected_str, "", "", "NG"
                            )))
                    continue  # 次のATTへ
            else:
                self._update_queue.put(('log', f"[CAL] スキップ"))

            if self._stop_event.is_set():
                break

            _latt_dbg(f"CAL phase done, entering MEASURE phase ATT={att}")

            # ATT区切り行をサマリーに表示
            self._update_queue.put(('row', (
                "", f"ATT{att}", "", "", "", "", "", ""
            )))

            # ===== 3. 各DEFの計測 =====
            for def_info in selected_defs:
                if self._stop_event.is_set():
                    break
                def_name = def_info['name']
                def_idx = def_info['index']

                for tp in points:
                    if self._stop_event.is_set():
                        break
                    self._update_queue.put(('progress',
                        f"{def_name} / LBC ATT{att} {tp.display_code}"))

                    self._datagen_send(f"alt a {tp.address_code} cii p")
                    time.sleep(settle)

                    # POS計測
                    self._switch_scanner(self._ch_addr(def_info['pos_channel']), switch_delay)
                    voltage_pos = self._measure_voltage()

                    # NEG計測
                    self._switch_scanner(self._ch_addr(def_info['neg_channel']), switch_delay)
                    voltage_neg = self._measure_voltage()

                    if voltage_pos is not None and voltage_neg is not None:
                        error_pos = voltage_pos - tp.expected_pos
                        error_neg = voltage_neg - tp.expected_neg
                        judge_pos = "OK" if abs(error_pos) <= tp.tolerance else "NG"
                        judge_neg = "OK" if abs(error_neg) <= tp.tolerance else "NG"
                        judge = "OK" if judge_pos == "OK" and judge_neg == "OK" else "NG"
                    else:
                        error_pos, error_neg = None, None
                        judge_pos, judge_neg, judge = "NG", "NG", "NG"

                    self._results[def_idx]['lbc'].append({
                        'att': tp.att, 'code': tp.display_code,
                        'voltage_pos': voltage_pos, 'voltage_neg': voltage_neg,
                        'expected_str': tp.expected_str,
                        'error_pos': error_pos, 'error_neg': error_neg, 'judge': judge,
                    })

                    # POS行
                    vp = f"{voltage_pos:.3f}" if voltage_pos is not None else "---"
                    ep = f"{error_pos:.3f}" if error_pos is not None else "---"
                    self._update_queue.put(('row', (
                        def_name, "POS", tp.display_code, vp,
                        tp.expected_str, ep, "", judge_pos
                    )))
                    # NEG行
                    vn = f"{voltage_neg:.3f}" if voltage_neg is not None else "---"
                    en = f"{error_neg:.3f}" if error_neg is not None else "---"
                    self._update_queue.put(('row', (
                        "", "NEG", tp.display_code, vn,
                        "", en, "", judge_neg
                    )))

    def _run_moni_test(self, def_info, settle, switch_delay):
        """moni計測"""
        def_name = def_info['name']
        def_idx = def_info['index']
        current_pole = None
        for mi, tp in enumerate(MONI_TEST_POINTS):
            if self._stop_event.is_set():
                break
            self._update_queue.put(('progress', f"{def_name} / moni / {tp.part} {tp.display_code}"))

            pole = "p" if tp.part == "POS" else "n"
            self._datagen_send(f"alt a {tp.address_code} ci {pole}")
            time.sleep(settle)

            if tp.part != current_pole:
                ch = def_info['pos_channel'] if tp.part == "POS" else def_info['neg_channel']
                self._switch_scanner(self._ch_addr(ch), switch_delay)
                current_pole = tp.part

            voltage = self._measure_voltage()

            if voltage is not None:
                error = voltage - tp.expected
                error_pct = (error / tp.expected * 100) if tp.expected != 0 else 0
                judge = "OK" if abs(error) <= tp.tolerance else "NG"
            else:
                error, error_pct, judge = None, None, "NG"

            self._results[def_idx]['moni'].append({
                'part': tp.part, 'code': tp.display_code,
                'voltage': voltage, 'expected_str': tp.expected_str,
                'error': error, 'error_pct': error_pct, 'judge': judge,
            })

            v_str = f"{voltage:.2f}" if voltage is not None else "---"
            e_str = f"{error:.2f}" if error is not None else "---"
            ep_str = f"{error_pct:.2f}" if error_pct is not None else "---"
            self._update_queue.put(('row', (
                def_name, tp.part, tp.display_code, v_str,
                tp.expected_str, e_str, ep_str, judge
            )))

    # ==================== 表示順リオーダー ====================
    def _reorder_for_display(self, data, order):
        """計測順の結果リストをExcel表示順に並べ替え"""
        return [data[i] for i in order]

    # ==================== XLSX保存 ====================
    def _save_xlsx_single(self, sheet_key, slots_data, save_dir, timestamp):
        """XLSX保存 (選択 DEF 全てを 1 ファイルに)

        slots_data: [(serial_no, display_data), ...] 最大 4 枠
        ファイル名は選択 DEF の S/N を全て `_` で連結する。
        """
        sheet_titles = {'position': 'POSTION', 'lbc': 'LBC', 'moni': 'moni'}
        sheet_title = sheet_titles.get(sheet_key, sheet_key)
        sn_joined = self._join_sns(slots_data)
        filepath = os.path.join(
            save_dir, f"{sn_joined}_DC特性_{sheet_title}_{timestamp}.xlsx")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title

        if sheet_key == 'position':
            self._write_position_sheet(ws, slots_data)
        elif sheet_key == 'lbc':
            self._write_lbc_sheet(ws, slots_data)
        else:
            self._write_moni_sheet(ws, slots_data)

        wb.save(filepath)
        return filepath

    @staticmethod
    def _join_sns(slots_data):
        """slots_data から S/N を順に `_` 連結 (空ならフォールバック)"""
        if not slots_data:
            return "unknown"
        return "_".join(sn for sn, _ in slots_data)

    def _write_position_sheet(self, ws, slots_data):
        """Position XLSX書き込み（docx仕様: 4DEF固定枠、セル結合、右寄せ、数式埋込）

        slots_data: [(serial_no, display_data), ...] 最大 4 枠
        """
        thin = Side(style='thin')
        thick = Side(style='medium')
        border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
        border_top_thick = Border(top=thick, left=thin, right=thin, bottom=thin)
        dat_font = Font(name='MS Pゴシック', size=10)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        # ヘッダー（太字なし）
        headers = ['ﾕﾆｯﾄ No.', '極', '入力ｺｰﾄﾞ', '測定電圧(V)', '許容誤差(V)(<0.1%)', '誤差(V)', '誤差(%)', '判定']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = dat_font; cell.alignment = center; cell.border = border_all

        expected_strs = [
            "160.000±0.160", "0.000±0.100", "-160.000±0.160",
            "160.000±0.160", "0.000±0.100", "-160.000±0.160",
        ]
        # 期待値数値（数式用）
        expected_vals = [160.0, 0.0, -160.0, 160.0, 0.0, -160.0]
        code_strs = ["FFFFF H", "80000 H", "00000 H", "00000 H", "80000 H", "FFFFF H"]

        for slot in range(4):
            base = 2 + slot * 6
            if slot < len(slots_data):
                sn_slot, slot_data = slots_data[slot]
                unit_label = f"1PB397MK2\n{sn_slot}"
            else:
                slot_data = None
                unit_label = ""

            for ri in range(6):
                r = base + ri
                is_center = (ri == 1 or ri == 4)
                bdr = border_top_thick if ri == 0 else border_all

                if slot_data and ri < len(slot_data):
                    d = slot_data[ri]
                    v = d['voltage'] if d['voltage'] is not None else ''
                    judge = d['judge']
                else:
                    v, judge = '', ''

                # 入力ｺｰﾄﾞ（右寄せ）
                c3 = ws.cell(row=r, column=3, value=code_strs[ri])
                c3.font = dat_font; c3.alignment = right_align; c3.border = bdr
                # 測定電圧（右寄せ）
                c4 = ws.cell(row=r, column=4, value=v)
                c4.font = dat_font; c4.alignment = right_align; c4.border = bdr
                if v != '': c4.number_format = '0.000'
                # 許容誤差（右寄せ）
                c5 = ws.cell(row=r, column=5, value=expected_strs[ri])
                c5.font = dat_font; c5.alignment = right_align; c5.border = bdr
                # 誤差V = 測定電圧 - 期待値（数式）
                exp_v = expected_vals[ri]
                if v != '':
                    c6 = ws.cell(row=r, column=6, value=f'=D{r}-({exp_v})')
                else:
                    c6 = ws.cell(row=r, column=6, value='')
                c6.font = dat_font; c6.alignment = right_align; c6.border = bdr
                c6.number_format = '0.000'
                # 誤差% = 誤差V / 期待値 * 100（数式、centerはグレー空欄）
                c7 = ws.cell(row=r, column=7)
                if is_center:
                    c7.value = ''
                    c7.fill = gray_fill
                elif v != '' and exp_v != 0:
                    c7.value = f'=F{r}/{exp_v}*100'
                    c7.number_format = '0.000'
                else:
                    c7.value = ''
                c7.font = dat_font; c7.alignment = right_align; c7.border = bdr
                # 判定（中央）
                c8 = ws.cell(row=r, column=8, value=judge)
                c8.font = dat_font; c8.alignment = center; c8.border = bdr

            # ﾕﾆｯﾄNo: 6行結合
            ws.merge_cells(start_row=base, start_column=1, end_row=base + 5, end_column=1)
            c1 = ws.cell(row=base, column=1, value=unit_label)
            c1.font = dat_font; c1.alignment = center
            for ri in range(6):
                ws.cell(row=base + ri, column=1).border = border_top_thick if ri == 0 else border_all

            # 極POS: 3行結合
            ws.merge_cells(start_row=base, start_column=2, end_row=base + 2, end_column=2)
            ws.cell(row=base, column=2, value='POS').font = dat_font
            ws.cell(row=base, column=2).alignment = center
            for ri in range(3):
                ws.cell(row=base + ri, column=2).border = border_top_thick if ri == 0 else border_all

            # 極NEG: 3行結合
            ws.merge_cells(start_row=base + 3, start_column=2, end_row=base + 5, end_column=2)
            ws.cell(row=base + 3, column=2, value='NEG').font = dat_font
            ws.cell(row=base + 3, column=2).alignment = center
            for ri in range(3, 6):
                ws.cell(row=base + ri, column=2).border = border_all

        for r in range(1, 26 + 1):
            ws.row_dimensions[r].height = 16
        for c, w in {1: 14, 2: 6, 3: 12, 4: 14, 5: 22, 6: 10, 7: 10, 8: 6}.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    def _write_lbc_sheet(self, ws, slots_data):
        """LBC XLSX書き込み（docx仕様: 9列、4DEF固定枠、セル結合、数式埋込、小数3桁）

        slots_data: [(serial_no, display_data), ...] 最大 4 枠
        """
        thin = Side(style='thin')
        thick = Side(style='medium')
        border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
        border_top_thick = Border(top=thick, left=thin, right=thin, bottom=thin)
        dat_font = Font(name='MS Pゴシック', size=10)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')

        headers = ['ﾕﾆｯﾄ No.', 'LBC\nATT', '入力ｺｰﾄﾞ', 'POS 出力電圧(V)', 'NEG 出力電圧(V)',
                   '許容差(V)', 'POS 誤差\n(V)', 'NEG 誤差\n(V)', '判定']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = dat_font; cell.alignment = center; cell.border = border_all

        # 期待値数値（POS/NEG）: ATT 1/1, 1/2, 1/4 の FFFF/0000
        expected_pos = [6.180, -6.180, 3.090, -3.090, 1.545, -1.545]
        expected_neg = [-6.180, 6.180, -3.090, 3.090, -1.545, 1.545]
        expected_strs = ['±6.180±0.062', '', '±3.090±0.031', '', '±1.545±0.015', '']
        code_strs = ['FFFF H', '0000 H', 'FFFF H', '0000 H', 'FFFF H', '0000 H']

        for slot in range(4):
            base = 2 + slot * 6
            if slot < len(slots_data):
                sn_slot, slot_data = slots_data[slot]
                unit_label = f"1PB397MK2\n{sn_slot}"
            else:
                slot_data = None
                unit_label = ""

            for ri in range(6):
                r = base + ri
                bdr = border_top_thick if ri == 0 else border_all

                if slot_data and ri < len(slot_data):
                    d = slot_data[ri]
                    is_cal_ng = (d['judge'] == 'CAL NG')
                    vp = d['voltage_pos'] if d['voltage_pos'] is not None else ''
                    vn = d['voltage_neg'] if d['voltage_neg'] is not None else ''
                    judge = d['judge']
                else:
                    is_cal_ng = False
                    vp, vn, judge = '', '', ''

                c3 = ws.cell(row=r, column=3, value=code_strs[ri])
                c3.font = dat_font; c3.alignment = right_align; c3.border = bdr

                if is_cal_ng:
                    # CAL NG: 電圧欄に"CAL NG"、誤差・判定も"CAL NG"
                    c4 = ws.cell(row=r, column=4, value='CAL NG')
                    c4.font = dat_font; c4.alignment = center; c4.border = bdr
                    c5 = ws.cell(row=r, column=5, value='CAL NG')
                    c5.font = dat_font; c5.alignment = center; c5.border = bdr
                    c7 = ws.cell(row=r, column=7, value='')
                    c7.font = dat_font; c7.alignment = right_align; c7.border = bdr
                    c8 = ws.cell(row=r, column=8, value='')
                    c8.font = dat_font; c8.alignment = right_align; c8.border = bdr
                    c9 = ws.cell(row=r, column=9, value='CAL NG')
                    c9.font = dat_font; c9.alignment = center; c9.border = bdr
                else:
                    # POS電圧
                    c4 = ws.cell(row=r, column=4, value=vp)
                    c4.font = dat_font; c4.alignment = right_align; c4.border = bdr
                    if vp != '': c4.number_format = '0.000'
                    # NEG電圧
                    c5 = ws.cell(row=r, column=5, value=vn)
                    c5.font = dat_font; c5.alignment = right_align; c5.border = bdr
                    if vn != '': c5.number_format = '0.000'
                    # POS誤差 = POS電圧 - 期待値（数式）
                    if vp != '':
                        c7 = ws.cell(row=r, column=7, value=f'=D{r}-({expected_pos[ri]})')
                    else:
                        c7 = ws.cell(row=r, column=7, value='')
                    c7.font = dat_font; c7.alignment = right_align; c7.border = bdr
                    c7.number_format = '0.000'
                    # NEG誤差 = NEG電圧 - 期待値（数式）
                    if vn != '':
                        c8 = ws.cell(row=r, column=8, value=f'=E{r}-({expected_neg[ri]})')
                    else:
                        c8 = ws.cell(row=r, column=8, value='')
                    c8.font = dat_font; c8.alignment = right_align; c8.border = bdr
                    c8.number_format = '0.000'
                    # 判定
                    c9 = ws.cell(row=r, column=9, value=judge)
                    c9.font = dat_font; c9.alignment = center; c9.border = bdr

            # ﾕﾆｯﾄNo: 6行結合
            ws.merge_cells(start_row=base, start_column=1, end_row=base + 5, end_column=1)
            ws.cell(row=base, column=1, value=unit_label).font = dat_font
            ws.cell(row=base, column=1).alignment = center
            for ri in range(6):
                ws.cell(row=base + ri, column=1).border = border_top_thick if ri == 0 else border_all

            # ATT: 各2行結合
            for ai, att_val in enumerate(['1/1', '1/2', '1/4']):
                sr = base + ai * 2
                ws.merge_cells(start_row=sr, start_column=2, end_row=sr + 1, end_column=2)
                ws.cell(row=sr, column=2, value=att_val).font = dat_font
                ws.cell(row=sr, column=2).alignment = center
                for ri in range(2):
                    ws.cell(row=sr + ri, column=2).border = border_top_thick if (ai == 0 and ri == 0) else border_all

            # 許容差: 各2行結合
            for ai, exp_val in enumerate(['±6.180±0.062', '±3.090±0.031', '±1.545±0.015']):
                sr = base + ai * 2
                ws.merge_cells(start_row=sr, start_column=6, end_row=sr + 1, end_column=6)
                ws.cell(row=sr, column=6, value=exp_val).font = dat_font
                ws.cell(row=sr, column=6).alignment = right_align
                for ri in range(2):
                    ws.cell(row=sr + ri, column=6).border = border_top_thick if (ai == 0 and ri == 0) else border_all

        ws.row_dimensions[1].height = 32
        for r in range(2, 26):
            ws.row_dimensions[r].height = 16
        for c, w in {1: 14, 2: 6, 3: 12, 4: 14, 5: 14, 6: 18, 7: 10, 8: 10, 9: 6}.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    def _write_moni_sheet(self, ws, slots_data):
        """moni XLSX書き込み（docx仕様: 4DEF固定枠、セル結合、右寄せ、数式埋込、小数2桁）

        slots_data: [(serial_no, display_data), ...] 最大 4 枠
        """
        thin = Side(style='thin')
        thick = Side(style='medium')
        border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
        border_top_thick = Border(top=thick, left=thin, right=thin, bottom=thin)
        dat_font = Font(name='MS Pゴシック', size=10)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')

        headers = ['ﾕﾆｯﾄ No.', '極', '入力ｺｰﾄﾞ', '測定電圧(V)', '許容誤差(V)(<1%)', '誤差(V)', '誤差(%)', '判定']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = dat_font; cell.alignment = center; cell.border = border_all

        # 表示順: POS FFFFF, POS 00000, NEG 00000, NEG FFFFF
        expected_strs = ['+5.00±0.05', '-5.00±0.05', '+5.00±0.05', '-5.00±0.05']
        expected_vals = [5.0, -5.0, 5.0, -5.0]
        code_strs = ['FFFFF H', '00000 H', '00000 H', 'FFFFF H']

        for slot in range(4):
            base = 2 + slot * 4
            if slot < len(slots_data):
                sn_slot, slot_data = slots_data[slot]
                unit_label = f"1PB397MK2\n{sn_slot}"
            else:
                slot_data = None
                unit_label = ""

            for ri in range(4):
                r = base + ri
                bdr = border_top_thick if ri == 0 else border_all

                if slot_data and ri < len(slot_data):
                    d = slot_data[ri]
                    v = d['voltage'] if d['voltage'] is not None else ''
                    judge = d['judge']
                else:
                    v, judge = '', ''

                c3 = ws.cell(row=r, column=3, value=code_strs[ri])
                c3.font = dat_font; c3.alignment = right_align; c3.border = bdr
                c4 = ws.cell(row=r, column=4, value=v)
                c4.font = dat_font; c4.alignment = right_align; c4.border = bdr
                if v != '': c4.number_format = '0.00'
                c5 = ws.cell(row=r, column=5, value=expected_strs[ri])
                c5.font = dat_font; c5.alignment = right_align; c5.border = bdr
                # 誤差V = 測定電圧 - 期待値（数式）
                exp_v = expected_vals[ri]
                if v != '':
                    c6 = ws.cell(row=r, column=6, value=f'=D{r}-({exp_v})')
                else:
                    c6 = ws.cell(row=r, column=6, value='')
                c6.font = dat_font; c6.alignment = right_align; c6.border = bdr
                c6.number_format = '0.00'
                # 誤差% = 誤差V / 期待値 * 100
                c7 = ws.cell(row=r, column=7)
                if v != '' and exp_v != 0:
                    c7.value = f'=F{r}/{exp_v}*100'
                    c7.number_format = '0.00'
                else:
                    c7.value = ''
                c7.font = dat_font; c7.alignment = right_align; c7.border = bdr
                c8 = ws.cell(row=r, column=8, value=judge)
                c8.font = dat_font; c8.alignment = center; c8.border = bdr

            # ﾕﾆｯﾄNo: 4行結合
            ws.merge_cells(start_row=base, start_column=1, end_row=base + 3, end_column=1)
            c1 = ws.cell(row=base, column=1, value=unit_label)
            c1.font = dat_font; c1.alignment = center
            for ri in range(4):
                ws.cell(row=base + ri, column=1).border = border_top_thick if ri == 0 else border_all

            # 極POS: 2行結合
            ws.merge_cells(start_row=base, start_column=2, end_row=base + 1, end_column=2)
            ws.cell(row=base, column=2, value='POS').font = dat_font
            ws.cell(row=base, column=2).alignment = center
            for ri in range(2):
                ws.cell(row=base + ri, column=2).border = border_top_thick if ri == 0 else border_all

            # 極NEG: 2行結合
            ws.merge_cells(start_row=base + 2, start_column=2, end_row=base + 3, end_column=2)
            ws.cell(row=base + 2, column=2, value='NEG').font = dat_font
            ws.cell(row=base + 2, column=2).alignment = center
            for ri in range(2, 4):
                ws.cell(row=base + ri, column=2).border = border_all

        for r in range(1, 18):
            ws.row_dimensions[r].height = 16
        for c, w in {1: 14, 2: 6, 3: 12, 4: 14, 5: 20, 6: 10, 7: 10, 8: 6}.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    # ==================== PNG生成（Excel COM経由スクリーンショット） ====================
    def _generate_table_png(self, sheet_name, slots_data, save_dir, timestamp):
        """XLSXを開いて表範囲をCopyPicture→PNG保存

        slots_data: [(serial_no, display_data), ...]
        ファイル名は選択 DEF の S/N を全て `_` 連結 (XLSX と揃える)。

        低性能 PC でフリーズする問題への対策:
        - DispatchEx で毎回新規 Excel.exe を起動 (前回の Quit 待ちのゾンビ回避)
        - CoInitialize/CoUninitialize の try/finally を正しく対応付け
        - クリップボード取得は短いリトライ (CopyPicture の遅延吸収)
        - 各 Excel COM ステップを [DBG] ログで出力
        """
        sheet_titles = {'position': 'POSTION', 'lbc': 'LBC', 'moni': 'moni'}
        sheet_title = sheet_titles.get(sheet_name, sheet_name)
        sn_joined = self._join_sns(slots_data)

        xlsx_path = os.path.join(
            save_dir, f"{sn_joined}_DC特性_{sheet_title}_{timestamp}.xlsx")
        png_path = os.path.join(
            save_dir, f"{sn_joined}_DC特性_{sheet_title}_{timestamp}.png")

        if not os.path.exists(xlsx_path):
            return png_path

        def _dbg(msg):
            self._update_queue.put((
                'log', f"[DBG] dc_png({sn_joined}/{sheet_title}): {msg}"))

        try:
            import win32com.client
            from PIL import ImageGrab
            import pythoncom

            _dbg("CoInitialize")
            pythoncom.CoInitialize()
            try:
                _dbg("DispatchEx")
                excel = win32com.client.DispatchEx("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                try:
                    _dbg(f"Open {os.path.basename(xlsx_path)}")
                    wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
                    ws = wb.Sheets(1)

                    # 表範囲
                    if sheet_name == 'position':
                        last_row = 25   # 1 header + 24 data
                        last_col = 'H'  # 8列
                    elif sheet_name == 'moni':
                        last_row = 17   # 1 header + 16 data
                        last_col = 'H'  # 8列
                    else:  # lbc
                        last_row = 25   # 1 header + 24 data
                        last_col = 'I'  # 9列

                    _dbg(f"CopyPicture A1:{last_col}{last_row}")
                    rng = ws.Range(f"A1:{last_col}{last_row}")
                    rng.CopyPicture(Appearance=1, Format=2)  # xlScreen, xlBitmap

                    # クリップボード取得（低性能 PC 対策: 0.1s × 最大 20 回）
                    _dbg("grabclipboard wait")
                    img = None
                    for attempt in range(20):
                        time.sleep(0.1)
                        try:
                            img = ImageGrab.grabclipboard()
                        except Exception as e_clip:
                            _dbg(f"grabclipboard exception #{attempt}: {e_clip}")
                            img = None
                        if img is not None:
                            _dbg(f"grabclipboard ok at #{attempt}")
                            break

                    if img:
                        _dbg("img.save")
                        img.save(os.path.abspath(png_path))
                        _dbg("img.save done")
                    else:
                        _dbg("grabclipboard None after retries")
                        self._update_queue.put((
                            'log',
                            f"[DC特性] PNG: クリップボード取得失敗 "
                            f"({serial_no}/{sheet_title})"))

                    _dbg("wb.Close")
                    wb.Close(False)
                finally:
                    _dbg("excel.Quit")
                    try:
                        excel.Quit()
                    except Exception as e_q:
                        _dbg(f"excel.Quit error: {e_q}")
                    _dbg("excel.Quit done")
            finally:
                _dbg("CoUninitialize")
                pythoncom.CoUninitialize()
                _dbg("CoUninitialize done")
        except Exception as e:
            self._update_queue.put((
                'log', f"[DC特性] PNG生成エラー: {e}"))

        return png_path
