import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import csv
import json
import os
from utils.graph_plotter import LSBGraphPlotter


class GraphTab(ttk.Frame):
    """
    グラフ描画タブ
    
    機能:
    - CSVファイル選択・読み込み
    - LSB変換設定（bit精度、基準電圧、LSB/div）- 自動保存機能付き
    - 表示データ選択（シリアルNo.、POS/NEG）- デフォルト全選択
    - グラフ表示・自動更新
    """
    
    SETTINGS_FILE = "graph_settings.json"
    
    def __init__(self, parent, gpib_controller):
        super().__init__(parent)
        self.gpib = gpib_controller
        
        # データ保持
        self.csv_data = None
        self.temp_csv_data = None  # 温度CSVデータ
        self.serial_numbers = []
        self.checkboxes = {}
        self.graph_windows = []  # 開いているグラフウィンドウのリスト

        # 温特グラフ詳細設定の保持（ウィンドウ外でも値を保持）
        self._png_scale_setting = "1.0"  # PNG保存縮尺
        self._show_temp_arrows = True  # 温度区間矢印表示
        # 温度区間Div範囲（23℃、28℃、18℃の開始・終了Div）
        self._temp_zone_divs = [
            (0, 6),    # 23℃: 0〜6 Div
            (6, 13),   # 28℃: 6〜13 Div
            (13, 19)   # 18℃: 13〜19 Div
        ]
        
        # 設定変更の監視フラグ（初期化中は保存しない）
        self._initializing = True
        
        self.create_widgets()
        self.load_settings()
        
        # 設定値の変更を監視
        self._setup_auto_save()
        
        # 初期化完了
        self._initializing = False
    
    def create_widgets(self):
        """ウィジェットを作成"""
        # ファイル選択フレーム
        self._create_file_selection_frame()

        # 設定フレーム（グラフ表示ボタン含む）
        self._create_settings_frame()

        # データ選択フレーム
        self._create_data_selection_frame()
    
    def _create_file_selection_frame(self):
        """ファイル選択フレームを作成"""
        file_frame = ttk.LabelFrame(self, text="CSVファイル選択", padding=10)
        file_frame.pack(fill=tk.X, padx=10, pady=5)

        # 測定CSVファイル選択
        measurement_frame = ttk.Frame(file_frame)
        measurement_frame.pack(fill=tk.X, pady=2)
        ttk.Label(measurement_frame, text="測定CSV:", width=10).pack(side=tk.LEFT, padx=5)
        self.file_path_var = tk.StringVar()
        ttk.Entry(measurement_frame, textvariable=self.file_path_var, state='readonly',
                  width=70).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(measurement_frame, text="読み込み",
                   command=self.load_csv_file).pack(side=tk.LEFT, padx=5)

        # 温度CSVファイル選択
        temp_frame = ttk.Frame(file_frame)
        temp_frame.pack(fill=tk.X, pady=2)
        ttk.Label(temp_frame, text="温度CSV:", width=10).pack(side=tk.LEFT, padx=5)
        self.temp_file_path_var = tk.StringVar()
        ttk.Entry(temp_frame, textvariable=self.temp_file_path_var, state='readonly',
                  width=70).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(temp_frame, text="読み込み",
                   command=self.load_temp_csv_file).pack(side=tk.LEFT, padx=5)
    
    def _create_settings_frame(self):
        """設定フレームを作成"""
        # 設定フレームを横並びにするためのコンテナ（均等幅）
        settings_container = ttk.Frame(self)
        settings_container.pack(fill=tk.X, padx=10, pady=5)
        settings_container.columnconfigure(0, weight=1)
        settings_container.columnconfigure(1, weight=1)

        # 左側: 変換設定
        settings_frame = ttk.LabelFrame(settings_container, text="変換設定", padding=10)
        settings_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))

        # Bit精度設定
        bit_frame = ttk.Frame(settings_frame)
        bit_frame.pack(fill=tk.X, pady=2)
        ttk.Label(bit_frame, text="Bit精度:").pack(side=tk.LEFT, padx=5)
        self.bit_precision_var = tk.StringVar(value="24")
        ttk.Entry(bit_frame, textvariable=self.bit_precision_var,
                  width=10).pack(side=tk.LEFT, padx=5)
        ttk.Label(bit_frame, text="bit").pack(side=tk.LEFT)

        # 基準電圧設定（幅を2/3に）
        ref_frame = ttk.Frame(settings_frame)
        ref_frame.pack(fill=tk.X, pady=2)
        ttk.Label(ref_frame, text="+Full基準電圧:").pack(side=tk.LEFT, padx=5)
        self.pos_full_var = tk.StringVar(value="10.0")
        ttk.Entry(ref_frame, textvariable=self.pos_full_var,
                  width=7).pack(side=tk.LEFT, padx=5)
        ttk.Label(ref_frame, text="V").pack(side=tk.LEFT, padx=(0, 10))

        ttk.Label(ref_frame, text="-Full基準電圧:").pack(side=tk.LEFT, padx=5)
        self.neg_full_var = tk.StringVar(value="-10.0")
        ttk.Entry(ref_frame, textvariable=self.neg_full_var,
                  width=7).pack(side=tk.LEFT, padx=5)
        ttk.Label(ref_frame, text="V").pack(side=tk.LEFT)

        # Y軸範囲設定（2行に分割）
        yaxis_frame1 = ttk.Frame(settings_frame)
        yaxis_frame1.pack(fill=tk.X, pady=2)
        ttk.Label(yaxis_frame1, text="Y軸範囲:", width=10).pack(side=tk.LEFT, padx=5)
        self.yaxis_mode_var = tk.StringVar(value="auto")
        ttk.Radiobutton(yaxis_frame1, text="オート", variable=self.yaxis_mode_var,
                        value="auto").pack(side=tk.LEFT)

        yaxis_frame2 = ttk.Frame(settings_frame)
        yaxis_frame2.pack(fill=tk.X, pady=2)
        ttk.Label(yaxis_frame2, text="", width=10).pack(side=tk.LEFT, padx=5)  # インデント用
        ttk.Radiobutton(yaxis_frame2, text="設定値", variable=self.yaxis_mode_var,
                        value="manual").pack(side=tk.LEFT)
        ttk.Label(yaxis_frame2, text="Min:").pack(side=tk.LEFT, padx=(10, 2))
        self.yaxis_min_var = tk.StringVar(value="-50")
        ttk.Entry(yaxis_frame2, textvariable=self.yaxis_min_var,
                  width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(yaxis_frame2, text="Max:").pack(side=tk.LEFT, padx=(10, 2))
        self.yaxis_max_var = tk.StringVar(value="50")
        ttk.Entry(yaxis_frame2, textvariable=self.yaxis_max_var,
                  width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(yaxis_frame2, text="LSB").pack(side=tk.LEFT)

        # LSB/Div設定 - 順序変更
        lsb_frame = ttk.Frame(settings_frame)
        lsb_frame.pack(fill=tk.X, pady=2)
        ttk.Label(lsb_frame, text="縦軸LSB/Div:").pack(side=tk.LEFT, padx=5)
        self.lsb_per_div_var = tk.StringVar(value="10")
        ttk.Entry(lsb_frame, textvariable=self.lsb_per_div_var,
                  width=10).pack(side=tk.LEFT, padx=5)
        ttk.Label(lsb_frame, text="LSB").pack(side=tk.LEFT)

        # 基準電圧計算方法
        ref_mode_frame = ttk.Frame(settings_frame)
        ref_mode_frame.pack(fill=tk.X, pady=2)
        ttk.Label(ref_mode_frame, text="基準電圧:").pack(side=tk.LEFT, padx=5)
        self.ref_mode_var = tk.StringVar(value="ideal")
        ttk.Radiobutton(ref_mode_frame, text="理想値", variable=self.ref_mode_var,
                        value="ideal").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(ref_mode_frame, text="全平均", variable=self.ref_mode_var,
                        value="all_avg").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(ref_mode_frame, text="区間別平均", variable=self.ref_mode_var,
                        value="section_avg").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(ref_mode_frame, text="初回平均", variable=self.ref_mode_var,
                        value="first_avg").pack(side=tk.LEFT, padx=5)

        # スキップ設定フレーム
        skip_frame = ttk.Frame(settings_frame)
        skip_frame.pack(fill=tk.X, pady=2)

        # コード切替後スキップ行数
        ttk.Label(skip_frame, text="切替後スキップ:").pack(side=tk.LEFT, padx=5)
        self.skip_after_change_var = tk.StringVar(value="0")
        ttk.Entry(skip_frame, textvariable=self.skip_after_change_var,
                  width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(skip_frame, text="行").pack(side=tk.LEFT, padx=(0, 15))

        # パターン開始最初のデータを省く
        self.skip_first_data_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(skip_frame, text="開始時スキップ",
                        variable=self.skip_first_data_var).pack(side=tk.LEFT, padx=5)

        # 切替わり1周回前を省く
        self.skip_before_change_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(skip_frame, text="切替前スキップ",
                        variable=self.skip_before_change_var).pack(side=tk.LEFT, padx=5)

        # グラフ表示ボタン（変換設定内）
        ttk.Separator(settings_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=5)
        ttk.Button(settings_frame, text="選択したデータをグラフ表示",
                   command=self.plot_selected_data).pack(pady=5)

        # 右側: 温特グラフ設定
        temp_settings_frame = ttk.LabelFrame(settings_container, text="温特グラフ設定", padding=15)
        temp_settings_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))

        # 固定値・初期値の注記（Y軸関連を上に）
        ttk.Label(temp_settings_frame, text="【固定値・初期値】",
                  font=('', 9, 'bold')).pack(anchor=tk.W, pady=(0, 8))
        ttk.Label(temp_settings_frame, text="・Y軸(LSB)初期値: ±8LSB, 2LSB/Div").pack(anchor=tk.W, padx=15, pady=1)
        ttk.Label(temp_settings_frame, text="・Y軸(温度): ±8℃, 2℃/Div").pack(anchor=tk.W, padx=15, pady=1)
        ttk.Label(temp_settings_frame, text="・基準電圧モード: 初回平均").pack(anchor=tk.W, padx=15, pady=1)
        ttk.Label(temp_settings_frame, text="・X軸: 10分/Div").pack(anchor=tk.W, padx=15, pady=1)

        # 温特グラフで使用する設定（共通設定を参照）
        ttk.Separator(temp_settings_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=12)
        ttk.Label(temp_settings_frame, text="【共通設定を使用】",
                  font=('', 9, 'bold')).pack(anchor=tk.W, pady=(0, 8))
        ttk.Label(temp_settings_frame, text="・Bit精度").pack(anchor=tk.W, padx=15, pady=1)
        ttk.Label(temp_settings_frame, text="・スキップ設定").pack(anchor=tk.W, padx=15, pady=1)

        # 温特グラフ表示ボタン（温特グラフ設定内）
        ttk.Separator(temp_settings_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=8)
        ttk.Button(temp_settings_frame, text="温特グラフ表示",
                   command=self.plot_temperature_graph).pack(pady=8)
    
    def _create_data_selection_frame(self):
        """データ選択フレームを作成"""
        selection_frame = ttk.LabelFrame(self, text="表示データ選択", padding=10)
        selection_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # スクロール可能なフレーム
        canvas = tk.Canvas(selection_frame, height=100)
        scrollbar = ttk.Scrollbar(selection_frame, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    def _setup_auto_save(self):
        """設定値の自動保存を設定"""
        # StringVarの変更を監視
        self.bit_precision_var.trace_add('write', self._on_setting_changed)
        self.pos_full_var.trace_add('write', self._on_setting_changed)
        self.neg_full_var.trace_add('write', self._on_setting_changed)
        self.lsb_per_div_var.trace_add('write', self._on_setting_changed)
        self.ref_mode_var.trace_add('write', self._on_setting_changed)
        self.yaxis_mode_var.trace_add('write', self._on_setting_changed)
        self.yaxis_min_var.trace_add('write', self._on_setting_changed)
        self.yaxis_max_var.trace_add('write', self._on_setting_changed)
        self.skip_after_change_var.trace_add('write', self._on_setting_changed)
        self.skip_first_data_var.trace_add('write', self._on_setting_changed)
        self.skip_before_change_var.trace_add('write', self._on_setting_changed)
    
    def _on_setting_changed(self, *args):
        """設定値が変更されたときに自動保存＆グラフ更新"""
        if not self._initializing:
            self.save_settings()
            self.update_all_graphs()
    
    def load_settings(self):
        """設定を読み込み"""
        if os.path.exists(self.SETTINGS_FILE):
            try:
                with open(self.SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                
                self.bit_precision_var.set(settings.get('bit_precision', '24'))
                self.pos_full_var.set(settings.get('pos_full_voltage', '10.0'))
                self.neg_full_var.set(settings.get('neg_full_voltage', '-10.0'))
                self.lsb_per_div_var.set(settings.get('lsb_per_div', '10'))
                self.ref_mode_var.set(settings.get('ref_mode', 'ideal'))
                self.yaxis_mode_var.set(settings.get('yaxis_mode', 'auto'))
                self.yaxis_min_var.set(settings.get('yaxis_min', '-50'))
                self.yaxis_max_var.set(settings.get('yaxis_max', '50'))
                self.skip_after_change_var.set(settings.get('skip_after_change', '0'))
                self.skip_first_data_var.set(settings.get('skip_first_data', False))
                self.skip_before_change_var.set(settings.get('skip_before_change', False))

                # CSVファイルパスを復元
                csv_path = settings.get('csv_file_path', '')
                if csv_path and os.path.exists(csv_path):
                    self.file_path_var.set(csv_path)
                    self._load_csv_from_path(csv_path, show_message=False)

                # 温度CSVファイルパスを復元
                temp_csv_path = settings.get('temp_csv_file_path', '')
                if temp_csv_path and os.path.exists(temp_csv_path):
                    self.temp_file_path_var.set(temp_csv_path)
                    self._load_temp_csv_from_path(temp_csv_path, show_message=False)

                # 温特グラフ詳細設定
                self._png_scale_setting = settings.get('png_scale', '1.0')
                self._show_temp_arrows = settings.get('show_temp_arrows', True)
                # Div範囲を読み込み（リストのリストからタプルのリストに変換）
                divs = settings.get('temp_zone_divs', [[0, 6], [6, 13], [13, 19]])
                self._temp_zone_divs = [(d[0], d[1]) for d in divs]

            except Exception as e:
                print(f"設定の読み込みに失敗しました: {e}")
    
    def save_settings(self):
        """設定を保存（自動）"""
        try:
            settings = {
                'bit_precision': self.bit_precision_var.get(),
                'pos_full_voltage': self.pos_full_var.get(),
                'neg_full_voltage': self.neg_full_var.get(),
                'lsb_per_div': self.lsb_per_div_var.get(),
                'ref_mode': self.ref_mode_var.get(),
                'yaxis_mode': self.yaxis_mode_var.get(),
                'yaxis_min': self.yaxis_min_var.get(),
                'yaxis_max': self.yaxis_max_var.get(),
                'skip_after_change': self.skip_after_change_var.get(),
                'skip_first_data': self.skip_first_data_var.get(),
                'skip_before_change': self.skip_before_change_var.get(),
                'csv_file_path': self.file_path_var.get(),
                'temp_csv_file_path': self.temp_file_path_var.get(),
                'png_scale': self._png_scale_setting,
                'show_temp_arrows': self._show_temp_arrows,
                'temp_zone_divs': [list(d) for d in self._temp_zone_divs]
            }
            
            with open(self.SETTINGS_FILE, 'w', encoding='utf-8') as f:
                json.dump(settings, f, indent=4)
            
        except Exception as e:
            print(f"設定の保存に失敗しました: {e}")
    
    def update_all_graphs(self):
        """開いている全てのグラフを更新"""
        # 設定値を取得
        try:
            bit_precision = int(self.bit_precision_var.get())
            pos_full = float(self.pos_full_var.get())
            neg_full = float(self.neg_full_var.get())
            lsb_per_div = float(self.lsb_per_div_var.get())
            ref_mode = self.ref_mode_var.get()
            yaxis_mode = self.yaxis_mode_var.get()
            yaxis_min = float(self.yaxis_min_var.get()) if yaxis_mode == "manual" else None
            yaxis_max = float(self.yaxis_max_var.get()) if yaxis_mode == "manual" else None
            skip_after_change = int(self.skip_after_change_var.get())
            skip_first_data = self.skip_first_data_var.get()
            skip_before_change = self.skip_before_change_var.get()
        except ValueError:
            return

        # 新しいプロッターを作成
        plotter = LSBGraphPlotter(bit_precision, pos_full, neg_full, lsb_per_div, ref_mode,
                                  yaxis_mode, yaxis_min, yaxis_max,
                                  skip_after_change, skip_first_data, skip_before_change)
        
        # 存在するウィンドウのみ更新
        valid_windows = []
        for window in self.graph_windows:
            try:
                if window.winfo_exists():
                    plotter.update_plot_window(window)
                    valid_windows.append(window)
            except:
                pass
        
        # 有効なウィンドウのリストを更新
        self.graph_windows = valid_windows
    
    def load_csv_file(self):
        """CSVファイルを選択して読み込み"""
        # ファイル選択ダイアログ
        filename = filedialog.askopenfilename(
            title="CSVファイルを選択",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )

        if not filename:
            return

        self.file_path_var.set(filename)
        self._load_csv_from_path(filename, show_message=True)
        # ファイルパスを保存
        self.save_settings()

    def _load_csv_from_path(self, filename, show_message=True):
        """指定パスからCSVファイルを読み込み"""
        try:
            with open(filename, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                self.csv_data = list(reader)
                headers = reader.fieldnames

            if not self.csv_data:
                if show_message:
                    messagebox.showerror("エラー", "CSVファイルにデータがありません")
                return

            # シリアルNo.を抽出
            self._extract_serial_numbers(headers)

            # チェックボックスを作成（デフォルト全選択）
            self._create_checkboxes()

            if show_message:
                messagebox.showinfo("成功", f"CSVファイルを読み込みました\n{len(self.csv_data)}行のデータ")

        except Exception as e:
            if show_message:
                messagebox.showerror("エラー", f"CSVファイルの読み込みに失敗しました:\n{str(e)}")
    
    def _extract_serial_numbers(self, headers):
        """ヘッダーからシリアルNo.を抽出"""
        self.serial_numbers = []
        for header in headers:
            if header.endswith('_POS') or header.endswith('_NEG'):
                parts = header.rsplit('_', 1)
                serial = parts[0]
                pole = parts[1]
                
                if serial not in [sn for sn, _ in self.serial_numbers]:
                    self.serial_numbers.append((serial, []))
                
                for i, (sn, poles) in enumerate(self.serial_numbers):
                    if sn == serial and pole not in poles:
                        self.serial_numbers[i] = (sn, poles + [pole])
    
    def _create_checkboxes(self):
        """データ選択用チェックボックスを作成（デフォルト全選択）"""
        # 既存のチェックボックスをクリア
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        self.checkboxes = {}
        
        # シリアルNo.ごとにチェックボックスを作成
        for serial, poles in self.serial_numbers:
            frame = ttk.Frame(self.scrollable_frame)
            frame.pack(fill=tk.X, pady=2)
            
            ttk.Label(frame, text=f"{serial}:", width=15).pack(side=tk.LEFT, padx=5)
            
            for pole in poles:
                var = tk.BooleanVar(value=True)  # デフォルト: 選択状態
                cb = ttk.Checkbutton(frame, text=pole, variable=var)
                cb.pack(side=tk.LEFT, padx=10)
                self.checkboxes[f"{serial}_{pole}"] = var
    
    def plot_selected_data(self):
        """選択されたデータをグラフ表示"""
        if not self.csv_data:
            messagebox.showerror("エラー", "CSVファイルを読み込んでください")
            return

        # 設定値を取得
        try:
            bit_precision = int(self.bit_precision_var.get())
            pos_full = float(self.pos_full_var.get())
            neg_full = float(self.neg_full_var.get())
            lsb_per_div = float(self.lsb_per_div_var.get())
            ref_mode = self.ref_mode_var.get()
            yaxis_mode = self.yaxis_mode_var.get()
            yaxis_min = float(self.yaxis_min_var.get()) if yaxis_mode == "manual" else None
            yaxis_max = float(self.yaxis_max_var.get()) if yaxis_mode == "manual" else None
            skip_after_change = int(self.skip_after_change_var.get())
            skip_first_data = self.skip_first_data_var.get()
            skip_before_change = self.skip_before_change_var.get()
        except ValueError:
            messagebox.showerror("エラー", "設定値が不正です")
            return

        # LSBGraphPlotterを作成
        plotter = LSBGraphPlotter(bit_precision, pos_full, neg_full, lsb_per_div, ref_mode,
                                  yaxis_mode, yaxis_min, yaxis_max,
                                  skip_after_change, skip_first_data, skip_before_change)
        
        # 選択されたデータをプロット
        plot_count = 0
        for key, var in self.checkboxes.items():
            if var.get():
                serial, pole = key.rsplit('_', 1)
                window = plotter.plot_csv_data(self, self.csv_data, serial, pole)
                if window:
                    self.graph_windows.append(window)
                    plot_count += 1
        
        if plot_count == 0:
            messagebox.showwarning("警告", "表示するデータを選択してください")

    def load_temp_csv_file(self):
        """温度CSVファイルを選択して読み込み"""
        filename = filedialog.askopenfilename(
            title="温度CSVファイルを選択",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )

        if not filename:
            return

        self.temp_file_path_var.set(filename)
        self._load_temp_csv_from_path(filename, show_message=True)
        self.save_settings()

    def _load_temp_csv_from_path(self, filename, show_message=True):
        """指定パスから温度CSVファイルを読み込み"""
        try:
            with open(filename, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                self.temp_csv_data = list(reader)

            if not self.temp_csv_data:
                if show_message:
                    messagebox.showerror("エラー", "温度CSVファイルにデータがありません")
                return

            if show_message:
                messagebox.showinfo("成功", f"温度CSVを読み込みました\n{len(self.temp_csv_data)}行のデータ")

        except Exception as e:
            if show_message:
                messagebox.showerror("エラー", f"温度CSVの読み込みに失敗しました:\n{str(e)}")

    def plot_temperature_graph(self):
        """温特グラフを表示（2軸: LSB変動 + 温度差）+ 設定画面"""
        if not self.csv_data:
            messagebox.showerror("エラー", "測定CSVファイルを読み込んでください")
            return

        if not self.temp_csv_data:
            messagebox.showerror("エラー", "温度CSVファイルを読み込んでください")
            return

        # 選択されたデータを確認
        selected_keys = [key for key, var in self.checkboxes.items() if var.get()]
        if not selected_keys:
            messagebox.showwarning("警告", "表示するデータを選択してください")
            return

        # 設定画面ウィンドウを作成
        self._create_temp_graph_settings_window(selected_keys)

    def _create_temp_graph_settings_window(self, selected_keys):
        """温特グラフ設定画面ウィンドウを作成"""
        # 既存のウィンドウがあれば閉じる
        if hasattr(self, 'temp_settings_window') and self.temp_settings_window:
            try:
                if self.temp_settings_window.winfo_exists():
                    self.temp_settings_window.destroy()
            except:
                pass

        # 設定画面ウィンドウ（温特グラフと被らないように右端に配置）
        self.temp_settings_window = tk.Toplevel(self)
        self.temp_settings_window.title("温特グラフ詳細設定")
        # 画面右端に配置（コンパクトなサイズ）
        screen_width = self.temp_settings_window.winfo_screenwidth()
        self.temp_settings_window.geometry(f"430x600+{screen_width - 460}+50")
        self.temp_settings_window.resizable(False, False)

        # 選択されたキーを保存
        self.temp_graph_selected_keys = selected_keys

        # Y軸設定用変数（デフォルト選択時用）
        self.temp_yaxis_select_var = tk.StringVar(value="default")

        # NEG絶対値無効オプション
        self.neg_no_abs_var = tk.BooleanVar(value=False)

        # X軸全表示オプション（デフォルト: 25div=250分）
        self.xaxis_full_var = tk.BooleanVar(value=False)

        main_frame = ttk.Frame(self.temp_settings_window, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 計算結果表示エリア（POS/NEG横並び）
        calc_frame = ttk.LabelFrame(main_frame, text="LSB電圧計算結果（実測値）", padding=5)
        calc_frame.pack(fill=tk.X, pady=(0, 8))

        # POS/NEG横並びフレーム
        calc_row = ttk.Frame(calc_frame)
        calc_row.pack(fill=tk.X)

        # POS計算結果（左側・上詰め）
        pos_frame = ttk.Frame(calc_row)
        pos_frame.pack(side=tk.LEFT, padx=(0, 20), anchor=tk.N)
        ttk.Label(pos_frame, text="【POS】", font=('', 9, 'bold')).pack(anchor=tk.W)
        self.calc_pos_fffff_label = ttk.Label(pos_frame, text="+Full: ---")
        self.calc_pos_fffff_label.pack(anchor=tk.W)
        self.calc_pos_zero_label = ttk.Label(pos_frame, text="-Full: ---")
        self.calc_pos_zero_label.pack(anchor=tk.W)
        self.calc_pos_lsb_label = ttk.Label(pos_frame, text="LSB: ---")
        self.calc_pos_lsb_label.pack(anchor=tk.W)

        # NEG計算結果（右側・上詰め）
        neg_frame = ttk.Frame(calc_row)
        neg_frame.pack(side=tk.LEFT, anchor=tk.N)
        ttk.Label(neg_frame, text="【NEG】", font=('', 9, 'bold')).pack(anchor=tk.W)
        self.calc_neg_fffff_label = ttk.Label(neg_frame, text="+Full: ---")
        self.calc_neg_fffff_label.pack(anchor=tk.W)
        self.calc_neg_zero_label = ttk.Label(neg_frame, text="-Full: ---")
        self.calc_neg_zero_label.pack(anchor=tk.W)
        self.calc_neg_lsb_label = ttk.Label(neg_frame, text="LSB: ---")
        self.calc_neg_lsb_label.pack(anchor=tk.W)
        # NEG絶対値無効チェックボックス（NEGの下に配置）
        self.neg_no_abs_check = ttk.Checkbutton(
            neg_frame,
            text="絶対値を使用しない",
            variable=self.neg_no_abs_var,
            command=self._on_neg_no_abs_changed
        )
        self.neg_no_abs_check.pack(anchor=tk.W)

        # Y軸(LSB)設定エリア
        yaxis_frame = ttk.LabelFrame(main_frame, text="Y軸(LSB)設定", padding=5)
        yaxis_frame.pack(fill=tk.X, pady=(0, 8))

        # Y軸範囲（1行目）
        yaxis_row1 = ttk.Frame(yaxis_frame)
        yaxis_row1.pack(fill=tk.X, pady=2)
        ttk.Radiobutton(yaxis_row1, text="デフォルト(±8LSB 2LSB/Div)",
                        variable=self.temp_yaxis_select_var, value="default").pack(side=tk.LEFT)

        # Y軸範囲（2行目：オートとLSB/Div）
        yaxis_row2 = ttk.Frame(yaxis_frame)
        yaxis_row2.pack(fill=tk.X, pady=2)
        ttk.Radiobutton(yaxis_row2, text="オート",
                        variable=self.temp_yaxis_select_var, value="auto").pack(side=tk.LEFT)
        ttk.Label(yaxis_row2, text="LSB/Div:").pack(side=tk.LEFT, padx=(15, 0))
        ttk.Entry(yaxis_row2, textvariable=self.lsb_per_div_var, width=5).pack(side=tk.LEFT, padx=3)
        ttk.Label(yaxis_row2, text="※オート/設定値時有効", foreground="gray").pack(side=tk.LEFT, padx=(5, 0))

        # Y軸範囲（3行目：設定値）
        yaxis_row3 = ttk.Frame(yaxis_frame)
        yaxis_row3.pack(fill=tk.X, pady=2)
        ttk.Radiobutton(yaxis_row3, text="設定値",
                        variable=self.temp_yaxis_select_var, value="manual").pack(side=tk.LEFT)
        ttk.Entry(yaxis_row3, textvariable=self.yaxis_min_var, width=5).pack(side=tk.LEFT, padx=(5, 0))
        ttk.Label(yaxis_row3, text="〜").pack(side=tk.LEFT)
        ttk.Entry(yaxis_row3, textvariable=self.yaxis_max_var, width=5).pack(side=tk.LEFT)

        # Y軸選択変更時に自動更新
        self.temp_yaxis_select_var.trace_add('write', lambda *args: self._redraw_temp_graph_preserve_position())

        # X軸設定エリア
        xaxis_frame = ttk.LabelFrame(main_frame, text="X軸設定", padding=5)
        xaxis_frame.pack(fill=tk.X, pady=(0, 8))
        ttk.Checkbutton(xaxis_frame, text="全表示（デフォルト: 25div=250分）",
                        variable=self.xaxis_full_var,
                        command=self._on_xaxis_full_changed).pack(anchor=tk.W)

        # 温度区間矢印設定エリア
        arrow_frame = ttk.LabelFrame(main_frame, text="温度区間矢印設定", padding=5)
        arrow_frame.pack(fill=tk.X, pady=(0, 8))

        # 矢印表示ON/OFF
        self.show_temp_arrows_var = tk.BooleanVar(value=self._show_temp_arrows)
        ttk.Checkbutton(arrow_frame, text="温度区間矢印を表示",
                        variable=self.show_temp_arrows_var,
                        command=self._on_temp_arrow_setting_changed).pack(anchor=tk.W)

        # Div範囲入力（23℃、28℃、18℃を横並び）
        div_row = ttk.Frame(arrow_frame)
        div_row.pack(fill=tk.X, pady=(3, 0))
        # 23℃
        ttk.Label(div_row, text="23℃:").pack(side=tk.LEFT)
        self.div_23_start_var = tk.StringVar(value=str(self._temp_zone_divs[0][0]))
        self.div_23_end_var = tk.StringVar(value=str(self._temp_zone_divs[0][1]))
        ttk.Entry(div_row, textvariable=self.div_23_start_var, width=3).pack(side=tk.LEFT, padx=1)
        ttk.Label(div_row, text="-").pack(side=tk.LEFT)
        ttk.Entry(div_row, textvariable=self.div_23_end_var, width=3).pack(side=tk.LEFT, padx=(1, 5))
        # 28℃
        ttk.Label(div_row, text="28℃:").pack(side=tk.LEFT)
        self.div_28_start_var = tk.StringVar(value=str(self._temp_zone_divs[1][0]))
        self.div_28_end_var = tk.StringVar(value=str(self._temp_zone_divs[1][1]))
        ttk.Entry(div_row, textvariable=self.div_28_start_var, width=3).pack(side=tk.LEFT, padx=1)
        ttk.Label(div_row, text="-").pack(side=tk.LEFT)
        ttk.Entry(div_row, textvariable=self.div_28_end_var, width=3).pack(side=tk.LEFT, padx=(1, 5))
        # 18℃
        ttk.Label(div_row, text="18℃:").pack(side=tk.LEFT)
        self.div_18_start_var = tk.StringVar(value=str(self._temp_zone_divs[2][0]))
        self.div_18_end_var = tk.StringVar(value=str(self._temp_zone_divs[2][1]))
        ttk.Entry(div_row, textvariable=self.div_18_start_var, width=3).pack(side=tk.LEFT, padx=1)
        ttk.Label(div_row, text="-").pack(side=tk.LEFT)
        ttk.Entry(div_row, textvariable=self.div_18_end_var, width=3).pack(side=tk.LEFT, padx=(1, 5))
        ttk.Button(div_row, text="適用", width=4,
                   command=self._on_temp_arrow_setting_changed).pack(side=tk.LEFT)

        # PNG保存エリア（1行にまとめる）
        save_frame = ttk.LabelFrame(main_frame, text="グラフ保存", padding=5)
        save_frame.pack(fill=tk.X, pady=(0, 8))
        scale_row = ttk.Frame(save_frame)
        scale_row.pack(fill=tk.X)
        ttk.Label(scale_row, text="PNG縮尺:").pack(side=tk.LEFT)
        self.png_scale_var = tk.StringVar(value=self._png_scale_setting)
        ttk.Entry(scale_row, textvariable=self.png_scale_var, width=5).pack(side=tk.LEFT, padx=3)
        ttk.Label(scale_row, text="倍").pack(side=tk.LEFT)
        ttk.Button(scale_row, text="PNG保存",
                   command=self._save_temp_graphs).pack(side=tk.LEFT, padx=(15, 0))

        # データ解析エリア（1行にまとめる）
        analysis_frame = ttk.LabelFrame(main_frame, text="データ解析", padding=5)
        analysis_frame.pack(fill=tk.X, pady=(0, 8))
        ttk.Button(analysis_frame, text="温度係数表示",
                   command=self._show_section_averages).pack(side=tk.LEFT)

        # 閉じるボタン
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(5, 0))
        ttk.Button(btn_frame, text="温特グラフを全て閉じる",
                   command=self._close_all_temp_graphs).pack(side=tk.LEFT)
        ttk.Button(btn_frame, text="閉じる",
                   command=self.temp_settings_window.destroy).pack(side=tk.RIGHT)

        # 初回描画（デフォルト値で）
        self._draw_temp_graph_and_update_calc()

    def _apply_yaxis_to_temp_graph(self):
        """縦軸更新ボタン: 共通設定の値でグラフを再描画（位置保持）"""
        import matplotlib.pyplot as plt

        # 既存のグラフウィンドウの位置を保存
        graph_positions = []
        for fig_num in plt.get_fignums():
            try:
                fig = plt.figure(fig_num)
                manager = fig.canvas.manager
                x = manager.window.winfo_x()
                y = manager.window.winfo_y()
                graph_positions.append((x, y))
            except:
                pass

        plt.close('all')
        self._draw_temp_graph_and_update_calc()

        # 新しいグラフウィンドウを保存した位置に移動
        if graph_positions:
            for i, fig_num in enumerate(plt.get_fignums()):
                if i < len(graph_positions):
                    try:
                        fig = plt.figure(fig_num)
                        manager = fig.canvas.manager
                        x, y = graph_positions[i]
                        manager.window.geometry(f"+{x}+{y}")
                    except:
                        pass

    def _draw_temp_graph_and_update_calc(self):
        """温特グラフを描画し、計算結果を更新"""
        try:
            bit_precision = int(self.bit_precision_var.get())
            pos_full = float(self.pos_full_var.get())
            neg_full = float(self.neg_full_var.get())
            ref_mode = "first_avg"  # 温特グラフは初回平均固定

            # Y軸設定を選択に基づいて決定
            yaxis_select = self.temp_yaxis_select_var.get()
            if yaxis_select == "default":
                # デフォルト: ±8LSB, 2LSB/Div
                lsb_per_div = 2
                yaxis_mode = "manual"
                yaxis_min = -8.0
                yaxis_max = 8.0
            elif yaxis_select == "auto":
                # オート
                lsb_per_div = float(self.lsb_per_div_var.get())
                yaxis_mode = "auto"
                yaxis_min = None
                yaxis_max = None
            else:
                # 設定値
                lsb_per_div = float(self.lsb_per_div_var.get())
                yaxis_mode = "manual"
                yaxis_min = float(self.yaxis_min_var.get())
                yaxis_max = float(self.yaxis_max_var.get())

            skip_after_change = int(self.skip_after_change_var.get())
            skip_first_data = self.skip_first_data_var.get()
            skip_before_change = self.skip_before_change_var.get()
        except ValueError:
            messagebox.showerror("エラー", "設定値が不正です")
            return

        # LSBGraphPlotterを作成
        plotter = LSBGraphPlotter(bit_precision, pos_full, neg_full, lsb_per_div, ref_mode,
                                  yaxis_mode, yaxis_min, yaxis_max,
                                  skip_after_change, skip_first_data, skip_before_change)

        # 選択されたデータをプロット（全グラフの計算結果とfigureを保存）
        # 辞書形式: {serial: calc_info}
        if not hasattr(self, 'temp_graph_all_info'):
            self.temp_graph_all_info = {}
        self.temp_graph_all_info = {}  # 毎回リセット

        # 互換性のため旧変数も維持
        self.temp_graph_pos_info = None
        self.temp_graph_neg_info = None
        self.temp_graph_pos_serial = None
        self.temp_graph_neg_serial = None

        # NEG絶対値無効オプションを取得
        neg_no_abs = self.neg_no_abs_var.get()
        # X軸全表示オプションを取得
        xaxis_full = self.xaxis_full_var.get()

        # 温度区間矢印設定を取得（設定画面から）
        temp_zone_divs = getattr(self, '_temp_zone_divs', [(0, 6), (6, 13), (13, 19)])
        show_temp_arrows = getattr(self, '_show_temp_arrows', True)

        for key in self.temp_graph_selected_keys:
            serial, pole = key.rsplit('_', 1)
            # NEGの場合のみneg_no_absを適用
            use_no_abs = neg_no_abs if pole == "NEG" else False
            calc_info = plotter.plot_temperature_characteristic(
                self.csv_data, self.temp_csv_data, serial, pole,
                yaxis_mode, yaxis_min, yaxis_max, use_no_abs, xaxis_full,
                temp_zone_divs=temp_zone_divs, show_temp_arrows=show_temp_arrows
            )
            if calc_info:
                # 全グラフ情報を保存
                self.temp_graph_all_info[f"{serial}_{pole}"] = calc_info

                # 互換性のため旧変数も更新
                if pole == "POS":
                    self.temp_graph_pos_info = calc_info
                    self.temp_graph_pos_serial = serial
                elif pole == "NEG":
                    self.temp_graph_neg_info = calc_info
                    self.temp_graph_neg_serial = serial

        pos_calc_info = self.temp_graph_pos_info
        neg_calc_info = self.temp_graph_neg_info

        # エラーチェック：必要なデータ（FFFFF/00000）があるか確認
        missing_data = []
        for calc_info, pole_name in [(pos_calc_info, "POS"), (neg_calc_info, "NEG")]:
            if calc_info:
                if not calc_info.get('has_fffff'):
                    missing_data.append(f"{pole_name}: +Full(FFFFF)データなし")
                if not calc_info.get('has_zero'):
                    missing_data.append(f"{pole_name}: -Full(00000)データなし")

        if missing_data:
            messagebox.showwarning("警告",
                "LSB電圧計算に必要なデータが不足しています:\n" + "\n".join(missing_data))

        # 計算結果を更新（POS/NEG別）
        self._update_calc_labels(pos_calc_info, neg_calc_info)

    def _update_calc_labels(self, pos_calc_info, neg_calc_info):
        """計算結果ラベルを更新（POS/NEG別、表示は小数第5位まで）"""
        # POS計算結果
        if pos_calc_info:
            fffff_avg = pos_calc_info.get('fffff_avg')
            zero_avg = pos_calc_info.get('00000_avg')
            lsb_voltage = pos_calc_info.get('lsb_voltage')

            if fffff_avg is not None:
                self.calc_pos_fffff_label.config(text=f"  +Full平均: {fffff_avg:.5f} V")
            else:
                self.calc_pos_fffff_label.config(text="  +Full平均: データなし")

            if zero_avg is not None:
                self.calc_pos_zero_label.config(text=f"  -Full平均: {zero_avg:.5f} V")
            else:
                self.calc_pos_zero_label.config(text="  -Full平均: データなし")

            if lsb_voltage is not None:
                lsb_mv = lsb_voltage * 1000  # V → mV
                self.calc_pos_lsb_label.config(text=f"  LSB電圧: {lsb_mv:.5f} mV")
            else:
                self.calc_pos_lsb_label.config(text="  LSB電圧: 計算不可")
        else:
            self.calc_pos_fffff_label.config(text="  +Full平均: ---")
            self.calc_pos_zero_label.config(text="  -Full平均: ---")
            self.calc_pos_lsb_label.config(text="  LSB電圧: ---")

        # NEG計算結果
        if neg_calc_info:
            fffff_avg = neg_calc_info.get('fffff_avg')
            zero_avg = neg_calc_info.get('00000_avg')
            lsb_voltage = neg_calc_info.get('lsb_voltage')

            if fffff_avg is not None:
                self.calc_neg_fffff_label.config(text=f"  +Full平均: {fffff_avg:.5f} V")
            else:
                self.calc_neg_fffff_label.config(text="  +Full平均: データなし")

            if zero_avg is not None:
                self.calc_neg_zero_label.config(text=f"  -Full平均: {zero_avg:.5f} V")
            else:
                self.calc_neg_zero_label.config(text="  -Full平均: データなし")

            if lsb_voltage is not None:
                lsb_mv = lsb_voltage * 1000  # V → mV
                self.calc_neg_lsb_label.config(text=f"  LSB電圧: {lsb_mv:.5f} mV")
            else:
                self.calc_neg_lsb_label.config(text="  LSB電圧: 計算不可")
        else:
            self.calc_neg_fffff_label.config(text="  +Full平均: ---")
            self.calc_neg_zero_label.config(text="  -Full平均: ---")
            self.calc_neg_lsb_label.config(text="  LSB電圧: ---")

    def _redraw_temp_graph(self):
        """温特グラフを再描画"""
        import matplotlib.pyplot as plt
        # 既存のグラフを閉じる
        plt.close('all')
        # 再描画
        self._draw_temp_graph_and_update_calc()

    def _on_neg_no_abs_changed(self):
        """NEG絶対値無効チェックボックスの変更時: グラフを再描画（位置保持）"""
        self._redraw_temp_graph_preserve_position()

    def _on_xaxis_full_changed(self):
        """X軸全表示チェックボックスの変更時: グラフを再描画（位置保持）"""
        self._redraw_temp_graph_preserve_position()

    def _on_temp_arrow_setting_changed(self):
        """温度区間矢印設定の変更時: グラフを再描画（位置保持）"""
        # 設定を保存
        self._show_temp_arrows = self.show_temp_arrows_var.get()
        try:
            self._temp_zone_divs = [
                (int(self.div_23_start_var.get()), int(self.div_23_end_var.get())),
                (int(self.div_28_start_var.get()), int(self.div_28_end_var.get())),
                (int(self.div_18_start_var.get()), int(self.div_18_end_var.get()))
            ]
        except ValueError:
            pass  # 無効な値の場合は前回値を維持
        self.save_settings()
        self._redraw_temp_graph_preserve_position()

    def _redraw_temp_graph_preserve_position(self):
        """温特グラフを位置を保持して再描画"""
        import matplotlib.pyplot as plt

        # 既存のグラフウィンドウの位置を保存
        graph_positions = []
        for fig_num in plt.get_fignums():
            try:
                fig = plt.figure(fig_num)
                manager = fig.canvas.manager
                x = manager.window.winfo_x()
                y = manager.window.winfo_y()
                graph_positions.append((x, y))
            except:
                pass

        plt.close('all')
        self._draw_temp_graph_and_update_calc()

        # 新しいグラフウィンドウを保存した位置に移動
        if graph_positions:
            for i, fig_num in enumerate(plt.get_fignums()):
                if i < len(graph_positions):
                    try:
                        fig = plt.figure(fig_num)
                        manager = fig.canvas.manager
                        x, y = graph_positions[i]
                        manager.window.geometry(f"+{x}+{y}")
                    except:
                        pass

    def _save_temp_graphs(self):
        """温特グラフをPNG保存（表示中の全グラフ）"""
        import os

        # 保存するグラフがあるか確認（全グラフ情報から取得）
        graphs_to_save = []
        if hasattr(self, 'temp_graph_all_info') and self.temp_graph_all_info:
            for key, calc_info in self.temp_graph_all_info.items():
                if calc_info and calc_info.get('figure'):
                    serial, pole = key.rsplit('_', 1)
                    graphs_to_save.append((serial, pole, calc_info['figure']))

        if not graphs_to_save:
            messagebox.showerror("エラー", "保存するグラフがありません")
            return

        # 縮尺を取得・検証（0.5～3.0）
        try:
            scale = float(self.png_scale_var.get())
            if scale < 0.5:
                scale = 0.5
            elif scale > 3.0:
                scale = 3.0
            # 小数第1位に丸める
            scale = round(scale, 1)
            self.png_scale_var.set(str(scale))
        except ValueError:
            scale = 1.0
            self.png_scale_var.set("1.0")

        # 設定を保存
        self._png_scale_setting = self.png_scale_var.get()
        self.save_settings()

        # フォルダ選択ダイアログ
        folder_path = filedialog.askdirectory(title="保存先フォルダを選択")
        if not folder_path:
            return

        # 保存実行（dpi = 150 × 縮尺）
        base_dpi = 150
        dpi = int(base_dpi * scale)
        saved_files = []
        errors = []
        for serial, pole, fig in graphs_to_save:
            base_filename = f"温特グラフ_{serial}_{pole}"
            filename = self._get_unique_png_filename(folder_path, base_filename)
            filepath = os.path.join(folder_path, filename)
            try:
                fig.savefig(filepath, dpi=dpi, bbox_inches='tight')
                saved_files.append(filename)
            except Exception as e:
                errors.append(f"{filename}: {str(e)}")

        # 結果表示
        if saved_files:
            msg = "保存しました:\n" + "\n".join(saved_files)
            if errors:
                msg += "\n\nエラー:\n" + "\n".join(errors)
                messagebox.showwarning("一部保存完了", msg)
            else:
                messagebox.showinfo("成功", msg)
        else:
            messagebox.showerror("エラー", "保存に失敗しました:\n" + "\n".join(errors))

    def _close_all_temp_graphs(self):
        """表示している温特グラフを全て閉じる"""
        import matplotlib.pyplot as plt

        graphs_closed = False

        if hasattr(self, 'temp_graph_all_info') and self.temp_graph_all_info:
            for calc_info in self.temp_graph_all_info.values():
                fig = calc_info.get('figure')
                if fig:
                    try:
                        plt.close(fig)
                        graphs_closed = True
                    except Exception:
                        pass
                    finally:
                        calc_info['figure'] = None

        if graphs_closed:
            messagebox.showinfo("完了", "表示中の温特グラフをすべて閉じました。")
        else:
            messagebox.showinfo("情報", "表示中の温特グラフはありません。")

    def _get_unique_png_filename(self, folder_path, base_filename):
        """同名ファイルが存在する場合、連番を付けた別名を返す"""
        filename = f"{base_filename}.png"
        filepath = os.path.join(folder_path, filename)

        if not os.path.exists(filepath):
            return filename

        # 連番を付けて重複しないファイル名を探す
        counter = 1
        while True:
            filename = f"{base_filename}_{counter}.png"
            filepath = os.path.join(folder_path, filename)
            if not os.path.exists(filepath):
                return filename
            counter += 1

    def _show_section_averages(self):
        """区間別平均電圧を表示するウィンドウを開く"""
        if not self.csv_data:
            messagebox.showerror("エラー", "測定CSVファイルを読み込んでください")
            return

        # 選択されたデータを確認
        selected_keys = [key for key, var in self.checkboxes.items() if var.get()]
        if not selected_keys:
            messagebox.showwarning("警告", "表示するデータを選択してください")
            return

        # 設定値を取得
        try:
            bit_precision = int(self.bit_precision_var.get())
            pos_full = float(self.pos_full_var.get())
            neg_full = float(self.neg_full_var.get())
            skip_after_change = int(self.skip_after_change_var.get())
            skip_first_data = self.skip_first_data_var.get()
            skip_before_change = self.skip_before_change_var.get()
        except ValueError:
            messagebox.showerror("エラー", "設定値が不正です")
            return

        # LSBGraphPlotterを作成
        plotter = LSBGraphPlotter(bit_precision, pos_full, neg_full, 2, "first_avg",
                                  "auto", None, None,
                                  skip_after_change, skip_first_data, skip_before_change)

        # 選択されたデータをシリアル番号ごとにグループ化
        serial_data = {}  # {serial: {'POS': sections, 'NEG': sections}}

        for key in selected_keys:
            serial, pole = key.rsplit('_', 1)
            column_name = f"{serial}_{pole}"
            sections = plotter.extract_section_averages(
                self.csv_data, serial, pole, column_name, last_minutes=10
            )
            if serial not in serial_data:
                serial_data[serial] = {'POS': None, 'NEG': None}
            serial_data[serial][pole] = sections

        # 結果表示ウィンドウを作成（タブ形式）
        self._create_section_averages_window_tabbed(serial_data)

    def _create_section_averages_window_tabbed(self, serial_data):
        """温度係数表示ウィンドウを作成（タブ形式）"""
        # 既存のウィンドウがあれば閉じる
        if hasattr(self, 'section_avg_window') and self.section_avg_window:
            try:
                if self.section_avg_window.winfo_exists():
                    self.section_avg_window.destroy()
            except:
                pass

        # ウィンドウ作成
        self.section_avg_window = tk.Toplevel(self)
        self.section_avg_window.title("温度係数")
        self.section_avg_window.geometry("750x750")
        self.section_avg_window.resizable(True, True)

        main_frame = ttk.Frame(self.section_avg_window, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # スペック値（内部で使用）
        self.temp_coef_spec_var = tk.StringVar(value="1.9")

        # タブコントロール作成
        self.temp_coef_notebook = ttk.Notebook(main_frame)
        self.temp_coef_notebook.pack(fill=tk.BOTH, expand=True)
        notebook = self.temp_coef_notebook

        # 各シリアル番号のデータを保存（PNG保存用）
        self.temp_coef_serial_data = serial_data
        self.temp_coef_table_frames = {}

        # シリアル番号ごとにタブを作成
        for serial in sorted(serial_data.keys()):
            data = serial_data[serial]
            pos_sections = data.get('POS')
            neg_sections = data.get('NEG')

            # タブフレーム作成
            tab_frame = ttk.Frame(notebook, padding=5)
            notebook.add(tab_frame, text=serial)

            # テーブル表示エリア（スクロール対応）
            table_container = ttk.Frame(tab_frame)
            table_container.pack(fill=tk.BOTH, expand=True)

            # キャンバスとスクロールバー
            canvas = tk.Canvas(table_container, bg='white')
            v_scrollbar = ttk.Scrollbar(table_container, orient=tk.VERTICAL, command=canvas.yview)
            h_scrollbar = ttk.Scrollbar(table_container, orient=tk.HORIZONTAL, command=canvas.xview)
            table_frame = ttk.Frame(canvas)

            table_frame.bind(
                "<Configure>",
                lambda e, c=canvas: c.configure(scrollregion=c.bbox("all"))
            )

            canvas.create_window((0, 0), window=table_frame, anchor="nw")
            canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

            v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

            # テーブル作成
            self._create_temp_coef_table_in_frame(table_frame, pos_sections, neg_sections, serial)

            # PNG保存用にフレーム、キャンバス、スクロールバーを保存
            self.temp_coef_table_frames[serial] = {
                'frame': table_frame,
                'canvas': canvas,
                'v_scrollbar': v_scrollbar,
                'h_scrollbar': h_scrollbar
            }

        # ボタンフレーム
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        ttk.Button(btn_frame, text="Excel&PNG保存",
                   command=self._save_temp_coef_tables_png).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="閉じる",
                   command=self.section_avg_window.destroy).pack(side=tk.RIGHT)

    def _create_section_averages_window(self, pos_sections, neg_sections, pos_serial, neg_serial):
        """温度係数表示ウィンドウを作成（後方互換性のため残す）"""
        serial_data = {}
        if pos_serial:
            if pos_serial not in serial_data:
                serial_data[pos_serial] = {'POS': None, 'NEG': None}
            serial_data[pos_serial]['POS'] = pos_sections
        if neg_serial:
            if neg_serial not in serial_data:
                serial_data[neg_serial] = {'POS': None, 'NEG': None}
            serial_data[neg_serial]['NEG'] = neg_sections
        self._create_section_averages_window_tabbed(serial_data)

    def _save_temp_coef_tables_png(self):
        """温度係数テーブルをExcel&PNG保存（複数シリアルは1ファイルに複数シート）"""
        import os

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
            import win32com.client
            from PIL import Image
            import io
        except ImportError as e:
            messagebox.showerror("エラー",
                f"必要なライブラリがありません: {e}\n\n"
                "pip install openpyxl pywin32 pillow")
            return

        if not hasattr(self, 'temp_coef_serial_data') or not self.temp_coef_serial_data:
            messagebox.showerror("エラー", "保存するデータがありません")
            return

        # フォルダ選択ダイアログ
        folder_path = filedialog.askdirectory(title="保存先フォルダを選択")
        if not folder_path:
            return

        # スペック値を取得
        try:
            spec_value = float(self.temp_coef_spec_var.get())
        except ValueError:
            spec_value = 1.9

        saved_files = []
        errors = []
        sheet_info = {}  # {serial: (last_row, last_col)}

        # 1つのワークブックに全シリアルのシートを作成
        wb = Workbook()
        wb.remove(wb.active)  # デフォルトシートを削除

        for serial in sorted(self.temp_coef_serial_data.keys()):
            try:
                data = self.temp_coef_serial_data[serial]
                pos_sections = data.get('POS')
                neg_sections = data.get('NEG')

                # 新しいシートを作成
                ws = wb.create_sheet(title=serial)

                # テーブルデータをExcelに書き込み
                last_row, last_col = self._write_excel_table(
                    ws, pos_sections, neg_sections, serial, spec_value)
                sheet_info[serial] = (last_row, last_col)

            except Exception as e:
                errors.append(f"{serial}: {str(e)}")

        # Excelファイルを保存（シリアルNo.をファイル名に含める）
        serial_names = "_".join(sorted(self.temp_coef_serial_data.keys()))
        excel_base = f"温度係数_{serial_names}"
        excel_filename = f"{excel_base}.xlsx"
        excel_path = os.path.join(folder_path, excel_filename)

        idx = 1
        while os.path.exists(excel_path):
            excel_filename = f"{excel_base}_{idx}.xlsx"
            excel_path = os.path.join(folder_path, excel_filename)
            idx += 1

        wb.save(excel_path)
        wb.close()
        saved_files.append(excel_filename)

        # 各シートをPNGとして保存
        for serial, (last_row, last_col) in sheet_info.items():
            try:
                base_filename = f"温度係数_{serial}"
                filename = f"{base_filename}.png"
                filepath = os.path.join(folder_path, filename)

                index = 1
                while os.path.exists(filepath):
                    filename = f"{base_filename}_{index}.png"
                    filepath = os.path.join(folder_path, filename)
                    index += 1

                # Excel COMでセル範囲を画像として保存
                self._excel_range_to_png(excel_path, serial, f"A1:{get_column_letter(last_col)}{last_row}", filepath)
                saved_files.append(filename)

            except Exception as e:
                errors.append(f"{serial} PNG: {str(e)}")

        # 結果表示
        if saved_files:
            msg = "保存しました:\n" + "\n".join(saved_files)
            if errors:
                msg += "\n\nエラー:\n" + "\n".join(errors)
                messagebox.showwarning("一部保存完了", msg)
            else:
                messagebox.showinfo("成功", msg)
        else:
            messagebox.showerror("エラー", "保存に失敗しました:\n" + "\n".join(errors))

    def _excel_range_to_png(self, excel_path, sheet_name, cell_range, output_path):
        """ExcelのセルをPNG画像として保存"""
        import win32com.client
        import pythoncom
        from PIL import Image
        import io
        import time

        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            # DispatchExで新しいExcelインスタンスを作成（既存のExcelに影響しない）
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Open(excel_path)
            ws = wb.Worksheets(sheet_name)  # 指定シートを選択

            # セル範囲を選択してコピー
            rng = ws.Range(cell_range)
            rng.CopyPicture(Format=2)  # xlBitmap = 2

            # クリップボードから画像を取得
            time.sleep(0.3)

            from PIL import ImageGrab
            img = ImageGrab.grabclipboard()

            if img:
                img.save(output_path, 'PNG')
            else:
                raise Exception("クリップボードから画像を取得できませんでした")

        finally:
            if wb:
                wb.Close(SaveChanges=False)
            if excel:
                excel.Quit()
            pythoncom.CoUninitialize()

    def _write_excel_table(self, ws, pos_sections, neg_sections, serial, spec_value):
        """Excelシートにテーブルを書き込み"""
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        # スタイル定義
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # 斜線ボーダー（左上から右下）
        diagonal_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            diagonal=Side(style='thin'),
            diagonalDown=True
        )
        header_fill = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid')
        ng_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')

        # ヘッダー
        headers = ['ユニットNo.', '入力コード', '温度\n(℃)', '測定電圧\n(V)', 'ΔT\n(℃)',
                   'ΔV\n(V)', '温度係数\n(ppm/℃)', 'スペック\n(ppm/℃)', '判定']
        col_widths = [14, 10, 6, 12, 6, 12, 11, 10, 6]

        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=9)
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
            ws.column_dimensions[get_column_letter(col)].width = width

        # データ行
        row_idx = 2
        total_data_rows = 0

        for sections, pole in [(pos_sections, "POS"), (neg_sections, "NEG")]:
            if not sections:
                continue

            code_data = self._organize_sections_by_code(sections)
            unit_name = f"1PB397MK2\nDFH_{serial}\n{pole}"
            unit_start_row = row_idx

            # フルスケール計算
            fffff_23 = code_data.get('FFFFF', {}).get(23, {}).get('voltage')
            zero_23 = code_data.get('00000', {}).get(23, {}).get('voltage')
            full_scale_23 = (fffff_23 - zero_23) if (fffff_23 is not None and zero_23 is not None) else None

            for code in ['FFFFF', '80000', '00000']:
                if code not in code_data:
                    continue

                temps_data = code_data[code]
                code_start_row = row_idx

                # 温度係数計算
                tc_28 = dv_28 = tc_18 = dv_18 = None
                if 28 in temps_data and 23 in temps_data:
                    dv_28 = temps_data[28]['voltage'] - temps_data[23]['voltage']
                    if full_scale_23:
                        tc_28 = (dv_28 / full_scale_23) * 1e6 / 5.0

                if 23 in temps_data and 18 in temps_data:
                    dv_18 = temps_data[18]['voltage'] - temps_data[23]['voltage']
                    if full_scale_23:
                        tc_18 = (dv_18 / full_scale_23) * 1e6 / (-5.0)

                jdg_28 = "OK" if tc_28 is not None and abs(tc_28) <= spec_value else ("NG" if tc_28 is not None else "")
                jdg_18 = "OK" if tc_18 is not None and abs(tc_18) <= spec_value else ("NG" if tc_18 is not None else "")

                # 6行作成（各温度2行）
                for temp_idx, temp in enumerate([28, 23, 18]):
                    voltage = temps_data.get(temp, {}).get('voltage')
                    temp_start_row = row_idx

                    for sub_row in range(2):
                        # 温度セル（2行結合、最初のみ）
                        if sub_row == 0:
                            ws.cell(row=row_idx, column=3, value=temp).alignment = center_align
                            ws.cell(row=row_idx, column=3).border = thin_border

                        # 測定電圧（2行結合、最初のみ）
                        if sub_row == 0 and voltage is not None:
                            cell = ws.cell(row=row_idx, column=4, value=voltage)
                            cell.alignment = right_align
                            cell.number_format = '0.000000'
                            cell.border = thin_border

                        # ΔT, ΔV, 温度係数, 判定
                        if temp_idx == 0 and sub_row == 1:  # 28→23
                            c = ws.cell(row=row_idx, column=5, value=5.0)
                            c.alignment = center_align
                            c.number_format = '0.0'
                            if dv_28 is not None:
                                c = ws.cell(row=row_idx, column=6, value=dv_28)
                                c.alignment = right_align
                                c.number_format = '0.000000'
                            if tc_28 is not None:
                                c = ws.cell(row=row_idx, column=7, value=tc_28)
                                c.alignment = center_align
                                c.number_format = '0.0'
                                if jdg_28 == "NG":
                                    c.fill = ng_fill
                            c = ws.cell(row=row_idx, column=9, value=jdg_28)
                            c.alignment = center_align
                            if jdg_28 == "NG":
                                c.fill = ng_fill
                        elif temp_idx == 1 and sub_row == 1:  # 23→18
                            c = ws.cell(row=row_idx, column=5, value=-5.0)
                            c.alignment = center_align
                            c.number_format = '0.0'
                            if dv_18 is not None:
                                c = ws.cell(row=row_idx, column=6, value=dv_18)
                                c.alignment = right_align
                                c.number_format = '0.000000'
                            if tc_18 is not None:
                                c = ws.cell(row=row_idx, column=7, value=tc_18)
                                c.alignment = center_align
                                c.number_format = '0.0'
                                if jdg_18 == "NG":
                                    c.fill = ng_fill
                            c = ws.cell(row=row_idx, column=9, value=jdg_18)
                            c.alignment = center_align
                            if jdg_18 == "NG":
                                c.fill = ng_fill
                        else:
                            # 空セルに斜線（左上から右下への斜線）
                            for col in [5, 6, 7, 9]:
                                cell = ws.cell(row=row_idx, column=col)
                                cell.border = diagonal_border

                        # 全セルにボーダー適用（斜線セル以外）
                        diagonal_cols = [5, 6, 7, 9] if not (temp_idx == 0 and sub_row == 1) and not (temp_idx == 1 and sub_row == 1) else []
                        for col in range(1, 10):
                            if col not in diagonal_cols:
                                ws.cell(row=row_idx, column=col).border = thin_border

                        row_idx += 1
                        total_data_rows += 1

                    # 温度・電圧セルの結合
                    if temp_start_row < row_idx - 1:
                        ws.merge_cells(start_row=temp_start_row, start_column=3,
                                      end_row=row_idx - 1, end_column=3)
                        ws.merge_cells(start_row=temp_start_row, start_column=4,
                                      end_row=row_idx - 1, end_column=4)

                # 入力コードセル結合（6行）
                ws.cell(row=code_start_row, column=2, value=f"{code} H").alignment = center_align
                if code_start_row < row_idx - 1:
                    ws.merge_cells(start_row=code_start_row, start_column=2,
                                  end_row=row_idx - 1, end_column=2)

                # ΔT/ΔV/温度係数/判定の結合（28→23: row 2-3, 23→18: row 4-5）
                for merge_row, cols in [
                    (code_start_row + 1, [5, 6, 7, 9]),  # 28→23
                    (code_start_row + 3, [5, 6, 7, 9])   # 23→18
                ]:
                    for col in cols:
                        ws.merge_cells(start_row=merge_row, start_column=col,
                                      end_row=merge_row + 1, end_column=col)

            # ユニットNo.セル結合（18行）
            ws.cell(row=unit_start_row, column=1, value=unit_name).alignment = center_align
            if unit_start_row < row_idx - 1:
                ws.merge_cells(start_row=unit_start_row, start_column=1,
                              end_row=row_idx - 1, end_column=1)

        # スペックセル結合（全データ行）
        if total_data_rows > 0:
            c = ws.cell(row=2, column=8, value=spec_value)
            c.alignment = center_align
            c.number_format = '0.0'
            ws.merge_cells(start_row=2, start_column=8, end_row=row_idx - 1, end_column=8)

        # 行の高さ設定
        ws.row_dimensions[1].height = 28  # ヘッダー行
        for r in range(2, row_idx):
            ws.row_dimensions[r].height = 15

        # 大外枠を太字に設定
        last_row = row_idx - 1
        last_col = 9
        thick_side = Side(style='medium')
        thin_side = Side(style='thin')

        for row in range(1, last_row + 1):
            for col in range(1, last_col + 1):
                cell = ws.cell(row=row, column=col)
                # 既存ボーダーを取得（斜線を保持）
                existing = cell.border
                new_left = thick_side if col == 1 else (existing.left or thin_side)
                new_right = thick_side if col == last_col else (existing.right or thin_side)
                new_top = thick_side if row == 1 else (existing.top or thin_side)
                new_bottom = thick_side if row == last_row else (existing.bottom or thin_side)

                cell.border = Border(
                    left=new_left, right=new_right,
                    top=new_top, bottom=new_bottom,
                    diagonal=existing.diagonal,
                    diagonalDown=existing.diagonalDown,
                    diagonalUp=existing.diagonalUp
                )

        return row_idx - 1, 9  # 最終行、最終列

    def _generate_html_table(self, pos_sections, neg_sections, serial, spec_value):
        """PNG保存用のHTMLテーブルを生成"""

        # CSSスタイル
        css = """
        <style>
            body { font-family: 'MS Gothic', 'Yu Gothic', 'Meiryo', sans-serif; margin: 5px; background: white; }
            table { border-collapse: collapse; font-size: 10px; }
            th, td { border: 1px solid #888; padding: 1px 3px; text-align: center; vertical-align: middle; }
            th { background-color: #D0D0D0; font-weight: bold; white-space: pre-line; }
            td { background-color: white; }
            .diagonal { background: linear-gradient(to bottom right, white 49%, #888 50%, white 51%); }
            .ng { background-color: #FFCCCC !important; }
            .right { text-align: right; }
            .unit-cell { white-space: pre-line; line-height: 1.2; }
        </style>
        """

        # 全データを収集
        all_rows = []  # [(unit_name, code, temp, voltage, dt, dv, tc, spec, judgment, tc_bg, jdg_bg), ...]

        for sections, pole in [(pos_sections, "POS"), (neg_sections, "NEG")]:
            if not sections:
                continue

            code_data = self._organize_sections_by_code(sections)
            unit_name = f"1PB397MK2\nDFH_{serial}\n{pole}"

            # 23℃のフルスケール電圧範囲
            fffff_23 = code_data.get('FFFFF', {}).get(23, {}).get('voltage')
            zero_23 = code_data.get('00000', {}).get(23, {}).get('voltage')
            full_scale_23 = (fffff_23 - zero_23) if (fffff_23 is not None and zero_23 is not None) else None

            for code in ['FFFFF', '80000', '00000']:
                if code not in code_data:
                    continue

                temps_data = code_data[code]

                # 温度係数計算
                tc_28 = dv_28 = tc_18 = dv_18 = None
                if 28 in temps_data and 23 in temps_data:
                    dv_28 = temps_data[28]['voltage'] - temps_data[23]['voltage']
                    if full_scale_23:
                        tc_28 = (dv_28 / full_scale_23) * 1e6 / 5.0

                if 23 in temps_data and 18 in temps_data:
                    dv_18 = temps_data[18]['voltage'] - temps_data[23]['voltage']
                    if full_scale_23:
                        tc_18 = (dv_18 / full_scale_23) * 1e6 / (-5.0)

                jdg_28 = "OK" if tc_28 is not None and abs(tc_28) <= spec_value else ("NG" if tc_28 is not None else "")
                jdg_18 = "OK" if tc_18 is not None and abs(tc_18) <= spec_value else ("NG" if tc_18 is not None else "")
                ng_28 = jdg_28 == "NG"
                ng_18 = jdg_18 == "NG"

                # 6行生成（各温度2行）
                for temp in [28, 23, 18]:
                    v = temps_data.get(temp, {}).get('voltage')
                    v_str = f"{v:.6f}" if v is not None else ""

                    if temp == 28:
                        # 行1: 斜線
                        all_rows.append((unit_name, code, temp, v_str, "／", "／", "／", spec_value, "／", False, False))
                        # 行2: 28→23計算結果
                        all_rows.append((None, None, None, None,
                                        "5.0", f"{dv_28:.6f}" if dv_28 else "", f"{tc_28:.1f}" if tc_28 else "",
                                        None, jdg_28, ng_28, ng_28))
                    elif temp == 23:
                        # 行3: 23℃データ (28→23の計算結果と同じ行に含まれる想定だが、2行構造なので分離)
                        all_rows.append((None, None, temp, v_str, None, None, None, None, None, False, False))
                        # 行4: 23→18計算結果
                        all_rows.append((None, None, None, None,
                                        "-5.0", f"{dv_18:.6f}" if dv_18 else "", f"{tc_18:.1f}" if tc_18 else "",
                                        None, jdg_18, ng_18, ng_18))
                    else:  # 18
                        # 行5: 18℃データ
                        all_rows.append((None, None, temp, v_str, None, None, None, None, None, False, False))
                        # 行6: 斜線
                        all_rows.append((None, None, None, None, "／", "／", "／", None, "／", False, False))

        # HTMLテーブル生成（シンプルな行ごと出力）
        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">{css}</head><body><table>
        <tr><th>ユニットNo.</th><th>入力<br>コード</th><th>温度<br>(℃)</th><th>測定電圧<br>(V)</th>
        <th>ΔT<br>(℃)</th><th>ΔV<br>(V)</th><th>温度係数<br>(ppm/℃)</th><th>スペック<br>(ppm/℃)</th><th>判定</th></tr>"""

        row_idx = 0
        current_unit = None
        current_code = None
        unit_start = 0
        code_start = 0

        # まずシンプルに全行出力
        for i, row in enumerate(all_rows):
            unit, code, temp, volt, dt, dv, tc, spec, jdg, tc_ng, jdg_ng = row

            html += "<tr>"

            # ユニットNo.
            if unit is not None:
                # 18行結合（POS/NEG各18行）
                rowspan = 18
                html += f'<td rowspan="{rowspan}" class="unit-cell">{unit.replace(chr(10), "<br>")}</td>'

            # 入力コード
            if code is not None:
                html += f'<td rowspan="6">{code} H</td>'

            # 温度
            if temp is not None:
                html += f'<td rowspan="2">{temp}</td>'

            # 測定電圧
            if volt is not None:
                html += f'<td rowspan="2" class="right">{volt}</td>'

            # ΔT
            if dt == "／":
                html += '<td class="diagonal"></td>'
            elif dt is not None:
                html += f'<td rowspan="2">{dt}</td>'

            # ΔV
            if dv == "／":
                html += '<td class="diagonal"></td>'
            elif dv is not None:
                html += f'<td rowspan="2" class="right">{dv}</td>'

            # 温度係数
            if tc == "／":
                html += '<td class="diagonal"></td>'
            elif tc is not None:
                cls = ' class="ng"' if tc_ng else ''
                html += f'<td rowspan="2"{cls}>{tc}</td>'

            # スペック
            if spec == "／":
                html += '<td class="diagonal"></td>'
            elif spec is not None:
                html += f'<td rowspan="18">{spec:.1f}</td>'

            # 判定
            if jdg == "／":
                html += '<td class="diagonal"></td>'
            elif jdg is not None:
                cls = ' class="ng"' if jdg_ng else ''
                html += f'<td rowspan="2"{cls}>{jdg}</td>'

            html += "</tr>"

        html += "</table></body></html>"
        return html

    def _create_temp_coef_table(self, pos_sections, neg_sections, pos_serial, neg_serial):
        """温度係数テーブルを作成（添付画像形式）"""
        parent = self.temp_coef_table_frame

        # スペック値を取得
        try:
            spec_value = float(self.temp_coef_spec_var.get())
        except ValueError:
            spec_value = 1.9

        # ヘッダー行
        headers = ['ユニットNo.', '入力コード', '温度\n(℃)', '測定電圧\n(V)', 'ΔT\n(℃)', 'ΔV\n(V)',
                   '温度係数\n(ppm/℃)', 'スペック\n(ppm/℃)', '判定']
        header_widths = [10, 7, 4, 10, 4, 10, 9, 7, 4]

        for col, (header, width) in enumerate(zip(headers, header_widths)):
            label = tk.Label(parent, text=header, relief=tk.RIDGE, width=width,
                           bg='#D0D0D0', font=('', 9))
            label.grid(row=0, column=col, sticky='nsew', padx=0, pady=0)

        # 斜線セル作成用ヘルパー関数
        def create_diagonal_cell(row, col, width, rowspan=1):
            """斜線付きの空セルを作成（左上から右下への対角線）"""
            # 文字幅をピクセルに変換（概算: 1文字 ≈ 8px）
            cell_width = width * 8
            cell_height = 15 * rowspan
            canvas = tk.Canvas(parent, width=cell_width, height=cell_height,
                             bg='white', highlightthickness=1, highlightbackground='gray')
            canvas.grid(row=row, column=col, rowspan=rowspan, sticky='nsew')
            # 左上から右下への斜線を描画
            canvas.create_line(0, 0, cell_width, cell_height, fill='gray')

        # データ行を作成
        row_idx = 1

        # 行高さを均一にするための最小高さ設定
        row_min_height = 15

        # POS/NEGそれぞれ処理
        for pole_idx, (sections, serial, pole) in enumerate([(pos_sections, pos_serial, "POS"),
                                                              (neg_sections, neg_serial, "NEG")]):
            if not sections:
                continue

            # コードごとにデータを整理（FFFFF, 80000, 00000）
            code_data = self._organize_sections_by_code(sections)

            unit_name = f"1PB397MK2\nDFH_{serial}\n{pole}"
            unit_row_start = row_idx

            # 23℃のフルスケール電圧範囲を計算
            fffff_23 = code_data.get('FFFFF', {}).get(23, {}).get('voltage')
            zero_23 = code_data.get('00000', {}).get(23, {}).get('voltage')
            full_scale_23 = (fffff_23 - zero_23) if (fffff_23 is not None and zero_23 is not None) else None

            for code_idx, code in enumerate(['FFFFF', '80000', '00000']):
                if code not in code_data:
                    continue

                temps_data = code_data[code]  # {28: voltage, 23: voltage, 18: voltage}
                code_row_start = row_idx

                # 温度データ（28, 23, 18の順）- 各温度2行で合計6行
                temp_order = [28, 23, 18]
                ref_voltage = temps_data.get(23, {}).get('voltage')  # 23℃が基準
                rows_per_temp = 2  # 各温度2行

                for temp_idx, temp in enumerate(temp_order):
                    if temp not in temps_data:
                        continue

                    voltage = temps_data[temp]['voltage']
                    temp_row = code_row_start + temp_idx * rows_per_temp

                    # 温度セル（2行結合）
                    tk.Label(parent, text=str(temp), relief=tk.RIDGE,
                            width=header_widths[2], font=('', 9), anchor='center', bg='white'
                            ).grid(row=temp_row, column=2, rowspan=rows_per_temp, sticky='nsew')

                    # 測定電圧セル（2行結合）
                    tk.Label(parent, text=f"{voltage:.6f}" if voltage else "",
                            relief=tk.RIDGE, width=header_widths[3],
                            font=('', 9), anchor='e', bg='white'
                            ).grid(row=temp_row, column=3, rowspan=rows_per_temp, sticky='nsew')

                row_idx = code_row_start + len(temp_order) * rows_per_temp

                # ΔT以降のセル - 28→23の計算（行1-2に跨る）
                if 28 in temps_data and 23 in temps_data:
                    v28 = temps_data[28]['voltage']
                    v23 = temps_data[23]['voltage']
                    delta_t_28 = 5.0
                    delta_v_28 = v28 - v23
                    temp_coef_28 = (delta_v_28 / full_scale_23) * 1e6 / delta_t_28 if full_scale_23 else 0
                    judgment_28 = "OK" if abs(temp_coef_28) <= spec_value else "NG"
                    bg_28 = '#FFCCCC' if judgment_28 == "NG" else 'white'

                    # 28℃の下半分と23℃の上半分に跨る（row 1-2）
                    delta_row = code_row_start + 1
                    tk.Label(parent, text=f"{delta_t_28:.1f}", relief=tk.RIDGE, width=header_widths[4],
                            font=('', 9), anchor='e', bg='white').grid(row=delta_row, column=4, rowspan=2, sticky='nsew')
                    tk.Label(parent, text=f"{delta_v_28:.6f}", relief=tk.RIDGE, width=header_widths[5],
                            font=('', 9), anchor='e', bg='white').grid(row=delta_row, column=5, rowspan=2, sticky='nsew')
                    tk.Label(parent, text=f"{temp_coef_28:.1f}", relief=tk.RIDGE, width=header_widths[6],
                            font=('', 9), anchor='e', bg=bg_28).grid(row=delta_row, column=6, rowspan=2, sticky='nsew')
                    tk.Label(parent, text=judgment_28, relief=tk.RIDGE, width=header_widths[8],
                            font=('', 9), anchor='center', bg=bg_28).grid(row=delta_row, column=8, rowspan=2, sticky='nsew')

                # ΔT以降のセル - 23→18の計算（行3-4に跨る）
                if 23 in temps_data and 18 in temps_data:
                    v23 = temps_data[23]['voltage']
                    v18 = temps_data[18]['voltage']
                    delta_t_18 = -5.0
                    delta_v_18 = v18 - v23
                    temp_coef_18 = (delta_v_18 / full_scale_23) * 1e6 / delta_t_18 if full_scale_23 else 0
                    judgment_18 = "OK" if abs(temp_coef_18) <= spec_value else "NG"
                    bg_18 = '#FFCCCC' if judgment_18 == "NG" else 'white'

                    # 23℃の下半分と18℃の上半分に跨る（row 3-4）
                    delta_row = code_row_start + 3
                    tk.Label(parent, text=f"{delta_t_18:.1f}", relief=tk.RIDGE, width=header_widths[4],
                            font=('', 9), anchor='e', bg='white').grid(row=delta_row, column=4, rowspan=2, sticky='nsew')
                    tk.Label(parent, text=f"{delta_v_18:.6f}", relief=tk.RIDGE, width=header_widths[5],
                            font=('', 9), anchor='e', bg='white').grid(row=delta_row, column=5, rowspan=2, sticky='nsew')
                    tk.Label(parent, text=f"{temp_coef_18:.1f}", relief=tk.RIDGE, width=header_widths[6],
                            font=('', 9), anchor='e', bg=bg_18).grid(row=delta_row, column=6, rowspan=2, sticky='nsew')
                    tk.Label(parent, text=judgment_18, relief=tk.RIDGE, width=header_widths[8],
                            font=('', 9), anchor='center', bg=bg_18).grid(row=delta_row, column=8, rowspan=2, sticky='nsew')

                # 空セル（ΔT列の上端と下端）- 斜線付き
                for col in [4, 5, 6, 8]:
                    create_diagonal_cell(code_row_start, col, header_widths[col])
                    create_diagonal_cell(code_row_start + 5, col, header_widths[col])

                # 入力コードのセル結合（縦6行）
                tk.Label(parent, text=f"{code} H", relief=tk.RIDGE, width=header_widths[1],
                        font=('', 9), bg='white').grid(row=code_row_start, column=1, rowspan=6, sticky='nsew')

                # 各行の高さを均一に設定
                for r in range(6):
                    parent.grid_rowconfigure(code_row_start + r, minsize=row_min_height)

            # 各コード6行 × 3コード = 18行
            total_rows = 6 * 3  # FFFFF, 80000, 00000
            row_idx = unit_row_start + total_rows

            # ユニットNo.のセル結合（縦）
            tk.Label(parent, text=unit_name, relief=tk.RIDGE, width=header_widths[0],
                    font=('', 9), bg='white').grid(row=unit_row_start, column=0, rowspan=total_rows, sticky='nsew')

        # スペックのセル結合（POS/NEG全体で共通）
        total_data_rows = row_idx - 1  # ヘッダー行を除いた全データ行数
        if total_data_rows > 0:
            tk.Label(parent, text=f"{spec_value:.1f}", relief=tk.RIDGE, width=header_widths[7],
                    font=('', 9), anchor='center', bg='white').grid(row=1, column=7, rowspan=total_data_rows, sticky='nsew')

    def _create_temp_coef_table_in_frame(self, parent, pos_sections, neg_sections, serial):
        """温度係数テーブルを指定フレームに作成（タブ用）"""
        # スペック値を取得
        try:
            spec_value = float(self.temp_coef_spec_var.get())
        except ValueError:
            spec_value = 1.9

        # ヘッダー行
        headers = ['ユニットNo.', '入力コード', '温度\n(℃)', '測定電圧\n(V)', 'ΔT\n(℃)', 'ΔV\n(V)',
                   '温度係数\n(ppm/℃)', 'スペック\n(ppm/℃)', '判定']
        header_widths = [10, 7, 4, 10, 4, 10, 9, 7, 4]

        for col, (header, width) in enumerate(zip(headers, header_widths)):
            label = tk.Label(parent, text=header, relief=tk.RIDGE, width=width,
                           bg='#D0D0D0', font=('', 9))
            label.grid(row=0, column=col, sticky='nsew', padx=0, pady=0)

        # 斜線セル作成用ヘルパー関数
        def create_diagonal_cell(row, col, width, rowspan=1):
            """斜線付きの空セルを作成（左上から右下への対角線）"""
            cell_width = width * 8
            cell_height = 15 * rowspan
            canvas = tk.Canvas(parent, width=cell_width, height=cell_height,
                             bg='white', highlightthickness=1, highlightbackground='gray')
            canvas.grid(row=row, column=col, rowspan=rowspan, sticky='nsew')
            canvas.create_line(0, 0, cell_width, cell_height, fill='gray')

        # データ行を作成
        row_idx = 1
        row_min_height = 15

        # POS/NEGそれぞれ処理
        for pole_idx, (sections, pole) in enumerate([(pos_sections, "POS"), (neg_sections, "NEG")]):
            if not sections:
                continue

            code_data = self._organize_sections_by_code(sections)
            unit_name = f"1PB397MK2\nDFH_{serial}\n{pole}"
            unit_row_start = row_idx

            # 23℃のフルスケール電圧範囲を計算
            fffff_23 = code_data.get('FFFFF', {}).get(23, {}).get('voltage')
            zero_23 = code_data.get('00000', {}).get(23, {}).get('voltage')
            full_scale_23 = (fffff_23 - zero_23) if (fffff_23 is not None and zero_23 is not None) else None

            for code_idx, code in enumerate(['FFFFF', '80000', '00000']):
                if code not in code_data:
                    continue

                temps_data = code_data[code]
                code_row_start = row_idx
                temp_order = [28, 23, 18]
                rows_per_temp = 2

                for temp_idx, temp in enumerate(temp_order):
                    if temp not in temps_data:
                        continue

                    voltage = temps_data[temp]['voltage']
                    temp_row = code_row_start + temp_idx * rows_per_temp

                    tk.Label(parent, text=str(temp), relief=tk.RIDGE,
                            width=header_widths[2], font=('', 9), anchor='center', bg='white'
                            ).grid(row=temp_row, column=2, rowspan=rows_per_temp, sticky='nsew')

                    tk.Label(parent, text=f"{voltage:.6f}" if voltage else "",
                            relief=tk.RIDGE, width=header_widths[3],
                            font=('', 9), anchor='e', bg='white'
                            ).grid(row=temp_row, column=3, rowspan=rows_per_temp, sticky='nsew')

                row_idx = code_row_start + len(temp_order) * rows_per_temp

                # ΔT以降のセル - 28→23の計算
                if 28 in temps_data and 23 in temps_data:
                    v28 = temps_data[28]['voltage']
                    v23 = temps_data[23]['voltage']
                    delta_t_28 = 5.0
                    delta_v_28 = v28 - v23
                    temp_coef_28 = (delta_v_28 / full_scale_23) * 1e6 / delta_t_28 if full_scale_23 else 0
                    judgment_28 = "OK" if abs(temp_coef_28) <= spec_value else "NG"
                    bg_28 = '#FFCCCC' if judgment_28 == "NG" else 'white'

                    delta_row = code_row_start + 1
                    tk.Label(parent, text=f"{delta_t_28:.1f}", relief=tk.RIDGE, width=header_widths[4],
                            font=('', 9), anchor='e', bg='white').grid(row=delta_row, column=4, rowspan=2, sticky='nsew')
                    tk.Label(parent, text=f"{delta_v_28:.6f}", relief=tk.RIDGE, width=header_widths[5],
                            font=('', 9), anchor='e', bg='white').grid(row=delta_row, column=5, rowspan=2, sticky='nsew')
                    tk.Label(parent, text=f"{temp_coef_28:.1f}", relief=tk.RIDGE, width=header_widths[6],
                            font=('', 9), anchor='e', bg=bg_28).grid(row=delta_row, column=6, rowspan=2, sticky='nsew')
                    tk.Label(parent, text=judgment_28, relief=tk.RIDGE, width=header_widths[8],
                            font=('', 9), anchor='center', bg=bg_28).grid(row=delta_row, column=8, rowspan=2, sticky='nsew')

                # ΔT以降のセル - 23→18の計算
                if 23 in temps_data and 18 in temps_data:
                    v23 = temps_data[23]['voltage']
                    v18 = temps_data[18]['voltage']
                    delta_t_18 = -5.0
                    delta_v_18 = v18 - v23
                    temp_coef_18 = (delta_v_18 / full_scale_23) * 1e6 / delta_t_18 if full_scale_23 else 0
                    judgment_18 = "OK" if abs(temp_coef_18) <= spec_value else "NG"
                    bg_18 = '#FFCCCC' if judgment_18 == "NG" else 'white'

                    delta_row = code_row_start + 3
                    tk.Label(parent, text=f"{delta_t_18:.1f}", relief=tk.RIDGE, width=header_widths[4],
                            font=('', 9), anchor='e', bg='white').grid(row=delta_row, column=4, rowspan=2, sticky='nsew')
                    tk.Label(parent, text=f"{delta_v_18:.6f}", relief=tk.RIDGE, width=header_widths[5],
                            font=('', 9), anchor='e', bg='white').grid(row=delta_row, column=5, rowspan=2, sticky='nsew')
                    tk.Label(parent, text=f"{temp_coef_18:.1f}", relief=tk.RIDGE, width=header_widths[6],
                            font=('', 9), anchor='e', bg=bg_18).grid(row=delta_row, column=6, rowspan=2, sticky='nsew')
                    tk.Label(parent, text=judgment_18, relief=tk.RIDGE, width=header_widths[8],
                            font=('', 9), anchor='center', bg=bg_18).grid(row=delta_row, column=8, rowspan=2, sticky='nsew')

                # 空セル（ΔT列の上端と下端）- 斜線付き
                for col in [4, 5, 6, 8]:
                    create_diagonal_cell(code_row_start, col, header_widths[col])
                    create_diagonal_cell(code_row_start + 5, col, header_widths[col])

                # 入力コードのセル結合（縦6行）
                tk.Label(parent, text=f"{code} H", relief=tk.RIDGE, width=header_widths[1],
                        font=('', 9), bg='white').grid(row=code_row_start, column=1, rowspan=6, sticky='nsew')

                # 各行の高さを均一に設定
                for r in range(6):
                    parent.grid_rowconfigure(code_row_start + r, minsize=row_min_height)

            # 各コード6行 × 3コード = 18行
            total_rows = 6 * 3
            row_idx = unit_row_start + total_rows

            # ユニットNo.のセル結合（縦）
            tk.Label(parent, text=unit_name, relief=tk.RIDGE, width=header_widths[0],
                    font=('', 9), bg='white').grid(row=unit_row_start, column=0, rowspan=total_rows, sticky='nsew')

        # スペックのセル結合（POS/NEG全体で共通）
        total_data_rows = row_idx - 1
        if total_data_rows > 0:
            tk.Label(parent, text=f"{spec_value:.1f}", relief=tk.RIDGE, width=header_widths[7],
                    font=('', 9), anchor='center', bg='white').grid(row=1, column=7, rowspan=total_data_rows, sticky='nsew')

    def _organize_sections_by_code(self, sections):
        """区間データをコードと温度で整理"""
        # 温特パターン: FFFFF→00000→80000 を繰り返し
        # 温度順: 23℃(1), 28℃, 18℃, 23℃(2), 23℃戻し
        temp_mapping = {
            0: 23,   # 23℃(1) - 基準値として使用
            1: 28,   # 28℃
            2: 18,   # 18℃
            3: 23,   # 23℃(2)
            4: 23,   # 23℃戻し - 参考値
        }

        code_data = {}  # {code: {temp: {'voltage': v, 'temp_set': idx}}}
        codes_per_temp = 3

        for i, section in enumerate(sections):
            temp_set_idx = i // codes_per_temp
            code = section['code']

            if temp_set_idx not in temp_mapping:
                continue

            temp = temp_mapping[temp_set_idx]

            if code not in code_data:
                code_data[code] = {}

            # 23℃は最初の値（temp_set_idx=0）を基準として使用
            # 28℃、18℃はそれぞれ1回のみ
            if temp == 23:
                # 23℃(1)のデータを基準として使用（temp_set_idx=0）
                if temp_set_idx == 0:
                    code_data[code][temp] = {
                        'voltage': section['avg_voltage'],
                        'temp_set': temp_set_idx
                    }
            else:
                code_data[code][temp] = {
                    'voltage': section['avg_voltage'],
                    'temp_set': temp_set_idx
                }

        return code_data

    def _format_minutes(self, minutes):
        """分を時:分形式にフォーマット"""
        hours = int(minutes // 60)
        mins = int(minutes % 60)
        if hours > 0:
            return f"{hours}h{mins:02d}m"
        else:
            return f"{mins}m"
