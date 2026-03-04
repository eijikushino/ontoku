"""テンプレートXLSX生成スクリプト

参照XLSXからPosition/LBC用の出荷シーケンステンプレートを生成する。
一度実行して template/ ディレクトリに保存した後、gitコミットする。

使い方:
    python create_templates.py
"""
import os
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule

# ── 定数 (LinearityTab と同じ) ──────────────────────────────
POSITION_POS = [
    0x00000, 0x00001, 0x00002, 0x00003, 0x00004, 0x00007, 0x00008,
    0x0000F, 0x00010, 0x0001F, 0x00020, 0x0003F, 0x00040, 0x0007F,
    0x00080, 0x000FF, 0x00100, 0x001FF, 0x00200, 0x003FF, 0x00400,
    0x007FF, 0x00800, 0x00FFF, 0x01000, 0x01FFF, 0x02000, 0x03FFF,
    0x04000, 0x07FFF, 0x08000, 0x0FFFF, 0x10000, 0x1FFFF, 0x20000,
    0x3FFFF, 0x40000, 0x5FFFF, 0x60000, 0x7FFFF, 0x80000, 0x9FFFF,
    0xA0000, 0xBFFFF, 0xC0000, 0xDFFFF, 0xE0000, 0xFFFFF,
]

LBC_POS = [
    0x0000, 0x0001, 0x0002, 0x0003, 0x0004, 0x0007, 0x0008,
    0x000F, 0x0010, 0x001F, 0x0020, 0x003F, 0x0040, 0x007F,
    0x0080, 0x00FF, 0x0100, 0x01FF, 0x0200, 0x03FF, 0x0400,
    0x07FF, 0x0800, 0x0FFF, 0x1000, 0x1FFF, 0x2000, 0x2FFF,
    0x3000, 0x3FFF, 0x4000, 0x4FFF, 0x5000, 0x5FFF, 0x6000,
    0x6FFF, 0x7000, 0x7FFF, 0x8000, 0x8FFF, 0x9000, 0x9FFF,
    0xA000, 0xAFFF, 0xB000, 0xBFFF, 0xC000, 0xCFFF, 0xD000,
    0xDFFF, 0xE000, 0xEFFF, 0xF000, 0xFFFF,
]

CRITERIA = {
    'Position': {'inl': 0.75, 'dnl': 0.50},
    'LBC':      {'inl': 0.50, 'dnl': 0.25},
}

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(SCRIPT_DIR, 'template')
REF_XLSX = os.path.join(
    SCRIPT_DIR, '..', 'Linearity結果',
    'DFH903_Position_POS_linearity_出荷シーケンス_20260304_102918_グラフ修正.xlsx')

JP_FONT = Font(name='ＭＳ Ｐゴシック', size=8, charset=128)
JP_FONT_BOLD = Font(name='ＭＳ Ｐゴシック', size=9, charset=128)


# ── ユーティリティ ──────────────────────────────────────────
def update_chart_refs(ws, old_name, new_name):
    """チャート内のシート参照を更新"""
    old_unquoted = f"{old_name}!"
    old_quoted = f"'{old_name}'!"
    new_ref = f"'{new_name}'!"

    for chart in ws._charts:
        for child in chart._charts:
            for series in child.series:
                if hasattr(series.val, 'numRef') and series.val.numRef and series.val.numRef.f:
                    f = series.val.numRef.f
                    f = f.replace(old_quoted, new_ref)
                    f = f.replace(old_unquoted, new_ref)
                    series.val.numRef.f = f
                if (hasattr(series, 'title') and hasattr(series.title, 'strRef')
                        and series.title.strRef and series.title.strRef.f):
                    f = series.title.strRef.f
                    f = f.replace(old_quoted, new_ref)
                    f = f.replace(old_unquoted, new_ref)
                    series.title.strRef.f = f


def add_conditional_formatting(ws, first_r, last_r, dac_name):
    """INL/DNL/判定結果の条件付き書式を追加"""
    red_font = Font(color="FF0000")
    crit = CRITERIA[dac_name]

    # INL (G列): |value| > threshold → 赤
    ws.conditional_formatting.add(
        f'G{first_r}:G{last_r}',
        CellIsRule(operator='greaterThan', formula=[str(crit['inl'])], font=red_font))
    ws.conditional_formatting.add(
        f'G{first_r}:G{last_r}',
        CellIsRule(operator='lessThan', formula=[str(-crit['inl'])], font=red_font))

    # DNL (H列): |value| > threshold → 赤
    ws.conditional_formatting.add(
        f'H{first_r}:H{last_r}',
        CellIsRule(operator='greaterThan', formula=[str(crit['dnl'])], font=red_font))
    ws.conditional_formatting.add(
        f'H{first_r}:H{last_r}',
        CellIsRule(operator='lessThan', formula=[str(-crit['dnl'])], font=red_font))

    # 判定結果セル: "NG" → 赤・サイズ26
    judge_r = last_r + 4
    ws.conditional_formatting.add(
        f'G{judge_r}',
        CellIsRule(operator='equal', formula=['"NG"'],
                   font=Font(size=26, color="FF0000")))


# ── Position テンプレート ────────────────────────────────────
def create_position_template():
    """参照XLSXからPosition用テンプレートを生成"""
    print(f"Loading reference: {REF_XLSX}")
    wb = openpyxl.load_workbook(REF_XLSX)
    ws = wb.active

    n_pts = len(POSITION_POS)  # 48
    first_r = 7
    last_r = first_r + n_pts - 1  # 54

    # シート名を "T" に変更
    old_name = ws.title
    ws.title = "T"

    # B,C列を非表示
    ws.column_dimensions['B'].hidden = True
    ws.column_dimensions['C'].hidden = True

    # E列のデータをクリア (値のみ、書式は保持)
    for r in range(first_r, last_r + 1):
        ws.cell(row=r, column=5).value = None

    # G,H列のフォントをリセット (手動設定の赤色フォントを除去)
    for r in range(first_r, last_r + 1):
        ws.cell(row=r, column=7).font = JP_FONT
        if ws.cell(row=r, column=8).value is not None:
            ws.cell(row=r, column=8).font = JP_FONT

    # 判定結果セルのフォントもリセット (デフォルトサイズ26、黒)
    judge_r = last_r + 4  # 58
    ws.cell(row=judge_r, column=7).font = Font(
        name='ＭＳ Ｐゴシック', size=26, charset=128)

    # チャート参照を更新
    update_chart_refs(ws, old_name, "T")

    # チャートタイトルのプレースホルダーをクリア
    chart = ws._charts[0]
    runs = chart.title.tx.rich.paragraphs[0].r
    runs[0].t = ""   # "{name} {pole} {dac_name} " - 実行時に設定

    # 条件付き書式を追加
    add_conditional_formatting(ws, first_r, last_r, 'Position')

    # 保存
    os.makedirs(TEMPLATE_DIR, exist_ok=True)
    out_path = os.path.join(TEMPLATE_DIR, 'linearity_position.xlsx')
    wb.save(out_path)
    print(f"Position template saved: {out_path}")
    return out_path


# ── LBC テンプレート ─────────────────────────────────────────
def create_lbc_template():
    """Position用テンプレートをベースにLBC用テンプレートを生成"""
    pos_path = os.path.join(TEMPLATE_DIR, 'linearity_position.xlsx')
    print(f"Loading Position template: {pos_path}")
    wb = openpyxl.load_workbook(pos_path)
    ws = wb.active

    lbc_n = len(LBC_POS)  # 54
    first_r = 7
    pos_last_r = 54   # Position: 7+48-1
    lbc_last_r = first_r + lbc_n - 1  # 60
    bits = 16
    hex_width = 4
    mask = 0xFFFF
    offset_val = 2 ** (bits - 1)

    center_align = Alignment(horizontal='center')
    right_align = Alignment(horizontal='right')
    thin = Side(style='thin')
    medium = Side(style='medium')
    fmt_v = '0.000000_ '
    fmt_err = '0.00_ '
    fmt_sci = '##0.0E+0'

    # ── Row 3: 理論直線係数 + Position 1LSB ──
    ws['F3'] = f'=(E{lbc_last_r}-E{first_r})/(C{lbc_last_r}-C{first_r})'
    ws['F3'].number_format = fmt_sci
    ws['G3'] = "Position　１LSB"
    ws['G3'].font = JP_FONT_BOLD
    ws['H3'] = '=IF(ABS(E7)>3,160/2^19,160/SQRT(2)/2^19)'
    ws['H3'].font = JP_FONT_BOLD
    ws['H3'].number_format = '##0.0E+0'

    # ── Headers (Row 5) ──
    ws['G5'] = 'INL'
    ws['G5'].font = JP_FONT_BOLD
    ws['G5'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H5'] = 'DNL'
    ws['H5'].font = JP_FONT_BOLD
    ws['H5'].alignment = Alignment(horizontal='center', vertical='center')

    # マージセルを先に解除 (G58:H58)
    try:
        ws.unmerge_cells(f'G{pos_last_r + 4}:H{pos_last_r + 4}')
    except Exception:
        pass

    # ── 既存データ行 + サマリー行をクリア (rows 7-58) ──
    for r in range(first_r, pos_last_r + 5):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.font = JP_FONT
            cell.border = Border()
            cell.number_format = 'General'
            cell.alignment = Alignment()
        # 行の高さをリセット (Position判定行の33.6を解消)
        ws.row_dimensions[r].height = None

    # ── 新データ行 (54点) ──
    for i, uval in enumerate(LBC_POS):
        r = first_r + i
        signed = uval - offset_val

        ws.cell(row=r, column=1, value=i + 1).font = JP_FONT
        ws.cell(row=r, column=1).alignment = center_align

        ws.cell(row=r, column=2, value=signed).font = JP_FONT
        ws.cell(row=r, column=2).alignment = center_align

        ws.cell(row=r, column=3, value=uval).font = JP_FONT
        ws.cell(row=r, column=3).alignment = center_align

        ws.cell(row=r, column=4, value=f"{uval & mask:0{hex_width}X}").font = JP_FONT
        ws.cell(row=r, column=4).alignment = center_align

        # E: 空 (実行時に書き込み)
        ws.cell(row=r, column=5).number_format = fmt_v
        ws.cell(row=r, column=5).font = JP_FONT
        ws.cell(row=r, column=5).alignment = right_align

        ws.cell(row=r, column=6, value=f'=$F$3*C{r}+$E${first_r}').font = JP_FONT
        ws.cell(row=r, column=6).number_format = fmt_v
        ws.cell(row=r, column=6).alignment = right_align

        # G: INL (LBCは $H$3 を使用)
        ws.cell(row=r, column=7, value=f'=(E{r}-F{r})/$H$3').font = JP_FONT
        ws.cell(row=r, column=7).number_format = fmt_err

        # H: DNL (コードステップ=1の場合のみ)
        if i < lbc_n - 1 and LBC_POS[i + 1] - uval == 1:
            ws.cell(row=r, column=8, value=f'=G{r + 1}-G{r}').font = JP_FONT
            ws.cell(row=r, column=8).number_format = fmt_err

        # 枠線
        is_last = (i == lbc_n - 1)
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = Border(
                top=thin,
                bottom=medium if is_last else thin,
                left=medium if c == 1 else thin,
                right=medium if c == 8 else thin)

    # ── サマリー行 ──
    sum_start = lbc_last_r + 1  # 61
    g_range = f'G{first_r}:G{lbc_last_r}'
    h_range = f'H{first_r}:H{lbc_last_r}'
    criteria = CRITERIA['LBC']

    summary_data = [
        ('+最大誤差', f'=MAX({g_range})', f'=MAX({h_range})'),
        ('-最大誤差', f'=MIN({g_range})', f'=MIN({h_range})'),
        ('判定基準', criteria['inl'], criteria['dnl']),
    ]
    for offset, (label, inl_v, dnl_v) in enumerate(summary_data):
        r = sum_start + offset
        ws.cell(row=r, column=6, value=label).font = JP_FONT
        ws.cell(row=r, column=6).alignment = center_align
        ws.cell(row=r, column=7, value=inl_v).font = JP_FONT
        ws.cell(row=r, column=7).number_format = fmt_err
        ws.cell(row=r, column=8, value=dnl_v).font = JP_FONT
        ws.cell(row=r, column=8).number_format = fmt_err
        if offset == 2:
            ws.cell(row=r, column=7).alignment = right_align
            ws.cell(row=r, column=8).alignment = right_align
        for c in range(6, 9):
            ws.cell(row=r, column=c).border = Border(
                top=medium if offset == 0 else thin,
                bottom=thin,
                left=medium if c == 6 else thin,
                right=medium if c == 8 else thin)

    # ── 合否判定行 ──
    r_judge = sum_start + 3  # 64
    r_plus = sum_start       # 61
    r_minus = sum_start + 1  # 62
    r_crit = sum_start + 2   # 63

    ws.cell(row=r_judge, column=6, value='判定結果')
    ws.cell(row=r_judge, column=6).font = Font(
        name='ＭＳ Ｐゴシック', size=11, charset=128)
    ws.cell(row=r_judge, column=6).alignment = Alignment(
        horizontal='center', vertical='center')
    ws.row_dimensions[r_judge].height = 33.6

    ws.merge_cells(f'G{r_judge}:H{r_judge}')
    judge_formula = (f'=IF((ABS(G{r_plus})<G{r_crit})'
                     f'+(ABS(H{r_plus})<H{r_crit})'
                     f'+(ABS(G{r_minus})<G{r_crit})'
                     f'+(ABS(H{r_minus})<H{r_crit})=4,"OK","NG")')
    ws.cell(row=r_judge, column=7, value=judge_formula)
    ws.cell(row=r_judge, column=7).font = Font(
        name='ＭＳ Ｐゴシック', size=26, charset=128)
    ws.cell(row=r_judge, column=7).alignment = center_align
    for c in range(6, 9):
        ws.cell(row=r_judge, column=c).border = Border(
            top=thin, bottom=medium,
            left=medium if c == 6 else thin,
            right=medium if c == 8 else thin)

    # ── チャートデータ範囲を更新 ──
    chart = ws._charts[0]
    chart._charts[0].series[0].val.numRef.f = f"'T'!$G${first_r}:$G${lbc_last_r}"
    chart._charts[1].series[0].val.numRef.f = f"'T'!$H${first_r}:$H${lbc_last_r}"

    # ── グラフ縦軸を±0.5 LSB、0.1刻みに設定（両軸） ──
    chart.y_axis.scaling.min = -0.5
    chart.y_axis.scaling.max = 0.5
    chart.y_axis.majorUnit = 0.1
    for child_chart in chart._charts:
        if hasattr(child_chart, 'y_axis'):
            child_chart.y_axis.scaling.min = -0.5
            child_chart.y_axis.scaling.max = 0.5
            child_chart.y_axis.majorUnit = 0.1

    # チャートタイトルの点数を更新
    runs = chart.title.tx.rich.paragraphs[0].r
    runs[2].t = f"({lbc_n}"

    # ── 条件付き書式 ──
    # Position用の条件付き書式をクリアしてLBC用に置換
    from collections import OrderedDict
    ws.conditional_formatting._cf_rules = OrderedDict()
    add_conditional_formatting(ws, first_r, lbc_last_r, 'LBC')

    # 保存
    out_path = os.path.join(TEMPLATE_DIR, 'linearity_lbc.xlsx')
    wb.save(out_path)
    print(f"LBC template saved: {out_path}")
    return out_path


if __name__ == '__main__':
    create_position_template()
    create_lbc_template()
    print("Done!")
